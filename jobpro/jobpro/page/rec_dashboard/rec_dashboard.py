import frappe
from frappe import _
from frappe.utils import getdate, flt, today, nowdate
from frappe.utils import now_datetime
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import json

import json
import frappe
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import frappe, io
import os
from frappe.utils.file_manager import get_file_path
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
import base64



# @frappe.whitelist()
# def get_ptsr_data_project_wise(status=None, sourcing_statu=None):
#     filters = {
#         "service": ["in", ['REC-I', 'REC-D']]
#     }

#     if status:
#         if isinstance(status, str):
#             try:
#                 status = json.loads(status)
#             except json.JSONDecodeError:
#                 status = [status]
#         elif not isinstance(status, list):
#             status = [status]
#         filters["status"] = ["in", status]

    # if sourcing_statu:
    #     if isinstance(sourcing_statu, str):
    #         try:
    #             sourcing_statu = json.loads(sourcing_statu)
    #         except json.JSONDecodeError:
    #             sourcing_statu = [sourcing_statu]
    #     elif not isinstance(sourcing_statu, list):
    #         sourcing_statu = [sourcing_statu]
    #     filters["sourcing_statu"] = ["in", sourcing_statu]


    # projects = frappe.get_all(
    #     "Project",
    #     filters=filters,
    #     fields=[
    #         'name', 'project_name', 'priority', 'remark','tvac','tsp','tfp','tsl','custom_t_lp','tpsl',
    #         'account_manager_remark', 'custom_spoc_remark',
    #         'sourcing_statu', 'territory', 'expected_value',
    #         'expected_psl', 'custom_psl_value', 'customer', 'creation'
    #     ],
    #     order_by="priority ASC, customer ASC"
    # )

    # return get_task_data(projects)


# @frappe.whitelist()
# def get_task_data(projects):
#     today = now_datetime()
#     data = []
#     total_counts = {
#         "vac": 0, "sp": 0, "fp": 0, "sl": 0, "psl": 0, "custom_lp": 0, "age": 0
#     }

#     for p in projects:
#         task_list = frappe.get_all("Task", {
#             "status": ("in", ['Working', 'Open', 'Overdue', 'Pending Review']),
#             "project": p.name
#         }, ['name','subject','priority','vac','sp','fp','sl','psl','custom_lp','creation','custom_task_sourcing_status'], order_by="priority ASC")

#         tasks = []
#         for t in task_list:
#             age = (today - t['creation']).days
#             vac = flt(t.get('vac'))
#             sp = flt(t.get('sp'))
#             fp = flt(t.get('fp'))
#             sl = flt(t.get('sl'))
#             psl = flt(t.get('psl'))
#             lp = flt(t.get('custom_lp'))

#             tasks.append({
#                 "name": t['name'],
#                 "task_name": t['subject'],
#                 "task_priority": t['priority'],
#                 "vac": vac,
#                 "sp": sp,
#                 "fp": fp,
#                 "sl": sl,
#                 "psl": psl,
#                 "custom_lp": lp,
#                 "age": age
#             })

#             # Accumulate totals
#             total_counts['vac'] += vac
#             total_counts['sp'] += sp
#             total_counts['fp'] += fp
#             total_counts['sl'] += sl
#             total_counts['psl'] += psl
#             total_counts['custom_lp'] += lp
#             total_counts['age'] += age

#         data.append({
#             "name": p['name'],
#             "project_name": p['project_name'],
#             "priority": p['priority'],
#             "remark": p['remark'],
#             "tvac":p['tvac'],
#             "tsp":p['tsp'],
#             "tfp":p['tfp'],
#             "tsl":p['tsl'],
#             "custom_t_lp":p['custom_t_lp'],
#             "tpsl":p['tpsl'],
#             "account_manager_remark": p['account_manager_remark'],
#             "custom_spoc_remark": p['custom_spoc_remark'],
#             "sourcing_statu": p['sourcing_statu'],
#             "territory": p['territory'],
#             "expected_value": p['expected_value'],
#             "expected_psl": p['expected_psl'],
#             "custom_psl_value": p['custom_psl_value'],
#             "customer": p['customer'],
#             "creation": p['creation'],
#             "task_count": len(tasks),
#             "tasks": tasks
#         })

#     return {
#         "projects": data,
#         "counts": total_counts
#     }


import frappe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from frappe.utils.file_manager import save_file


@frappe.whitelist()
def download_closure_summary_excel():

    # 🔹 Fetch data
    data = frappe.db.sql("""
        SELECT customer, status
        FROM `tabClosure`
    """, as_dict=1)

    # 🔹 Build summary
    summary = {}

    for d in data:
        customer = d.customer or "-"

        if customer not in summary:
            summary[customer] = {
                "Signed Offer Letter": 0,
                "Premedical": 0,
                "PCC / Visa": 0,
                "Final Medical": 0,
                "Biometric": 0,
                "QVP": 0,
                "Trade Test": 0,
                "Visa Stamping": 0,
                "Emigration": 0,
                "Ticket": 0,
                "Onboarding": 0
            }

        if d.status == "PCC" or d.status == "Visa":
            summary[customer]["PCC / Visa"] += 1
        elif d.status in summary[customer]:
            summary[customer][d.status] += 1

    # 🔹 Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Closure Summary"

    headers = [
        "Customer",
        "SOL",
        "Premedical",
        "PCC / Visa",
        "Final Medical",
        "Biometric",
        "QVP",
        "Trade Test",
        "Visa Stamping",
        "Emigration",
        "Ticket",
        "Onboarding"
    ]

    ws.append(headers)

    # 🔵 Header Style
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F1568")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 🔹 Data rows
    row_no = 2

    for customer, val in summary.items():
        ws.append([
            customer,
            val["Signed Offer Letter"],
            val["Premedical"],
            val["PCC / Visa"],
            val["Final Medical"],
            val["Biometric"],
            val["QVP"],
            val["Trade Test"],
            val["Visa Stamping"],
            val["Emigration"],
            val["Ticket"],
            val["Onboarding"]
        ])

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_no, column=col)

            if col == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

        row_no += 1

    # 🔹 Save file
    file_name = "closure_summary.xlsx"

    file_url = save_file(
        file_name,
        wb,
        "File",
        None,
        is_private=0
    ).file_url

    return file_url

@frappe.whitelist()
def get_ptsr_data_project_wise(status=None, sourcing_statu=None):
    """
    Fetch tasks first, then append project info if the project status matches.
    """
    today = now_datetime()
    
    # Task filters
    task_filters = {
        "status": ("in", ['Working', 'Open', 'Overdue', 'Pending Review']),
        "service":("in",['REC-I', 'REC-D'])
    }
    
    # Filter tasks by sourcing status if provided
    if sourcing_statu:
        if isinstance(sourcing_statu, str):
            try:
                sourcing_statu = json.loads(sourcing_statu)
            except json.JSONDecodeError:
                sourcing_statu = [sourcing_statu]
        elif not isinstance(sourcing_statu, list):
            sourcing_statu = [sourcing_statu]
        task_filters["custom_task_sourcing_status"] = ["in", sourcing_statu]
    
    # Fetch all tasks first
    tasks = frappe.get_all(
        "Task",
        filters=task_filters,
        fields=['name','subject','priority','vac','sp','fp','sl','psl','custom_lp',
                'project','custom_task_sourcing_status','creation'],
        order_by="priority ASC"
    )

    # Prepare totals and project mapping
    total_counts = {"vac": 0, "sp": 0, "fp": 0, "sl": 0, "psl": 0, "custom_lp": 0, "age": 0}
    projects_map = {}

    for t in tasks:
        # Get project info
        project_name = t['project']
        if not project_name:
            continue
        
        # Fetch project only once
        if project_name not in projects_map:
            p = frappe.get_doc("Project", project_name)
            
            # If project status filter is given, skip if doesn't match
            if status:
                if isinstance(status, str):
                    try:
                        status_list = json.loads(status)
                    except json.JSONDecodeError:
                        status_list = [status]
                elif not isinstance(status, list):
                    status_list = [status]
                else:
                    status_list = status
                if p.status not in status_list:
                    continue
            
            projects_map[project_name] = {
                "name": p.name,
                "project_name": p.project_name,
                "priority": p.priority,
                "remark": p.remark,
                "tvac": p.tvac,
                "tsp": p.tsp,
                "tfp": p.tfp,
                "tsl": p.tsl,
                "custom_t_lp": p.custom_t_lp,
                "tpsl": p.tpsl,
                "account_manager_remark": p.account_manager_remark,
                "custom_spoc_remark": p.custom_spoc_remark,
                "territory": p.territory,
                "expected_value": p.expected_value,
                "expected_psl": p.expected_psl,
                "custom_psl_value": p.custom_psl_value,
                "customer": p.customer,
                "creation": p.creation,
                "tasks": []
            }

        # Compute task age
        age = (today - t['creation']).days
        vac = flt(t.get('vac'))
        sp = flt(t.get('sp'))
        fp = flt(t.get('fp'))
        sl = flt(t.get('sl'))
        psl = flt(t.get('psl'))
        lp = flt(t.get('custom_lp'))

        # Append task to project
        projects_map[project_name]["tasks"].append({
            "name": t['name'],
            "task_name": t['subject'],
            "task_priority": t['priority'],
            "vac": vac,
            "sp": sp,
            "fp": fp,
            "sl": sl,
            "psl": psl,
            "custom_lp": lp,
            "custom_task_sourcing_status": t['custom_task_sourcing_status'],
            "age": age
        })

        # Update totals
        total_counts['vac'] += vac
        total_counts['sp'] += sp
        total_counts['fp'] += fp
        total_counts['sl'] += sl
        total_counts['psl'] += psl
        total_counts['custom_lp'] += lp
        total_counts['age'] += age

    # Convert project map to list
    projects_data = []
    for proj in projects_map.values():
        proj["task_count"] = len(proj["tasks"])
        proj["sourcing_statu"] = ", ".join([t['custom_task_sourcing_status'] for t in proj["tasks"] if t.get('custom_task_sourcing_status')])
        projects_data.append(proj)

    return {"projects": projects_data, "counts": total_counts}







@frappe.whitelist()
def download_all_closures_excel(status=None):
    """
    Generates one Excel file with 6 sheets:
    1. DIRECT CLOSURE (Indian)
    2. DIRECT FOLLOW UP CLOSURE (Indian)
    3. SUPPLIER FOLLOW UP CLOSURE (Indian)
    4. CLIENT CLOSURE (Indian)
    5. NEPAL CLOSURE (Nepali)
    6. SRILANKA CLOSURE (Srilankan)
    """

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "INTERNAL CLOSURE"
    ws2 = wb.create_sheet("CANDIDATE CLOSURE")
    ws3 = wb.create_sheet("AGENT CLOSURE")
    ws4 = wb.create_sheet("SUPPLIER CLOSURE")
    ws5 = wb.create_sheet("CLIENT CLOSURE")
    ws6 = wb.create_sheet("NEPAL CLOSURE")
    ws7 = wb.create_sheet("SRILANKA CLOSURE")

    
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")

    
    def get_closures_internal(nationality=None):
        filters = {}
        # Handle status
        if status:
            if isinstance(status, str):
                try:
                    s = json.loads(status)
                except json.JSONDecodeError:
                    s = [status]
            else:
                s = status if isinstance(status, list) else [status]
            filters["status"] = ["in", s]

        if nationality:
            filters["nationality"] = nationality
            
        filters["status"] = ["in", ["PSL", "Emigration","Ticket", "Onboarding"]]    

        closures = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name", "given_name", "passport_no", "customer", "territory",
                "status", "remark", "last_updated_on",
                "sa_name", "sa_mobile_number", "associate", "nationality"
            ],
            order_by="last_updated_on asc"
        )

        for c in closures:
            history = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )
            c["custom_history"] = history
            c["age"] = calculate_age_from_history(c.get("custom_history"))

        def safe_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        closures.sort(key=lambda x: safe_int(x.get("age")), reverse=True)
        return closures
    
    def get_closures_can(nationality=None):
        filters = {}
        # Handle status
        if status:
            if isinstance(status, str):
                try:
                    s = json.loads(status)
                except json.JSONDecodeError:
                    s = [status]
            else:
                s = status if isinstance(status, list) else [status]
            filters["status"] = ["in", s]

        if nationality:
            filters["nationality"] = nationality
            
        filters["sa_name"] = ["is", "not set"]
        filters["status"] = ["in", ["Signed Offer Letter","Premedical","PCC","Final Medical"]]     

        closures = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name", "given_name", "passport_no", "customer", "territory",
                "status", "remark", "last_updated_on",
                "sa_name", "sa_mobile_number", "associate", "nationality","mobile"
            ],
            order_by="last_updated_on asc"
        )

        for c in closures:
            history = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )
            c["custom_history"] = history
            c["age"] = calculate_age_from_history(c.get("custom_history"))

        def safe_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        closures.sort(key=lambda x: safe_int(x.get("age")), reverse=True)
        return closures
    
    def get_closures_agent(nationality=None):
        filters = {}
        # Handle status
        if status:
            if isinstance(status, str):
                try:
                    s = json.loads(status)
                except json.JSONDecodeError:
                    s = [status]
            else:
                s = status if isinstance(status, list) else [status]
            filters["status"] = ["in", s]

        if nationality:
            filters["nationality"] = nationality
            
        filters["sa_name"] = ["is", "set"]
        filters["status"] = ["in", ["Signed Offer Letter","Premedical","PCC","Final Medical"]]      

        closures = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name", "given_name", "passport_no", "customer", "territory",
                "status", "remark", "last_updated_on",
                "sa_name", "sa_mobile_number", "associate", "nationality","mobile"
            ],
            order_by="last_updated_on asc"
        )

        for c in closures:
            history = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )
            c["custom_history"] = history
            c["age"] = calculate_age_from_history(c.get("custom_history"))

        def safe_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        closures.sort(key=lambda x: safe_int(x.get("age")), reverse=True)
        return closures

    def get_closures_supp(nationality=None):
        filters = {}
        # Handle status
        if status:
            if isinstance(status, str):
                try:
                    s = json.loads(status)
                except json.JSONDecodeError:
                    s = [status]
            else:
                s = status if isinstance(status, list) else [status]
            filters["status"] = ["in", s]

        if nationality:
            filters["nationality"] = nationality
            
        
        filters["status"] = ["in", ["Certificate Attestation","Biometric","Trade Test","Visa Stamping"]]      

        closures = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name", "given_name", "passport_no", "customer", "territory",
                "status", "remark", "last_updated_on",
                "sa_name", "sa_mobile_number", "associate", "nationality","mobile"
            ],
            order_by="last_updated_on asc"
        )

        for c in closures:
            history = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )
            c["custom_history"] = history
            c["age"] = calculate_age_from_history(c.get("custom_history"))

        def safe_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        closures.sort(key=lambda x: safe_int(x.get("age")), reverse=True)
        return closures

    
    def get_closures_client(nationality=None):
        filters = {}
        # Handle status
        if status:
            if isinstance(status, str):
                try:
                    s = json.loads(status)
                except json.JSONDecodeError:
                    s = [status]
            else:
                s = status if isinstance(status, list) else [status]
            filters["status"] = ["in", s]

        if nationality:
            filters["nationality"] = nationality
            
        
        filters["status"] = ["in", ["Client Offer Letter","Visa"]]      

        closures = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name", "given_name", "passport_no", "customer", "territory",
                "status", "remark", "last_updated_on",
                "sa_name", "sa_mobile_number", "associate", "nationality","mobile"
            ],
            order_by="last_updated_on asc"
        )

        for c in closures:
            history = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )
            c["custom_history"] = history
            c["age"] = calculate_age_from_history(c.get("custom_history"))

        def safe_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        closures.sort(key=lambda x: safe_int(x.get("age")), reverse=True)
        return closures

    
    
    def get_closures_nepal(nationality=None):
        filters = {}
        # Handle status
        if status:
            if isinstance(status, str):
                try:
                    s = json.loads(status)
                except json.JSONDecodeError:
                    s = [status]
            else:
                s = status if isinstance(status, list) else [status]
            filters["status"] = ["in", s]

        if nationality:
            filters["nationality"] = nationality
            
        
        filters["status"] = ["in", ["PSL", "Emigration","Ticket", "Onboarding","Signed Offer Letter","Premedical","PCC","Final Medical","Certificate Attestation","Biometric","Trade Test","Visa Stamping"]]      

        closures = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name", "given_name", "passport_no", "customer", "territory",
                "status", "remark", "last_updated_on",
                "sa_name", "sa_mobile_number", "associate", "nationality","mobile"
            ],
            order_by="last_updated_on asc"
        )

        for c in closures:
            history = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )
            c["custom_history"] = history
            c["age"] = calculate_age_from_history(c.get("custom_history"))

        def safe_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        closures.sort(key=lambda x: safe_int(x.get("age")), reverse=True)
        return closures

    
    
    def get_closures_srilanka(nationality=None):
        filters = {}
        # Handle status
        if status:
            if isinstance(status, str):
                try:
                    s = json.loads(status)
                except json.JSONDecodeError:
                    s = [status]
            else:
                s = status if isinstance(status, list) else [status]
            filters["status"] = ["in", s]

        if nationality:
            filters["nationality"] = nationality
            
        
        filters["status"] = ["in", ["PSL", "Emigration","Ticket", "Onboarding","Signed Offer Letter","Premedical","PCC","Final Medical","Certificate Attestation","Biometric","Trade Test","Visa Stamping"]]      

        closures = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name", "given_name", "passport_no", "customer", "territory",
                "status", "remark", "last_updated_on",
                "sa_name", "sa_mobile_number", "associate", "nationality","mobile"
            ],
            order_by="last_updated_on asc"
        )

        for c in closures:
            history = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )
            c["custom_history"] = history
            c["age"] = calculate_age_from_history(c.get("custom_history"))

        def safe_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        closures.sort(key=lambda x: safe_int(x.get("age")), reverse=True)
        return closures

    
    
    
    
    # def write_sheet(ws, title, headers, rows, merge_range, col_widths):
    #     ws.merge_cells(merge_range)
    #     ws["A1"] = title
    #     ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    #     ws["A1"].alignment = center
    #     ws["A1"].fill = header_fill

    #     for idx, h in enumerate(headers, start=1):
    #         cell = ws.cell(row=3, column=idx, value=h)
    #         cell.font = Font(bold=True, size=12, color="FFFFFF")
    #         cell.alignment = center
    #         cell.fill = header_fill

    #     for i, w in enumerate(col_widths, start=1):
    #         ws.column_dimensions[chr(64 + i)].width = w

    #     for row_data in rows:
    #         ws.append(row_data)

    #     # Apply borders, alignment, wrap text
    #     center_cols = ["A", "B", "D", "G", "H"]  
    #     contact_col_index = headers.index("Contact") + 1 if "Contact" in headers else None
    #     latest_remark_col_index = headers.index("Latest Remark") + 1 if "Latest Remark" in headers else None
    #     remain_cols = ["C","E","F","J"]

    #     for r in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    #         for c_idx, cell in enumerate(r, start=1):
    #             cell.border = border
    #             col_letter = get_column_letter(c_idx)

    #             if col_letter in center_cols or (contact_col_index and c_idx == contact_col_index):
    #                 cell.alignment = Alignment(horizontal="center", vertical="center")
    #             if latest_remark_col_index and c_idx == latest_remark_col_index:
    #                 cell.alignment = Alignment(wrap_text=True, vertical="center")
                
    #             if col_letter in remain_cols:
    #                 cell.alignment = Alignment(wrap_text=True, vertical="center")
                    
    #     for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
    #                             min_col=1, max_col=len(headers)):
    #         for cell in row:
    #             ws.row_dimensions[cell.row].height = 30                
    
    

    def write_sheet(ws, title, headers, rows, merge_range, col_widths):
        
        even_row_fill = PatternFill(start_color="E6F2F1", end_color="E6F2F1", fill_type="solid")
        
        ws.merge_cells(merge_range)
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")

        
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=h)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")

        
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        
        for row_idx, row_data in enumerate(rows, start=4): 
            if row_data:  
                
                ws.append(row_data)

                
                row = ws[row_idx]  
                if row_idx % 2 != 0:  
                    for cell in row:
                        cell.fill = even_row_fill

                
                for c_idx, cell in enumerate(row, start=1):
                    cell.border = Border(left=Side(border_style="thin", color="000000"),
                                        right=Side(border_style="thin", color="000000"),
                                        top=Side(border_style="thin", color="000000"),
                                        bottom=Side(border_style="thin", color="000000"))
                    col_letter = get_column_letter(c_idx)
                    if col_letter in ["A", "B", "D", "G", "H"]:  
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_letter in ["E"]:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                    elif col_letter in ["C","F"]:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif headers[c_idx - 1] == "Latest Remark":
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                    elif headers[c_idx - 1] == "Contact":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        

        
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30
                        


    # -------------------- Headers & Column Widths --------------------
    headers_basic = ["S#", "CLID", "Name", "PP#", "Client", "Status", "Age", "Last Update On", "Latest Remark"]
    headers_followup = headers_basic + ["Contact"]
    headers_agent_followup = headers_basic + ["Agent","Contact"]
    headers_supplier = headers_basic + ["Supplier", "Contact"]

    col_widths_basic = [10, 12, 45, 20, 40, 18, 10, 15, 60]
    col_widths_followup = [10, 12, 45, 20, 40, 18, 10, 15, 60,  16]
    col_widths_agent_followup = [10, 12, 45, 20, 40, 18, 10, 15, 60, 20,  12]
    col_widths_supplier = [10, 12, 45, 20, 40, 18, 10, 15, 60, 20, 12]

    # -------------------- 1. DIRECT CLOSURE --------------------
    closures_direct = get_closures_internal("Indian")
    rows_direct = [
        [idx + 1, c.get("name"), c.get("given_name"), c.get("passport_no"), c.get("customer"),
         c.get("status"), c.get("age"), frappe.format(c.get("last_updated_on"), {"fieldtype": "Date"}), c.get("remark")]
        for idx, c in enumerate(closures_direct)
    ]
    write_sheet(ws1, "INTERNAL CLOSURE", headers_basic, rows_direct, "A1:I2", col_widths_basic)

    # -------------------- 2. DIRECT FOLLOW UP CLOSURE --------------------
    closures_followup = get_closures_can("Indian")
    rows_followup = [
        [idx + 1, c.get("name"), c.get("given_name"), c.get("passport_no"), c.get("customer"),
         c.get("status"), c.get("age"), frappe.format(c.get("last_updated_on"), {"fieldtype": "Date"}),
         c.get("remark"), c.get("mobile")]
        for idx, c in enumerate(closures_followup)
    ]
    write_sheet(ws2, "CANDIDATE CLOSURE", headers_followup, rows_followup, "A1:J2", col_widths_followup)
    
    # -------------------- 3. AGENT --------------------
    closures_followup = get_closures_agent("Indian")
    rows_followup = [
        [idx + 1, c.get("name"), c.get("given_name"), c.get("passport_no"), c.get("customer"),
         c.get("status"), c.get("age"), frappe.format(c.get("last_updated_on"), {"fieldtype": "Date"}),
         c.get("remark"), c.get("sa_name"), c.get("sa_mobile_number")]
        for idx, c in enumerate(closures_followup)
    ]
    write_sheet(ws3, "AGENT CLOSURE", headers_agent_followup, rows_followup, "A1:K2", col_widths_agent_followup)

    # -------------------- 4. SUPPLIER FOLLOW UP CLOSURE --------------------
    closures_supplier = get_closures_supp("Indian")
    rows_supplier = []
    for idx, c in enumerate(closures_supplier, start=1):
        mob = frappe.db.get_value("Supplier", {"name": c.get("associate", "")}, "mobile_no")
        row_data = [
            idx,
            c.get("name", ""),
            c.get("given_name", ""),
            c.get("passport_no", ""),
            c.get("customer", ""),
            c.get("status", ""),
            c.get("age", ""),
            frappe.format(c.get("last_updated_on"), {"fieldtype": "Date"}),
            c.get("remark", ""),
            c.get("associate", ""),  # Supplier
            mob,                     # Contact
        ]
        rows_supplier.append(row_data)
    write_sheet(ws4, "SUPPLIER CLOSURE", headers_supplier, rows_supplier, "A1:K2", col_widths_supplier)

    # -------------------- 5. CLIENT CLOSURE --------------------
    closures_client = get_closures_client("Indian")
    rows_client = [
        [idx + 1, c.get("name"), c.get("given_name"), c.get("passport_no"), c.get("customer"),
         c.get("status"), c.get("age"), frappe.format(c.get("last_updated_on"), {"fieldtype": "Date"}), c.get("remark")]
        for idx, c in enumerate(closures_client)
    ]
    write_sheet(ws5, "CLIENT CLOSURE", headers_basic, rows_client, "A1:I2", col_widths_basic)

    # -------------------- 6. NEPAL CLOSURE --------------------
    closures_nepal = get_closures_nepal("Nepali")
    rows_nepal = [
        [idx + 1, c.get("name"), c.get("given_name"), c.get("passport_no"), c.get("customer"),
         c.get("status"), c.get("age"), frappe.format(c.get("last_updated_on"), {"fieldtype": "Date"}), c.get("remark")]
        for idx, c in enumerate(closures_nepal)
    ]
    write_sheet(ws6, "NEPAL CLOSURE", headers_basic, rows_nepal, "A1:I2", col_widths_basic)

    # -------------------- 7. SRILANKA CLOSURE --------------------
    closures_srilanka = get_closures_srilanka("Srilankan")
    rows_srilanka = [
        [idx + 1, c.get("name"), c.get("given_name"), c.get("passport_no"), c.get("customer"),
         c.get("status"), c.get("age"), frappe.format(c.get("last_updated_on"), {"fieldtype": "Date"}), c.get("remark")]
        for idx, c in enumerate(closures_srilanka)
    ]
    write_sheet(ws7, "SRILANKA CLOSURE", headers_basic, rows_srilanka, "A1:I2", col_widths_basic)

    # -------------------- Save & Return --------------------
    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    frappe.response["filename"] = "All_Closures_Report.xlsx"
    frappe.response["filecontent"] = file_data.getvalue()
    frappe.response["type"] = "binary"







@frappe.whitelist()
def get_ptsr_data_closure_wise(status=None,territory=None,client=None):
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    if territory:
        filters["territory"] = territory
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    
    filters["nationality"] = "Indian"
            
        
        
    
            
      

    if client:
        filters["customer"] = client

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","mobile","std_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on desc"
        
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        

        

        
        c["custom_history"] = history

    frappe.log_error(message=closures, title="Closure with History")

    return { "closure": closures }

@frappe.whitelist()
def get_ptsr_data_closure_wise_candidate(status=None,territory=None,client=None):
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    if territory:
        filters["territory"] = territory
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    
    filters["nationality"] = "Indian"
    filters["sa_name"] = ["is", "not set"]

    if client:
        filters["customer"] = client
    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","mobile","std_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on desc"
        
    )
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        c["custom_history"] = history

    frappe.log_error(message=closures, title="Closure with History")

    return { "closure": closures }

@frappe.whitelist()
def get_ptsr_data_closure_wise_agent(status=None,territory=None,client=None):
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    if territory:
        filters["territory"] = territory
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    
    filters["nationality"] = "Indian"
    filters["sa_name"] = ["is", "set"]
   
            
        
        
    
            
      

    if client:
        filters["customer"] = client

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","mobile","std_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on desc"
        
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        

        

        
        c["custom_history"] = history

    frappe.log_error(message=closures, title="Closure with History")

    return { "closure": closures }

@frappe.whitelist()
def get_ptsr_data_closure_wise_nepal(status=None,territory=None,client=None):
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    if territory:
        filters["territory"] = territory
    # else:
    #     filters["territory"] = "Nepal"
    
    filters["nationality"] = "Nepali"
    
            
    
            
        
        
    
            
      

    if client:
        filters["customer"] = client

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","mobile","std_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on desc"
        
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        

        

        
        c["custom_history"] = history

    frappe.log_error(message=closures, title="Closure with History")

    return { "closure": closures }

@frappe.whitelist()
def get_ptsr_data_closure_wise_srilanka(status=None,territory=None,client=None):
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    if territory:
        filters["territory"] = territory
    # else:
    #     filters["territory"] = "Sri Lanka"
            
    filters["nationality"] = "Srilankan"
            
        
        
    
            
      

    if client:
        filters["customer"] = client

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","mobile","std_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
        
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        

        

        
        c["custom_history"] = history

    frappe.log_error(message=closures, title="Closure with History")

    return { "closure": closures }



# @frappe.whitelist()
# def get_ptsr_data_closure_wise_all():

#     # =====================================================
#     # STATUS GROUPS
#     # =====================================================

#     groups = {
#         "internal": {
#             "status": ["PSL", "Emigration", "Ticket", "Onboarding"],
#         },

#         "candidate": {
#             "status": ["Signed Offer Letter", "Premedical", "PCC", "Final Medical"],
#             "sa_name": ["is", "not set"]
#         },

#         "agent": {
#             "status": ["Signed Offer Letter", "Premedical", "PCC", "Final Medical"],
#             "sa_name": ["is", "set"]
#         },

#         "supplier": {
#             "status": ["Certificate Attestation", "Biometric", "Trade Test", "Visa Stamping"],
#             "sa_name": None
#         },

#         "client": {
#             "status": ["Client Offer Letter", "Visa"],
#             "sa_name": None
#         }
#     }

#     result = {}
#     project_cache = {}

#     def process_rows(rows):

#         for c in rows:

#             # history
#             history = frappe.get_all(
#                 "Closure Status History",
#                 filters={
#                     "parent": c["name"],
#                     "parenttype": "Closure",
#                     "parentfield": "custom_history"
#                 },
#                 fields=["date"]
#             )

#             c["custom_history"] = history

#             # project mapping
#             project = c.get("project")

#             if project and project not in project_cache:

#                 project_cache[project] = frappe.db.get_value(
#                     "Project",
#                     project,
#                     [
#                         "name",
#                         "project_name",
#                         "tvac",
#                         "tsp",
#                         "tfp",
#                         "tsl",
#                         "custom_t_lp",
#                         "custom_spoc_remark"
#                     ],
#                     as_dict=True
#                 )

#             p = project_cache.get(project) or {}

#             c["project_id"] = p.get("name") or project
#             c["project_name"] = p.get("project_name") or project

#             c["tvac"] = p.get("tvac") or 0
#             c["tsp"] = p.get("tsp") or 0
#             c["tfp"] = p.get("tfp") or 0
#             c["tsl"] = p.get("tsl") or 0
#             c["custom_t_lp"] = p.get("custom_t_lp") or 0
#             c["custom_spoc_remark"] = p.get("custom_spoc_remark") or "-"

#             # task
#             c["task_subject"] = frappe.db.get_value(
#                 "Task",
#                 c.get("task"),
#                 "subject"
#             ) or "No Position"

#         return rows

#     # =====================================================
#     # LOOP ALL GROUPS
#     # =====================================================

#     for key, cfg in groups.items():

#         filters = {
#             "status": ["in", cfg["status"]],
#             "nationality": "Indian"
#         }

#         if cfg.get("sa_name") is not None:
#             filters["sa_name"] = cfg["sa_name"]

#         data = frappe.db.get_all(
#             "Closure",
#             filters=filters,
#             fields=[
#                 "name","given_name","passport_no","customer",
#                 "territory","status","remark","last_updated_on",
#                 "sa_name","sa_mobile_number","associate","mobile",
#                 "std_remarks","custom_next_follow_up_on",
#                 "task","project"
#             ],
#             order_by="last_updated_on desc"
#         )

#         result[key] = process_rows(data)

#     return result



@frappe.whitelist()
def get_ptsr_data_closure_wise_all():

    groups = {
        "internal": {
            "status": ["PSL", "Emigration", "Ticket", "Onboarding"],
            "nationality": "Indian",
            "sa_name": None
        },

        "candidate": {
            "status": ["Signed Offer Letter", "Premedical", "PCC", "Final Medical"],
            "nationality": "Indian",
            "sa_name": ["is", "not set"]
        },

        "agent": {
            "status": ["Signed Offer Letter", "Premedical", "PCC", "Final Medical"],
            "nationality": "Indian",
            "sa_name": ["is", "set"]
        },

        "supplier": {
            "status": ["Certificate Attestation", "Biometric", "Trade Test", "Visa Stamping"],
            "nationality": "Indian",
            "sa_name": None
        },

        "client": {
            "status": ["Client Offer Letter", "Visa"],
            "nationality": "Indian",
            "sa_name": None
        },

        "nepal": {
            "status": [
                "PSL", "Emigration", "Ticket", "Onboarding",
                "Signed Offer Letter", "Premedical", "PCC", "Final Medical",
                "Certificate Attestation", "Biometric", "Trade Test", "Visa Stamping"
            ],
            "nationality": "Nepali",
            "sa_name": None
        },

        "srilanka": {
            "status": [
                "PSL", "Emigration", "Ticket", "Onboarding",
                "Signed Offer Letter", "Premedical", "PCC", "Final Medical",
                "Certificate Attestation", "Biometric", "Trade Test", "Visa Stamping"
            ],
            "nationality": "Srilankan",
            "sa_name": None
        }
    }

    result = {}
    project_cache = {}

    def process_rows(rows):

        for c in rows:

            c["custom_history"] = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )

            project = c.get("project")

            if project and project not in project_cache:

                project_cache[project] = frappe.db.get_value(
                    "Project",
                    project,
                    [
                        "name",
                        "project_name",
                        "tvac",
                        "tsp",
                        "tfp",
                        "tsl",
                        "custom_t_lp",
                        "custom_spoc_remark"
                    ],
                    as_dict=True
                )

            p = project_cache.get(project) or {}

            c["project_id"] = p.get("name") or project
            c["project_name"] = p.get("project_name") or project

            c["tvac"] = p.get("tvac") or 0
            c["tsp"] = p.get("tsp") or 0
            c["tfp"] = p.get("tfp") or 0
            c["tsl"] = p.get("tsl") or 0
            c["custom_t_lp"] = p.get("custom_t_lp") or 0
            c["custom_spoc_remark"] = p.get("custom_spoc_remark") or "-"

            c["task_subject"] = frappe.db.get_value(
                "Task",
                c.get("task"),
                "subject"
            ) or "No Position"

        return rows

    for key, cfg in groups.items():

        filters = {
            "status": ["in", cfg["status"]],
            "nationality": cfg["nationality"]
        }

        if cfg.get("sa_name"):
            filters["sa_name"] = cfg["sa_name"]

        data = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name","given_name","passport_no","customer",
                "territory","status","remark","last_updated_on",
                "sa_name","sa_mobile_number","associate","mobile",
                "std_remarks","custom_next_follow_up_on",
                "task","project"
            ],
            order_by="last_updated_on desc"
        )

        result[key] = process_rows(data)

    return result



@frappe.whitelist()
def get_territories():
    return frappe.get_all("Territory", fields=["name"])

@frappe.whitelist()
def get_client():
    return frappe.get_all("Customer", fields=["name"])

# @frappe.whitelist()
# def get_filtered_data(territory=None, client=None, table_type=None):
#     filters = {}
#     if territory:
#         filters["territory"] = territory
#     if client:
#         filters["Customer"] = client

#     if table_type == "teampro-btn":
#         data = frappe.get_all("Teampro", filters=filters, fields=["name", "status"])
#     elif table_type == "candidate-btn":
#         data = frappe.get_all("Candidate", filters=filters, fields=["name", "status"])
#     else:
#         data = frappe.get_all("Client", filters=filters, fields=["name", "status"])

#     return data



@frappe.whitelist()
def download_teampro_excel(status,territory=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "INTERNAL CLOSURE"
    
    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center",wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    even_row_fill = PatternFill(start_color="e6f2f1", end_color="e6f2f1", fill_type="solid")

    
    
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    # if territory:    
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    
    filters["nationality"] = "Indian"        

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","standard_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        
        c["custom_history"] = history
        
    for c in closures:
        c["age"] = calculate_age_from_history(c.get("custom_history"))
    
    # Sort closures by age (ascending)
    closures.sort(key=lambda x: x["age"], reverse=True)  
        
        
    
    
        
    
    ws.merge_cells("A1:I2")    
    ws["A1"]="INTERNAL CLOSURE"
    ws["A1"].font = Font(bold=True, size=14,color="FFFFFF")
    ws["A1"].alignment = center    
    ws["A1"].fill = header_fill   
        
        
    ws["A3"]="S#" 
    ws["A3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["A3"].alignment = center
    ws["A3"].fill = header_fill
       
    ws["B3"]="CLID"
    ws["B3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["B3"].alignment = center
    ws["B3"].fill = header_fill
         
    ws["C3"]="Name"
    ws["C3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["C3"].alignment = center
    ws["C3"].fill = header_fill
         
    ws["D3"]="PP#"
    ws["D3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["D3"].alignment = center
    ws["D3"].fill = header_fill
         
    ws["E3"]="Client"
    ws["E3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["E3"].alignment = center
    ws["E3"].fill = header_fill
    
         
    ws["F3"]="Status"
    ws["F3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["F3"].alignment = center
    ws["F3"].fill = header_fill
         
    ws["G3"]="Age"
    ws["G3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["G3"].alignment = center
    ws["G3"].fill = header_fill
         
    ws["H3"]="Last Update On"
    ws["H3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["H3"].alignment = center
    ws["H3"].fill = header_fill
        
    ws["I3"]="Latest Remark"
    ws["I3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["I3"].alignment = center
    ws["I3"].fill = header_fill    
       
    
    column_widths = {
        "A": 10, "B": 12, "C": 35, "D": 20, "E": 40, "F": 18, "G": 10, "H": 15,"I":60,
        
    }
 
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    
    
    
    for idx, closure in enumerate(closures, start=1):
        age = calculate_age_from_history(closure.get("custom_history"))
        last_update = frappe.format(closure.get("last_updated_on"),{"fieldtype":"Date"})

        row_data = [
            idx,
            closure.get("name", ""),
            closure.get("given_name", ""),
            closure.get("passport_no", ""),
            closure.get("customer", ""),
            closure.get("status", ""),
            closure.get("age", ""),
            last_update,
            closure.get("remark", "")
        ]

        ws.append(row_data)
        row = ws.max_row
        
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if idx % 2 == 0: 
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws[f"{col}{row}"].fill = even_row_fill
        
    for r in ws.iter_rows(min_row=1, max_row=row,
                                min_col=1, max_col=9):
            for cell in r:
                cell.border = border 
                
    for row in ws.iter_rows(min_row=4, max_row=row,
                                min_col=1, max_col=9):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30               
    

    
    
    
    
    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    encoded = base64.b64encode(file_data.getvalue()).decode()
    filename = "Teampro_Closure.xlsx"
    return {"filename": filename, "data": encoded}



@frappe.whitelist()
def download_candidate_excel(status,territory=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CANDIDATE CLOSURE"
    
    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    even_row_fill = PatternFill(start_color="e6f2f1", end_color="e6f2f1", fill_type="solid")

    
    
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    # if territory:    
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    filters["nationality"] = "Indian"
    filters["sa_name"] = ["is", "not set"]         

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","mobile","standard_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        
        c["custom_history"] = history
        
    for c in closures:
        c["age"] = calculate_age_from_history(c.get("custom_history"))
    
    # Sort closures by age (ascending)
    closures.sort(key=lambda x: x["age"], reverse=True)    
    
    
        
    
    ws.merge_cells("A1:J2")    
    ws["A1"]="CANDIDATE CLOSURE"
    ws["A1"].font = Font(bold=True, size=14,color="FFFFFF")
    ws["A1"].alignment = center    
    ws["A1"].fill = header_fill   
        
        
    ws["A3"]="S#" 
    ws["A3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["A3"].alignment = center
    ws["A3"].fill = header_fill
       
    ws["B3"]="CLID"
    ws["B3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["B3"].alignment = center
    ws["B3"].fill = header_fill
         
    ws["C3"]="Name"
    ws["C3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["C3"].alignment = center
    ws["C3"].fill = header_fill
         
    ws["D3"]="PP#"
    ws["D3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["D3"].alignment = center
    ws["D3"].fill = header_fill
         
    ws["E3"]="Client"
    ws["E3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["E3"].alignment = center
    ws["E3"].fill = header_fill
    
         
    ws["F3"]="Status"
    ws["F3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["F3"].alignment = center
    ws["F3"].fill = header_fill
         
    ws["G3"]="Age"
    ws["G3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["G3"].alignment = center
    ws["G3"].fill = header_fill
         
    ws["H3"]="Last Update On"
    ws["H3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["H3"].alignment = center
    ws["H3"].fill = header_fill
        
    ws["I3"]="Latest Remark"
    ws["I3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["I3"].alignment = center
    ws["I3"].fill = header_fill
    
        
    ws["J3"]="Contact"
    ws["J3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["J3"].alignment = center
    ws["J3"].fill = header_fill 
       
      
    
    column_widths = {
        "A": 10, "B": 12, "C": 35, "D": 20, "E": 40, "F": 18, "G": 10, "H": 20,"I":60,"J":11
        
    }
 
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    
    
    
    for idx, closure in enumerate(closures, start=1):
        age = calculate_age_from_history(closure.get("custom_history"))
        last_update = frappe.format(closure.get("last_updated_on"),{"fieldtype":"Date"})

        row_data = [
            idx,
            closure.get("name", ""),
            closure.get("given_name", ""),
            closure.get("passport_no", ""),
            closure.get("customer", ""),
            closure.get("status", ""),
            closure.get("age", ""),
            last_update,
            closure.get("remark", ""),
            closure.get("mobile", ""),
        ]

        ws.append(row_data)
        row = ws.max_row
        
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        if idx % 2 == 0: 
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','J']:
                ws[f"{col}{row}"].fill = even_row_fill
        
    for r in ws.iter_rows(min_row=1, max_row=row,
                                min_col=1, max_col=10):
            for cell in r:
                cell.border = border  
                
                  
    for row in ws.iter_rows(min_row=4, max_row=row,
                                min_col=1, max_col=10):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30 

    
    
    
    
    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    encoded = base64.b64encode(file_data.getvalue()).decode()
    filename = "Direct_follow_up_Closure.xlsx"
    return {"filename": filename, "data": encoded}


@frappe.whitelist()
def download_agent_excel(status,territory=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "AGENT CLOSURE"
    
    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    even_row_fill = PatternFill(start_color="e6f2f1", end_color="e6f2f1", fill_type="solid")

    
    
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
        
    # if territory:    
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    filters["nationality"] = "Indian"
    filters["sa_name"] = ["is", "set"]          

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","mobile","standard_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        
        c["custom_history"] = history
        
    for c in closures:
        c["age"] = calculate_age_from_history(c.get("custom_history"))
    
    # Sort closures by age (ascending)
    closures.sort(key=lambda x: x["age"], reverse=True)    
    
    
        
    
    ws.merge_cells("A1:K2")    
    ws["A1"]="AGENT CLOSURE"
    ws["A1"].font = Font(bold=True, size=14,color="FFFFFF")
    ws["A1"].alignment = center    
    ws["A1"].fill = header_fill   
        
        
    ws["A3"]="S#" 
    ws["A3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["A3"].alignment = center
    ws["A3"].fill = header_fill
       
    ws["B3"]="CLID"
    ws["B3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["B3"].alignment = center
    ws["B3"].fill = header_fill
         
    ws["C3"]="Name"
    ws["C3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["C3"].alignment = center
    ws["C3"].fill = header_fill
         
    ws["D3"]="PP#"
    ws["D3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["D3"].alignment = center
    ws["D3"].fill = header_fill
         
    ws["E3"]="Client"
    ws["E3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["E3"].alignment = center
    ws["E3"].fill = header_fill
    
         
    ws["F3"]="Status"
    ws["F3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["F3"].alignment = center
    ws["F3"].fill = header_fill
         
    ws["G3"]="Age"
    ws["G3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["G3"].alignment = center
    ws["G3"].fill = header_fill
         
    ws["H3"]="Last Update On"
    ws["H3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["H3"].alignment = center
    ws["H3"].fill = header_fill
        
    ws["I3"]="Latest Remark"
    ws["I3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["I3"].alignment = center
    ws["I3"].fill = header_fill
    
        
    ws["J3"]="Agent"
    ws["J3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["J3"].alignment = center
    ws["J3"].fill = header_fill 
       
    ws["K3"]="Contact"
    ws["K3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["K3"].alignment = center
    ws["K3"].fill = header_fill    
    
    column_widths = {
        "A": 10, "B": 12, "C": 35, "D": 20, "E": 40, "F": 18, "G": 10, "H": 20,"I":60,"J":20,"K":15
        
    }
 
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    
    
    
    for idx, closure in enumerate(closures, start=1):
        age = calculate_age_from_history(closure.get("custom_history"))
        last_update = frappe.format(closure.get("last_updated_on"),{"fieldtype":"Date"})

        row_data = [
            idx,
            closure.get("name", ""),
            closure.get("given_name", ""),
            closure.get("passport_no", ""),
            closure.get("customer", ""),
            closure.get("status", ""),
            closure.get("age", ""),
            last_update,
            closure.get("remark", ""),
            closure.get("sa_name", ""),
            closure.get("sa_mobile_number", ""),
        ]

        ws.append(row_data)
        row = ws.max_row
        
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"J{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        if idx % 2 == 0: 
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','J','K']:
                ws[f"{col}{row}"].fill = even_row_fill
        
    for r in ws.iter_rows(min_row=1, max_row=row,
                                min_col=1, max_col=11):
            for cell in r:
                cell.border = border  
                
    for row in ws.iter_rows(min_row=4, max_row=row,
                                min_col=1, max_col=11):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30              
    

    
    
    
    
    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    encoded = base64.b64encode(file_data.getvalue()).decode()
    filename = "Direct_follow_up_Closure.xlsx"
    return {"filename": filename, "data": encoded}



@frappe.whitelist()
def download_supplier_excel(status,territory=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SUPPLIER CLOSURE"
    
    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    even_row_fill = PatternFill(start_color="e6f2f1", end_color="e6f2f1", fill_type="solid")

    
    
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
    # if territory:    
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    
    filters["nationality"] = "Indian"  
               

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","standard_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        
        c["custom_history"] = history
        
        
    for c in closures:
        c["age"] = calculate_age_from_history(c.get("custom_history"))
    
    # Sort closures by age (ascending)
    closures.sort(key=lambda x: x["age"], reverse=True)
    
        
    
    ws.merge_cells("A1:K2")    
    ws["A1"]="SUPPLIER CLOSURE"
    ws["A1"].font = Font(bold=True, size=14,color="FFFFFF")
    ws["A1"].alignment = center    
    ws["A1"].fill = header_fill   
        
        
    ws["A3"]="S#" 
    ws["A3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["A3"].alignment = center
    ws["A3"].fill = header_fill
       
    ws["B3"]="CLID"
    ws["B3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["B3"].alignment = center
    ws["B3"].fill = header_fill
         
    ws["C3"]="Name"
    ws["C3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["C3"].alignment = center
    ws["C3"].fill = header_fill
         
    ws["D3"]="PP#"
    ws["D3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["D3"].alignment = center
    ws["D3"].fill = header_fill
         
    ws["E3"]="Client"
    ws["E3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["E3"].alignment = center
    ws["E3"].fill = header_fill
    
         
    ws["F3"]="Status"
    ws["F3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["F3"].alignment = center
    ws["F3"].fill = header_fill
         
    ws["G3"]="Age"
    ws["G3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["G3"].alignment = center
    ws["G3"].fill = header_fill
         
    ws["H3"]="Last Update On"
    ws["H3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["H3"].alignment = center
    ws["H3"].fill = header_fill
        
    ws["I3"]="Latest Remark"
    ws["I3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["I3"].alignment = center
    ws["I3"].fill = header_fill
    
        
    ws["J3"]="Supplier"
    ws["J3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["J3"].alignment = center
    ws["J3"].fill = header_fill 
       
    ws["K3"]="Contact"
    ws["K3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["K3"].alignment = center
    ws["K3"].fill = header_fill    
    
    column_widths = {
        "A": 10, "B": 12, "C": 35, "D": 20, "E": 40, "F": 18, "G": 10, "H": 15,"I":60,"J":20,"K":40
        
    }
 
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    
    
    
    for idx, closure in enumerate(closures, start=1):
        age = calculate_age_from_history(closure.get("custom_history"))
        last_update = frappe.format(closure.get("last_updated_on"),{"fieldtype":"Date"})
        mob = frappe.db.get_value("Supplier",{"name": closure.get("associate", "")},"mobile_no")

        row_data = [
            idx,
            closure.get("name", ""),
            closure.get("given_name", ""),
            closure.get("passport_no", ""),
            closure.get("customer", ""),
            closure.get("status", ""),
            closure.get("age", ""),
            last_update,
            closure.get("remark", ""),
            closure.get("associate", ""),
            mob,
        ]

        ws.append(row_data)
        row = ws.max_row
        
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"J{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        if idx % 2 == 0: 
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','J','K']:
                ws[f"{col}{row}"].fill = even_row_fill
        
    for r in ws.iter_rows(min_row=1, max_row=row,
                                min_col=1, max_col=11):
            for cell in r:
                cell.border = border 
                
    for row in ws.iter_rows(min_row=4, max_row=row,
                                min_col=1, max_col=11):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30               
    

    
    
    
    
    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    encoded = base64.b64encode(file_data.getvalue()).decode()
    filename = "supplier_follow_up_Closure.xlsx"
    return {"filename": filename, "data": encoded}



@frappe.whitelist()
def download_client_excel(status,territory=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CLIENT CLOSURE"
    
    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    even_row_fill = PatternFill(start_color="e6f2f1", end_color="e6f2f1", fill_type="solid")

    
    
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
    # if territory:    
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    # else:
    #     filters["territory"] = ["not in", "Nepal","Sri Lanka"]
    
    filters["nationality"] = "Indian"         

    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","standard_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        
        c["custom_history"] = history
        
        
    for c in closures:
        c["age"] = calculate_age_from_history(c.get("custom_history"))
    
    # Sort closures by age (ascending)
    closures.sort(key=lambda x: x["age"], reverse=True)
    
        
    
    ws.merge_cells("A1:I2")    
    ws["A1"]="CLIENT CLOSURE"
    ws["A1"].font = Font(bold=True, size=14,color="FFFFFF")
    ws["A1"].alignment = center    
    ws["A1"].fill = header_fill   
        
        
    ws["A3"]="S#" 
    ws["A3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["A3"].alignment = center
    ws["A3"].fill = header_fill
       
    ws["B3"]="CLID"
    ws["B3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["B3"].alignment = center
    ws["B3"].fill = header_fill
         
    ws["C3"]="Name"
    ws["C3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["C3"].alignment = center
    ws["C3"].fill = header_fill
         
    ws["D3"]="PP#"
    ws["D3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["D3"].alignment = center
    ws["D3"].fill = header_fill
         
    ws["E3"]="Client"
    ws["E3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["E3"].alignment = center
    ws["E3"].fill = header_fill
    
         
    ws["F3"]="Status"
    ws["F3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["F3"].alignment = center
    ws["F3"].fill = header_fill
         
    ws["G3"]="Age"
    ws["G3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["G3"].alignment = center
    ws["G3"].fill = header_fill
         
    ws["H3"]="Last Update On"
    ws["H3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["H3"].alignment = center
    ws["H3"].fill = header_fill
        
    ws["I3"]="Latest Remark"
    ws["I3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["I3"].alignment = center
    ws["I3"].fill = header_fill    
    
    
    column_widths = {
        "A": 10, "B": 12, "C": 35, "D": 20, "E": 40, "F": 18, "G": 10, "H": 15,"I":60,
        
    }
 
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    
    
    for idx, closure in enumerate(closures, start=1):
        age = calculate_age_from_history(closure.get("custom_history"))
        last_update = frappe.format(closure.get("last_updated_on"),{"fieldtype":"Date"})

        row_data = [
            idx,
            closure.get("name", ""),
            closure.get("given_name", ""),
            closure.get("passport_no", ""),
            closure.get("customer", ""),
            closure.get("status", ""),
            closure.get("age", ""),
            last_update,
            closure.get("remark", "")
        ]


        ws.append(row_data)
        row = ws.max_row
        
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if idx % 2 == 0: 
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws[f"{col}{row}"].fill = even_row_fill
        
        
    for r in ws.iter_rows(min_row=1, max_row=row,
                                min_col=1, max_col=9):
            for cell in r:
                cell.border = border  
                
    for row in ws.iter_rows(min_row=4, max_row=row,
                                min_col=1, max_col=9):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30              
    

    
    
    
    
    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    encoded = base64.b64encode(file_data.getvalue()).decode()
    filename = "Client_Closure.xlsx"
    return {"filename": filename, "data": encoded}


@frappe.whitelist()
def download_nepal_excel(status,territory=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NEPAL CLOSURE"
    
    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    even_row_fill = PatternFill(start_color="e6f2f1", end_color="e6f2f1", fill_type="solid")

    
    
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
    
    # if territory:
    #     filters["territory"] = ["in", ["Nepal"]]
    # else:      
    #     filters["territory"] = ["in", ["Nepal"]]
    
    filters["nationality"] = "Nepali"  
    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","standard_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        
        c["custom_history"] = history
        
        
    for c in closures:
        c["age"] = calculate_age_from_history(c.get("custom_history"))
    
    # Sort closures by age (ascending)
    closures.sort(key=lambda x: x["age"], reverse=True)
    
        
    
    ws.merge_cells("A1:I2")    
    ws["A1"]="NEPAL CLOSURE"
    ws["A1"].font = Font(bold=True, size=14,color="FFFFFF")
    ws["A1"].alignment = center    
    ws["A1"].fill = header_fill   
        
        
    ws["A3"]="S#" 
    ws["A3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["A3"].alignment = center
    ws["A3"].fill = header_fill
       
    ws["B3"]="CLID"
    ws["B3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["B3"].alignment = center
    ws["B3"].fill = header_fill
         
    ws["C3"]="Name"
    ws["C3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["C3"].alignment = center
    ws["C3"].fill = header_fill
         
    ws["D3"]="PP#"
    ws["D3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["D3"].alignment = center
    ws["D3"].fill = header_fill
         
    ws["E3"]="Client"
    ws["E3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["E3"].alignment = center
    ws["E3"].fill = header_fill
    
         
    ws["F3"]="Status"
    ws["F3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["F3"].alignment = center
    ws["F3"].fill = header_fill
         
    ws["G3"]="Age"
    ws["G3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["G3"].alignment = center
    ws["G3"].fill = header_fill
         
    ws["H3"]="Last Update On"
    ws["H3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["H3"].alignment = center
    ws["H3"].fill = header_fill
        
    ws["I3"]="Latest Remark"
    ws["I3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["I3"].alignment = center
    ws["I3"].fill = header_fill    
        
    
    column_widths = {
        "A": 10, "B": 12, "C": 35, "D": 20, "E": 40, "F": 18, "G": 10, "H": 15,"I":60,
        
    }
 
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    
    if closures:
        for idx, closure in enumerate(closures, start=1):
            age = calculate_age_from_history(closure.get("custom_history"))
            last_update = frappe.format(closure.get("last_updated_on"),{"fieldtype":"Date"})

            row_data = [
                idx,
                closure.get("name", ""),
                closure.get("given_name", ""),
                closure.get("passport_no", ""),
                closure.get("customer", ""),
                closure.get("status", ""),
                closure.get("age", ""),
                last_update,
                closure.get("remark", "")
            ]


            ws.append(row_data)
            row = ws.max_row
            
            ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if idx % 2 == 0: 
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    ws[f"{col}{row}"].fill = even_row_fill
            
        for r in ws.iter_rows(min_row=1, max_row=row,
                                    min_col=1, max_col=9):
                for cell in r:
                    cell.border = border   
                    
        for row in ws.iter_rows(min_row=4, max_row=row,
                                min_col=1, max_col=9):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30             
        

        
        
        
        
        file_data = io.BytesIO()
        wb.save(file_data)
        file_data.seek(0)

        encoded = base64.b64encode(file_data.getvalue()).decode()
        filename = "Nepal_Closure.xlsx"
        return {"filename": filename, "data": encoded}


@frappe.whitelist()
def download_srilanka_excel(status,territory=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SRILANKA CLOSURE"
    
    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    even_row_fill = PatternFill(start_color="e6f2f1", end_color="e6f2f1", fill_type="solid")

    
    
    filters = {}

    
    if status:
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except json.JSONDecodeError:
                status = [status]
        elif not isinstance(status, list):
            status = [status]
        filters["status"] = ["in", status]
    # if territory:    
    #     filters["territory"] = ["in", ["Sri Lanka"]]
    # else:
    #     filters["territory"] = ["in", ["Sri Lanka"]]
    
    filters["nationality"] = "Srilankan"    
    
    closures = frappe.db.get_all(
        "Closure",
        filters=filters,
        fields=["name", "given_name", "passport_no", "customer","territory","status","remark","last_updated_on","sa_name","sa_mobile_number","associate","standard_remarks","custom_next_follow_up_on"],
        order_by="last_updated_on asc"
    )

    
    for c in closures:
        history = frappe.get_all(
            "Closure Status History",  
            filters={
                "parent": c["name"],
                "parenttype": "Closure",
                "parentfield": "custom_history"
            },
            fields=["date"]
        )
        
        c["custom_history"] = history
        
        
    for c in closures:
        c["age"] = calculate_age_from_history(c.get("custom_history"))
    
    # Sort closures by age (ascending)
    closures.sort(key=lambda x: x["age"], reverse=True)
    
        
    
    ws.merge_cells("A1:I2")    
    ws["A1"]="SRILANKA CLOSURE"
    ws["A1"].font = Font(bold=True, size=14,color="FFFFFF")
    ws["A1"].alignment = center    
    ws["A1"].fill = header_fill   
        
        
    ws["A3"]="S#" 
    ws["A3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["A3"].alignment = center
    ws["A3"].fill = header_fill
       
    ws["B3"]="CLID"
    ws["B3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["B3"].alignment = center
    ws["B3"].fill = header_fill
         
    ws["C3"]="Name"
    ws["C3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["C3"].alignment = center
    ws["C3"].fill = header_fill
         
    ws["D3"]="PP#"
    ws["D3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["D3"].alignment = center
    ws["D3"].fill = header_fill
         
    ws["E3"]="Client"
    ws["E3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["E3"].alignment = center
    ws["E3"].fill = header_fill
    
         
    ws["F3"]="Status"
    ws["F3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["F3"].alignment = center
    ws["F3"].fill = header_fill
         
    ws["G3"]="Age"
    ws["G3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["G3"].alignment = center
    ws["G3"].fill = header_fill
         
    ws["H3"]="Last Update On"
    ws["H3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["H3"].alignment = center
    ws["H3"].fill = header_fill
        
    ws["I3"]="Latest Remark"
    ws["I3"].font = Font(bold=True, size=12,color="FFFFFF")
    ws["I3"].alignment = center
    ws["I3"].fill = header_fill    
      
    
    column_widths = {
        "A": 10, "B": 12, "C": 35, "D": 20, "E": 40, "F": 18, "G": 10, "H": 15,"I":60,
        
    }
 
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    
    if closures:
        for idx, closure in enumerate(closures, start=1):
            age = calculate_age_from_history(closure.get("custom_history"))
            last_update = frappe.format(closure.get("last_updated_on"),{"fieldtype":"Date"})

            row_data = [
                idx,
                closure.get("name", ""),
                closure.get("given_name", ""),
                closure.get("passport_no", ""),
                closure.get("customer", ""),
                closure.get("status", ""),
                closure.get("age", ""),
                last_update,
                closure.get("remark", "")
            ]


            ws.append(row_data)
            row = ws.max_row
            
            ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if idx % 2 == 0: 
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','J']:
                    ws[f"{col}{row}"].fill = even_row_fill
            
        for r in ws.iter_rows(min_row=1, max_row=row,
                                    min_col=1, max_col=9):
                for cell in r:
                    cell.border = border   
                    
        for row in ws.iter_rows(min_row=4, max_row=row,
                                min_col=1, max_col=9):
            for cell in row:
                ws.row_dimensions[cell.row].height = 30             
        

        
        
        
        
        file_data = io.BytesIO()
        wb.save(file_data)
        file_data.seek(0)

        encoded = base64.b64encode(file_data.getvalue()).decode()
        filename = "Srilanka_Closure.xlsx"
        return {"filename": filename, "data": encoded}







from datetime import datetime

def calculate_age_from_history(history):
    if not history:
        return "-"
    try:
        # Get the latest date instead of the oldest
        valid_dates = [h["date"] for h in history if h.get("date")]
        if not valid_dates:
            return "-"
        
        # Sort and pick the latest date
        latest_date = max(valid_dates)
        
        # If the date is in string format, parse it
        if isinstance(latest_date, str):
            latest_date = datetime.strptime(latest_date, "%Y-%m-%d").date()
        else:
            latest_date = latest_date.date() if isinstance(latest_date, datetime) else latest_date
        
        age = (datetime.now().date() - latest_date).days
        return age
    except Exception as e:
        print("Error calculating age:", e)
        return "-"





# @frappe.whitelist()
# def get_candidate_data_candidate_dashboard(customer=None,project=None,pending_for=None,name=None):
#     filters = {}

        
#     if customer:
#         filters["customer"] = customer
    
#     if project:
#         filters["project"] = project
        
#     if name:
#         filters["name"] = name
    
#     if pending_for:
#         filters["pending_for"] = pending_for
#     else:        
        
#         filters["pending_for"] = ["in", ["Submit(SPOC)","Submitted(Client)","Shortlisted","Linedup","Linedup Confirmed","Reported","Interviewed","Result Pending"]]    
    
#     candidate = frappe.db.get_all(
#         "Candidate",
#         filters=filters,
#         fields=["name","given_name", "passport_number", "project", "position","pending_for","mobile_number","whatsapp_number","customer"],
#         order_by="modified desc"
        
#     )

    
   

#     return { "candidate": candidate }


@frappe.whitelist()
def get_candidate_data_candidate_dashboard_old(customer=None,project=None,pending_for=None,name=None):
    filters = {}

    if customer:
        filters["customer"] = customer
    
    if project:
        filters["project"] = project
    else:
        filters["project"] = ["is", "set"]    
        
    if name:
        filters["name"] = name
    
    if pending_for:
        filters["pending_for"] = pending_for
    else:
        filters["pending_for"] = [
            "in",
            [
                "Submit(SPOC)", "Submitted(Client)", "Shortlisted",
                "Linedup", "Linedup Confirmed", "Reported",
                "Interviewed", "Result Pending"
            ]
        ]

    candidates = frappe.db.get_all(
        "Candidate",
        filters=filters,
        fields=[
            "name","given_name","passport_number","project","position",
            "pending_for","mobile_number","whatsapp_number","customer"
        ],
        order_by="modified desc",
        limit_page_length=0
    )
      

    for c in candidates:
        from frappe.utils import now_datetime, get_datetime
        if c.project=="PROJ-2026":
            frappe.log_error(f"Processing candidate: {c.name} with project PROJ-2026")
        status_row = frappe.db.sql("""
            SELECT parent, status, sourced_date
            FROM `tabCandidate status`
            WHERE parent IN %(names)s
        """, {"names": tuple([c.name for c in candidates])}, as_dict=True)

        if status_row and status_row[0].sourced_date:
            status_datetime = get_datetime(status_row[0].sourced_date)
            age_days = (now_datetime() - status_datetime).days
            c["status_age"] = age_days
        else:
            c["status_age"] = 0  
        if c.project:
            proj = frappe.get_value(
                "Project",
                c.project,
                ["tvac", "tsp", "tfp", "tsl", "custom_t_lp"],
                as_dict=True
            )
            c.update(proj)

    return {"candidate": candidates}



@frappe.whitelist()
def get_customer():
    return frappe.get_all("Customer", fields=["name"])

@frappe.whitelist()
def get_project():
    return frappe.get_all("Project", fields=["name"])


@frappe.whitelist()
def get_active_client_data_candidate_dashboard(service=None):

    filters = {}

    if service:
        filters["service"] = service
    else:
        filters["service"] = ["in", ["REC-I", "REC-D", "IT-SW"]]

    filters["party_from"] = "Customer"
    filters["status"] = ["!=", "Do Not Contact"]

    clients = frappe.db.get_all(
        "Sales Follow Up",
        filters=filters,
        fields=["name", "service", "party_name", "status", "remarks"]
    )

    active_rows = []

    for c in clients:
        if not c.party_name:
            continue

        
        active_projects = frappe.db.get_all(
            "Project",
            filters={
                "customer": c.party_name,
                "service": c.service,
                "status": ["in", ["Draft", "Open"]]
            },
            fields=["name", "status"]
        )

        
        if not active_projects:
            continue

        
        for p in active_projects:
            row = c.copy()
            row["project"] = p.name
            row["project_status"] = p.status
            active_rows.append(row)
            
    print(f"Active Rows: {len(active_rows)}")
    return {"client": active_rows}




@frappe.whitelist()
def get_inactive_client_data_candidate_dashboard(service=None):

    filters = {}

    if service:
        filters["service"] = service
    else:
        filters["service"] = ["in", ["REC-I", "REC-D", "IT-SW"]]

    filters["party_from"] = "Customer"
    filters["status"] = ["!=", "Do Not Contact"]

    clients = frappe.db.get_all(
        "Sales Follow Up",
        filters=filters,
        fields=["name", "service", "party_name", "status", "remarks"]
    )

    inactive_rows = []

    for c in clients:
        if not c.party_name:
            continue

        
        active_project = frappe.db.exists(
            "Project",
            {
                "customer": c.party_name,
                "service": c.service,
                "status": ["in", ["Draft", "Open"]]
            }
        )

        
        if active_project:
            continue

        
        inactive_project = frappe.db.exists(
            "Project",
            {
                "customer": c.party_name,
                "service": c.service,
            }
        )

        if inactive_project:
            inactive_rows.append(c)

    return {"client": inactive_rows}



# @frappe.whitelist()
# def get_candidate_by_project_customer(project=None, customer=None):
#     filters = {}

#     if project:
#         filters["project"] = project

#     if customer:
#         filters["customer"] = customer
        
#     filters["pending_for"] = [
#             "in",
#             [
#                 "Submit(SPOC)", "Submitted(Client)", "Shortlisted",
#                 "Linedup", "Linedup Confirmed", "Reported",
#                 "Interviewed", "Result Pending"
#             ]
#         ]    

#     data = frappe.db.get_all(
#         "Candidate",
#         filters=filters,
#         fields=[
#             "name", "given_name", "passport_number",
#             "position", "pending_for", "mobile_number",
#             "whatsapp_number"
#         ],
#         order_by="modified desc"
#     )

#     return data

from frappe.utils import now_datetime, get_datetime
@frappe.whitelist()
def get_candidate_data_candidate_dashboard(project=None):

    filters = {}


    if project:
        filters["project"] = project
    else:
        filters["project"] = ["is", "set"]

    filters["pending_for"] = [
        "in",
        [
            "Submit(SPOC)", "Submitted(Client)", "Shortlisted",
            "Linedup", "Linedup Confirmed", "Reported",
            "Interviewed", "Result Pending"
        ]
    ]

    candidates = frappe.db.get_all(
        "Candidate",
        filters=filters,
        fields=[
            "name","given_name","passport_number","project","position",
            "pending_for","mobile_number","whatsapp_number","customer"
        ],
        order_by="modified desc",
        # limit_page_length=10
    )

    if not candidates:
        return {"candidate": []}

    candidate_names = [c.name for c in candidates]

    status_rows = frappe.db.sql("""
        SELECT parent, sourced_date
        FROM `tabCandidate status`
        WHERE parent IN %(names)s
    """, {"names": tuple(candidate_names)}, as_dict=True)

    status_map = {s.parent: s.sourced_date for s in status_rows}

    project_names = list(set([c.project for c in candidates if c.project]))

    project_rows = frappe.db.get_all(
        "Project",
        filters={"name": ["in", project_names]},
        fields=["name","tvac","tsp","tfp","tsl","custom_t_lp"]
    )

    project_map = {p.name: p for p in project_rows}

    from frappe.utils import now_datetime, get_datetime

    for c in candidates:

        sourced_date = status_map.get(c.name)

        if sourced_date:
            age_days = (now_datetime() - get_datetime(sourced_date)).days
            c["status_age"] = age_days
        else:
            c["status_age"] = 0

        # if c.project and c.project in project_map:
        #     c.update(project_map[c.project])
    return {"candidate": candidates}


# @frappe.whitelist()
# def get_candidate_by_project_customer(project=None, customer=None, pending_for=None, name=None):
#     filters = {}

#     if project:
#         filters["project"] = project

#     if customer:
#         filters["customer"] = customer

#     if name:
#         filters["name"] = name

#     if pending_for:
#         filters["pending_for"] = pending_for
#     else:
#         filters["pending_for"] = [
#             "in",
#             [
#                 "Submit(SPOC)", "Submitted(Client)", "Shortlisted",
#                 "Linedup", "Linedup Confirmed", "Reported",
#                 "Interviewed", "Result Pending"
#             ]
#         ]

#     data = frappe.db.get_all(
#         "Candidate",
#         filters=filters,
#         fields=[
#             "name", "given_name", "passport_number",
#             "position", "pending_for", "mobile_number",
#             "whatsapp_number"
#         ],
#         order_by="modified desc"
#     )
    
#     for c in data:
#         status_row = frappe.db.get_all(
#             "Candidate status",
#             filters={
#                 "parent": c.name,
#                 "status": c.pending_for
#             },
#             fields=["sourced_date"],
#             order_by="sourced_date desc",
#             limit=1
#         )

#         if status_row and status_row[0].sourced_date:
#             status_datetime = get_datetime(status_row[0].sourced_date)
#             diff = now_datetime() - status_datetime
#             c["status_age"] = diff.days
#         else:
#             c["status_age"] = 0
#     return data





@frappe.whitelist()
def update_candidate_status(name, data):
    data = frappe.parse_json(data)
    doc = frappe.get_doc("Candidate", name)

    
    for key, val in data.items():

        
        if isinstance(val, list) and key == "custom_status_transition":
            for row in val:
                doc.append("custom_status_transition", row)

        else:
            doc.set(key, val)

    doc.save(ignore_permissions=True)
    return {"status": "success"}

    
import json

@frappe.whitelist()
def update_closure_remark(closures):

    if isinstance(closures, str):
        closures = json.loads(closures)

    for name, data in closures.items():
        doc = frappe.get_doc("Closure", name)
        if data.get("remark"):
            doc.remark = data.get("remark")
        if data.get("custom_next_follow_up_on"):
            doc.custom_next_follow_up_on = data.get("custom_next_follow_up_on")
        if data.get("standard_remarks"):
            std_remarks = frappe.db.get_value("Standard Remarks", {"standard_remarks": data.get("standard_remarks")}, "name")
            doc.standard_remarks = std_remarks

        doc.save(ignore_permissions=True)

    return {"status": "success"}


@frappe.whitelist()
def get_customers_from_candidate():

    return frappe.db.sql("""
        SELECT
            p.customer,

            SUM(IFNULL(p.tvac,0)) AS vao,
            SUM(IFNULL(p.tsp,0)) AS sp,
            SUM(IFNULL(p.tfp,0)) AS fp,
            SUM(IFNULL(p.tsl,0)) AS sl,
            SUM(IFNULL(p.custom_t_lp,0)) AS lp

        FROM `tabProject` p

        WHERE p.status = 'Open'
        AND p.service IN ('REC-I', 'REC-D')
        AND p.name IN (
            SELECT DISTINCT c.project
            FROM `tabCandidate` c
            WHERE c.customer IS NOT NULL
            AND c.customer != ''
            AND c.pending_for != 'IDB'
            AND c.project IS NOT NULL
        )

        GROUP BY p.customer
        ORDER BY p.customer
    """, as_dict=True)

# @frappe.whitelist()
# def get_projects_from_candidate(customer):

#     return frappe.db.sql("""
#         SELECT
#             p.name AS project,
#             p.project_name,

#             IFNULL(p.tvac,0) AS vao,
#             IFNULL(p.tsp,0) AS sp,
#             IFNULL(p.tfp,0) AS fp,
#             IFNULL(p.tsl,0) AS sl,
#             IFNULL(p.custom_t_lp,0) AS lp

#         FROM `tabProject` p

#         WHERE p.status = 'Open'
#         AND p.service IN ('REC-I', 'REC-D')
#         AND p.customer = %s
#         AND p.name IN (
#             SELECT DISTINCT c.project
#             FROM `tabCandidate` c
#             WHERE c.customer = %s
#             AND c.pending_for != 'IDB'
#             AND c.project IS NOT NULL
#         )

#         ORDER BY p.project_name
#     """, (customer, customer), as_dict=True)

@frappe.whitelist()
def get_projects_from_candidate(customer):

    return frappe.db.sql("""
        SELECT
            p.name AS project,
            p.project_name,
            p.custom_spoc_remark AS remark,
            IFNULL(p.tvac,0) AS vao,
            IFNULL(p.tsp,0) AS sp,
            IFNULL(p.tfp,0) AS fp,
            IFNULL(p.tsl,0) AS sl,
            IFNULL(p.custom_t_lp,0) AS lp

        FROM `tabProject` p
        INNER JOIN `tabCandidate` c
            ON c.project = p.name

        WHERE p.status = 'Open'
        AND p.service IN ('REC-I','REC-D')
        AND c.customer = %s
        AND c.pending_for != 'IDB'

        GROUP BY p.name
        ORDER BY p.project_name
    """, customer, as_dict=True)


# @frappe.whitelist()
# def get_tasks_from_project(project):

#     return frappe.db.sql("""
#         SELECT name, subject, status
#         FROM `tabTask`
#         WHERE project = %s
#         AND status NOT IN ('Hold','Completed','Cancelled')
#         ORDER BY subject
#     """, project, as_dict=True)


@frappe.whitelist()
def get_candidate_status_summary(project):

    return frappe.db.sql("""
        SELECT
            c.task,
            t.subject,
            t.vac AS vac,
            t.sp AS sp,
            c.pending_for,
            COUNT(c.name) as count
        FROM `tabCandidate` c
        JOIN `tabTask` t
            ON t.name = c.task
            AND t.status NOT IN ('Hold','Completed','Cancelled')
        WHERE c.project = %s
        GROUP BY c.task, t.subject, t.vac, t.sp, c.pending_for
    """, (project,), as_dict=True)