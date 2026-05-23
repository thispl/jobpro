import frappe
from frappe.utils import getdate, nowdate

@frappe.whitelist()
def get_order_booking_rec(from_date=None, to_date=None):
    conditions = ["service = 'REC-I'", "docstatus = 1", "status NOT IN ('On Hold', 'Cancelled', 'Closed','Completed')"]
    # if not from_date and not to_date:
    #     today = frappe.utils.today()
    #     fiscal_year = frappe.db.get_value("Fiscal Year", 
    #         filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
    #         fieldname=["year_start_date", "year_end_date"],
    #         as_dict=True
    #     )
    #     from_date = fiscal_year["year_start_date"]
    #     to_date = fiscal_year["year_end_date"]
    if from_date:
        conditions.append("transaction_date >= %(from_date)s")
    if to_date:
        conditions.append("transaction_date <= %(to_date)s")

    query = f"""
        SELECT SUM(base_net_total)
        FROM `tabSales Order`
        WHERE {' AND '.join(conditions)}
    """

    result = frappe.db.sql(query, {'from_date': from_date, 'to_date': to_date})[0][0] or 0
    return result

@frappe.whitelist()
def get_turnover_rec(from_date=None, to_date=None):
    conditions = ["services = 'REC-I'", "docstatus = 1", "status NOT IN ('Return','Draft','Cancelled')"]
    if not from_date and not to_date:
        today = frappe.utils.today()
        fiscal_year = frappe.db.get_value("Fiscal Year", 
            filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
            fieldname=["year_start_date", "year_end_date"],
            as_dict=True
        )
        from_date = fiscal_year["year_start_date"]
        to_date = fiscal_year["year_end_date"]
    if from_date:
        conditions.append("posting_date >= %(from_date)s")
    if to_date:
        conditions.append("posting_date <= %(to_date)s")

    query = f"""
        SELECT SUM(base_net_total)
        FROM `tabSales Invoice`
        WHERE {' AND '.join(conditions)}
    """

    result = frappe.db.sql(query, {'from_date': from_date, 'to_date': to_date})[0][0] or 0
    return result

@frappe.whitelist()
def get_collection_value_rec(from_date=None, to_date=None):
    if not from_date and not to_date:
        today = frappe.utils.today()
        fiscal_year = frappe.db.get_value(
            "Fiscal Year",
            filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
            fieldname=["year_start_date", "year_end_date"],
            as_dict=True
        )
        from_date = fiscal_year["year_start_date"]
        to_date = fiscal_year["year_end_date"]

    conditions = [
        "pe.payment_type = 'Receive'",
        "pe.docstatus = 1",
        "per.service = 'REC-I'"
    ]

    if from_date:
        conditions.append("pe.posting_date >= %(from_date)s")
    if to_date:
        conditions.append("pe.posting_date <= %(to_date)s")

    query = f"""
        SELECT SUM(pe.paid_amount)
        FROM `tabPayment Entry` pe
        INNER JOIN `tabPayment Entry Reference` per ON per.parent = pe.name
        WHERE {' AND '.join(conditions)}
    """

    return frappe.db.sql(query, {'from_date': from_date, 'to_date': to_date})[0][0] or 0

@frappe.whitelist()
def rec_receivable():
    today = frappe.utils.today()
    fiscal_year = frappe.db.get_value(
        "Fiscal Year",
        filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
        fieldname=["year_start_date", "year_end_date"],
        as_dict=True
    )
    from_date = fiscal_year["year_start_date"]
    to_date = fiscal_year["year_end_date"]

    filters = {
        'from_date': from_date,
        'to_date': to_date
    }
    total_invoice = frappe.db.sql("""
        SELECT SUM(outstanding_amount)
        FROM `tabSales Invoice`
        WHERE docstatus = 1
          AND services="REC-I"
          AND outstanding_amount > 0
    """)[0][0] or 0
    return total_invoice

@frappe.whitelist()
def rec_to_bill_value():
    today = frappe.utils.today()
    fiscal_year = frappe.db.get_value(
        "Fiscal Year",
        filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
        fieldname=["year_start_date", "year_end_date"],
        as_dict=True
    )
    from_date = fiscal_year["year_start_date"]
    to_date = fiscal_year["year_end_date"]

    filters = {
        'from_date': from_date,
        'to_date': to_date
    }
    total_invoice = frappe.db.sql("""
        SELECT SUM(base_grand_total)
        FROM `tabSales Order`
        WHERE docstatus = 1
          AND service='REC-I'
          AND status='To Bill'
    """)[0][0] or 0    
    return total_invoice

@frappe.whitelist()
def rec_to_deliver_bill_value():
    today = frappe.utils.today()
    fiscal_year = frappe.db.get_value(
        "Fiscal Year",
        filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
        fieldname=["year_start_date", "year_end_date"],
        as_dict=True
    )
    from_date = fiscal_year["year_start_date"]
    to_date = fiscal_year["year_end_date"]

    filters = {
        'from_date': from_date,
        'to_date': to_date
    }

    total_invoice = frappe.db.sql("""
        SELECT SUM(base_grand_total)
        FROM `tabSales Order`
        WHERE docstatus = 1
          AND service='REC-I'
          AND status='To Deliver and Bill'
    """)[0][0] or 0    

    return total_invoice

@frappe.whitelist()
def rec_payable():
    # if not from_date and not to_date:
    today = frappe.utils.today()
    fiscal_year = frappe.db.get_value(
        "Fiscal Year",
        filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
        fieldname=["year_start_date", "year_end_date"],
        as_dict=True
    )
    from_date = fiscal_year["year_start_date"]
    to_date = fiscal_year["year_end_date"]

    # filters = {
    #     'from_date': from_date,
    #     'to_date': to_date
    # }

    # Total Submitted Purchase Invoice Amount
    # total_invoice = frappe.db.sql("""
    #     SELECT SUM(outstanding_amount)
    #     FROM `tabPurchase Invoice`
    #     WHERE docstatus = 1
    #       AND services="REC-I"
    #       AND posting_date BETWEEN %(from_date)s AND %(to_date)s
    # """, filters)[0][0] or 0
    total_invoice = frappe.db.sql("""
        SELECT SUM(outstanding_amount)
        FROM `tabPurchase Invoice`
        WHERE docstatus = 1
          AND services="REC-I"
          AND outstanding_amount > 0
    """)[0][0] or 0
    return total_invoice

@frappe.whitelist()
def rec_receivable_table():
    from frappe.utils import today, getdate, nowdate, fmt_money
    from datetime import datetime
   
    data = frappe.db.sql("""
        SELECT name, customer, outstanding_amount, posting_date
        FROM `tabSales Invoice`
        WHERE docstatus = 1
          AND services="REC-I"
          AND outstanding_amount > 0
        ORDER BY posting_date
    """, as_dict=True)

    total_outstanding = 0
    today_date = getdate(nowdate())

    html = """
    <div style='max-height: 340px; overflow-y: auto; overflow-x: auto;'>
        <div style='min-width: 500px;'>
            <table class='table table-bordered' style='width: 100%; border-collapse: collapse;'>
                <thead>
                    <tr>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">S.No</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Customer</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Value</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Age</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Sales Invoice</th>
                    </tr>
                </thead>
                <tbody>
    """

    for idx, row in enumerate(data, 1):
        age = (today_date - getdate(row.posting_date)).days
        row_style = "color: red;" if age > 30 else ""
        name_style = "color: red;" if age > 30 else ""
        total_outstanding += row.outstanding_amount or 0

        html += f"""
            <tr style="{row_style}">
                <td style="text-align: center;">{idx}</td>
                <td style="white-space: nowrap;text-align: left;">{row.customer}</td>
                <td style='text-align:right;'>{fmt_money(row.outstanding_amount)}</td>
                <td style='text-align:right;'>{age}</td>
                <td style="white-space: nowrap;text-align: left;"><a href="/app/sales-invoice/{ row.name }" target="_blank" style="{name_style}">{ row.name }</a></td>
            </tr>
        """

    # Grand Total row
    html += f"""
        <tr style="background: #f0f0f0; font-weight: bold;">
            <td colspan="2" style="text-align: center;">Total</td>
            <td style="text-align: right;">{fmt_money(total_outstanding)}</td>
            <td></td>
            <td></td>
        </tr>
    """

    html += """
                </tbody>
            </table>
        </div>
    </div>
    """
    return html

@frappe.whitelist()
def rec_payable_table():
    from frappe.utils import today, getdate, nowdate
    from datetime import datetime
    from frappe.utils import today, getdate, nowdate, fmt_money
    # fiscal_year = frappe.db.get_value(
    #     "Fiscal Year",
    #     filters={"year_start_date": ["<=", today()], "year_end_date": [">=", today()]},
    #     fieldname=["year_start_date", "year_end_date"],
    #     as_dict=True
    # )
    # from_date = fiscal_year["year_start_date"]
    # to_date = fiscal_year["year_end_date"]

    # filters = {
    #     'from_date': from_date,
    #     'to_date': to_date
    # }

    # data = frappe.db.sql("""
    #     SELECT name, supplier, outstanding_amount, posting_date
    #     FROM `tabPurchase Invoice`
    #     WHERE docstatus = 1
    #       AND services="REC-I"
    #       AND posting_date BETWEEN %(from_date)s AND %(to_date)s
    #       AND outstanding_amount > 0
    #     ORDER BY posting_date
    # """, filters, as_dict=True)
    data = frappe.db.sql("""
        SELECT name, supplier, outstanding_amount, posting_date
        FROM `tabPurchase Invoice`
        WHERE docstatus = 1
          AND services="REC-I"
          AND outstanding_amount > 0
        ORDER BY posting_date
    """, as_dict=True)

    html = """
    <div style='max-height: 340px; overflow-y: auto; overflow-x: auto;'>
        <div style='min-width: 500px;'>
            <table class='table table-bordered' style='width: 100%; border-collapse: collapse;'>
                <thead>
                    <tr>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">S.No</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center; white-space: nowrap;">Supplier</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Value</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Age</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center; white-space: nowrap;">Purchase Invoice</th>
                    </tr>
                </thead>
                <tbody>
    """
    today_date = getdate(nowdate())
    total_outstanding = 0
    for idx, row in enumerate(data, 1):
        age = (today_date - getdate(row.posting_date)).days
        row_style = "color: red;" if age > 30 else ""  # Apply to whole row
        name_style = "color: red;" if age > 30 else ""
        total_outstanding += row.outstanding_amount or 0
        html += f"""
            <tr style="{row_style}">
            <td style="text-align: center;">{idx}</td>
                <td style="white-space: nowrap;text-align: left;">{row.supplier}</td>
                <td style='text-align:right;'>{frappe.utils.fmt_money(row.outstanding_amount)}</td>
                <td style='text-align:right;'>{age}</td>
                <td style="white-space: nowrap;text-align: left;"><a href="/app/purchase-invoice/{ row.name }" target="_blank" style="{name_style}">{ row.name }</a></td>
            </tr>
        """
    html += f"""
        <tr style="background: #f0f0f0; font-weight: bold;">
            <td colspan="2" style="text-align: center;">Total</td>
            <td style="text-align: right;">{fmt_money(total_outstanding)}</td>
            <td></td>
            <td></td>
        </tr>
    """

    html += """
                </tbody>
            </table>
        </div>
    </div>
    """
    return html

@frappe.whitelist()
def rec_tobill_table():
    from frappe.utils import today, getdate, nowdate
    from datetime import datetime
    from frappe.utils import today, getdate, nowdate, fmt_money
    from datetime import datetime
   
    data = frappe.db.sql("""
        SELECT name, customer, base_grand_total, transaction_date
        FROM `tabSales Order`
        WHERE docstatus = 1
          AND service='REC-I'
          AND status='To Bill'
          AND base_grand_total > 0
        ORDER BY transaction_date
    """, as_dict=True)
    html = """
    <div style='max-height: 340px; overflow-y: auto; overflow-x: auto;'>
        <div style='min-width: 500px;'>
            <table class='table table-bordered' style='width: 100%; border-collapse: collapse;'>
                <thead>
                    <tr>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">S.No</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center; white-space: nowrap;">Customer</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Value</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center;">Age</th>
                        <th style="position: sticky; top: 0; background: #002060; color: white; text-align: center; white-space: nowrap;">Sales Order</th>
                    </tr>
                </thead>
                <tbody>
    """

    total_outstanding = 0
    today_date = getdate(nowdate())
    for idx, row in enumerate(data, 1):
        age = (today_date - getdate(row.transaction_date)).days
        row_style = "color: red;" if age > 30 else ""  # Apply to whole row
        name_style = "color: red;" if age > 30 else ""
        total_outstanding += row.base_grand_total or 0
        html += f"""
            <tr style="{row_style}">
                <td style="text-align: center;">{idx}</td>
                <td style="white-space: nowrap;text-align: left;">{row.customer}</td>
                <td style='text-align:right;'>{frappe.utils.fmt_money(row.base_grand_total)}</td>
                <td style='text-align:right;'>{age}</td>
                <td style="white-space: nowrap;text-align: left;"><a href="/app/sales-order/{ row.name }" target="_blank" style="{name_style}">{ row.name }</a></td>
            </tr>
        """
    html += f"""
        <tr style="background: #f0f0f0; font-weight: bold;">
            <td colspan="2" style="text-align: center;" >Total</td>
            <td style="text-align: right;">{fmt_money(total_outstanding)}</td>
            <td></td>
            <td></td>
        </tr>
    """

    html += """
                </tbody>
            </table>
        </div>
    </div>
    """
    return html

@frappe.whitelist()
def get_project_count():
    return frappe.db.count("Project", filters={
        "status": ["in", ["Open"]],
        "service": "REC-I"
    })
    
@frappe.whitelist()
def get_teampro_closure_count():
    return frappe.db.count("Closure",filters={"status":["in",["PSL", "Emigration","Ticket", "Onboarding"]],"nationality":"Indian"})  
     
@frappe.whitelist()
def get_candidate_agent_closure_count():
    return frappe.db.count("Closure",filters={"status":["in",["Signed Offer Letter","Premedical","PCC","Final Medical"]],"nationality":"Indian","sa_name": ["is", "not set"]}) 

@frappe.whitelist()
def get_agent_closure_count():
    return frappe.db.count("Closure",filters={"status":["in",["Signed Offer Letter","Premedical","PCC","Final Medical"]],"nationality":"Indian","sa_name": ["is", "set"]}) 
 
@frappe.whitelist()
def get_supp_agent_closure_count():
    return frappe.db.count("Closure",filters={"status":["in",["Certificate Attestation","Biometric","Trade Test","Visa Stamping"]],"nationality":"Indian"})  
     
@frappe.whitelist()
def get_client_closure_count():
    return frappe.db.count("Closure",filters={"status":["in",["Client Offer Letter","Visa"]],"nationality":"Indian"}) 



@frappe.whitelist()
def get_so_pending_count():
    
    count = frappe.db.count("Closure", filters={
        "status": ["not in", ["Dropped", "Arrived"]],
        "so_created": 0,
        "nationality":"Indian"
    })
    
    
    so_pending_curr = frappe.db.sql("""
        SELECT 
            SUM(client_payment_company_currency) AS client_sum,
            SUM(candidate_payment_company_currenc) AS candidate_sum,
            SUM(custom_associate_payment_company_currency) AS associate_sum
        FROM `tabClosure`
        WHERE status NOT IN ("Dropped", "Arrived")
        AND nationality = 'Indian'
        AND so_created = 0
    """, as_dict=True)  
    
    if so_pending_curr:
        
        client_sum = so_pending_curr[0].get("client_sum", 0) or 0
        candidate_sum = so_pending_curr[0].get("candidate_sum", 0) or 0
        associate_sum = so_pending_curr[0].get("associate_sum", 0) or 0

        
        total_sum = client_sum + candidate_sum + associate_sum

        return {
            "count": count,  
            "total": total_sum  
        }
    else:
        return {
            "count": count,  
            "total": 0  
        }

      
@frappe.whitelist()
def get_nepal_closure_count():
    return frappe.db.count("Closure",filters={"status":["in",["PSL", "Emigration","Ticket", "Onboarding","Signed Offer Letter","Premedical","PCC","Final Medical","Certificate Attestation","Biometric","Trade Test","Visa Stamping"]],"nationality":"Nepali"})   
    
@frappe.whitelist()
def get_srilanka_closure_count():
    return frappe.db.count("Closure",filters={"status":["in",["PSL", "Emigration","Ticket", "Onboarding","Signed Offer Letter","Premedical","PCC","Final Medical","Certificate Attestation","Biometric","Trade Test","Visa Stamping"]],"nationality":"Srilankan"})       

@frappe.whitelist()
def get_task_count():
    count = frappe.db.sql("""
        SELECT COUNT(*)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return count


@frappe.whitelist()
def get_vacancy_count():
    total = frappe.db.sql("""
        SELECT SUM(t.vac)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_sp_count():
   
    total = frappe.db.sql("""
        SELECT SUM(t.sp)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_task_sp_count():
   
    total = frappe.db.sql("""
        SELECT COUNT(t.name)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
            AND t.custom_task_sourcing_status = 'SP'
            AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
            AND t.status IN ('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0

    return total

@frappe.whitelist()
def get_task_fp_count():
   
    total = frappe.db.sql("""
        SELECT COUNT(t.name)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
            AND t.custom_task_sourcing_status='FP'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_task_spfp_count():
   
    total = frappe.db.sql("""
        SELECT COUNT(t.name)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
            AND t.custom_task_sourcing_status='SP/FP'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_canidate_ip_count():
   
    total = frappe.db.sql("""
        SELECT count(c.name)
        FROM `tabCandidate` c
        WHERE 
          c.pending_for in ("Sourced","Pending QC")
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_canidate_fp_count():
   
    total = frappe.db.sql("""
        SELECT count(c.name)
        FROM `tabCandidate` c
        WHERE 
          c.pending_for in ("Submit(SPOC)","Submitted(Client)","Interviewed","Reported","Result Pending")
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_canidate_sl_count():
   
    total = frappe.db.sql("""
        SELECT count(c.name)
        FROM `tabCandidate` c
        WHERE 
          c.pending_for in ("Linedup","Linedup Confirmed","Shortlisted")
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_canidate_psl_count():
   
    total = frappe.db.sql("""
        SELECT count(c.name)
        FROM `tabCandidate` c
        WHERE 
          c.pending_for="Proposed PSL"
    """)[0][0] or 0
    return total


@frappe.whitelist()
def get_fp_count():
    
    total = frappe.db.sql("""
        SELECT SUM(t.fp)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_sl_count():
    total = frappe.db.sql("""
        SELECT SUM(t.sl)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_psl_count():
    total = frappe.db.sql("""
        SELECT SUM(t.psl)
        FROM `tabTask` t
        JOIN `tabProject` p ON t.project = p.name
        WHERE p.service = 'REC-I'
          AND p.status NOT IN ('Draft','Enquiry','Hold','Completed','Cancelled')
          AND t.status IN('Open','Working','Overdue','Pending Review')
    """)[0][0] or 0
    return total

@frappe.whitelist()
def get_territory_status_matrix():
    from datetime import datetime
    from collections import defaultdict

    settings = frappe.get_single("Closure Settings")

    tat_map = {}
    global_tat_map = {}

    for row in settings.closure_tat_days:
        if row.territory:
            tat_map[(row.status, row.territory)] = row.tat_days
        else:
            global_tat_map[row.status] = row.tat_days

    history_map = defaultdict(list)
    history_data = frappe.db.sql("""
        SELECT parent, status, date
        FROM `tabClosure Status History`
        ORDER BY parent, date DESC
    """, as_dict=True)

    for row in history_data:
        history_map[row.parent].append(row)
        if len(history_map[row.parent]) > 2:
            history_map[row.parent] = history_map[row.parent][:2]

    # Fetch all closures with so_created flag
    closures = frappe.get_all(
        "Closure",
        filters={"nationality":"Indian"},
        fields=["name", "territory", "status", "so_created"]
    )

    matrix = {}

    for closure in closures:
        closure_id = closure.name
        territory = closure.territory
        status = closure.status
        so_created = closure.so_created or 0  # checkbox → 0 or 1
        history = history_map.get(closure_id, [])

        if not territory or not status:
            continue

        # Determine applicable TAT
        tat_days = tat_map.get((status, territory)) or global_tat_map.get(status)
        tat_crossed = False

        if tat_days and history:
            start_date_raw = history[1]["date"] if len(history) >= 2 else history[0]["date"]
            try:
                start_date = start_date_raw.date() if isinstance(start_date_raw, datetime) else \
                    datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S.%f").date()
            except ValueError:
                start_date = datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S").date()

            days_elapsed = (datetime.now().date() - start_date).days
            tat_crossed = days_elapsed > tat_days

        # Initialize matrix
        if territory not in matrix:
            matrix[territory] = {}
        if status not in matrix[territory]:
            matrix[territory][status] = {
                "count": 0,
                "crossed_ids": [],
                "so_not_created_count": 0
            }

        # Count logic
        matrix[territory][status]["count"] += 1
        if tat_crossed:
            matrix[territory][status]["crossed_ids"].append(closure_id)
        if not so_created:  # so_created == 0
            matrix[territory][status]["so_not_created_count"] += 1

    # Prepare result list
    result = []
    for territory, statuses in matrix.items():
        for status, data in statuses.items():
            result.append({
                "territory": territory,
                "status": status,
                "count": data["count"],
                "tat_crossed_count": len(data["crossed_ids"]),
                "closure_ids": data["crossed_ids"],
                "so_not_created_count": data["so_not_created_count"]
            })

    return result









# @frappe.whitelist()
# def get_project_details_for_territory(territory):
#     from datetime import datetime
#     from collections import defaultdict

#     settings = frappe.get_single("Closure Settings")
#     tat_map = {}
#     global_tat_map = {}

#     # Build territory-specific and global TAT maps
#     for row in settings.closure_tat_days:
#         if row.territory:
#             tat_map[(row.status, row.territory)] = row.tat_days
#         else:
#             global_tat_map[row.status] = row.tat_days

#     # Load closure history sorted by latest date
#     history_map = defaultdict(list)
#     history_data = frappe.db.sql("""
#         SELECT parent, status, date
#         FROM `tabClosure Status History`
#         ORDER BY parent, date DESC
#     """, as_dict=True)

#     for row in history_data:
#         history_map[row.parent].append(row)
#         if len(history_map[row.parent]) > 2:
#             history_map[row.parent] = history_map[row.parent][:2]

#     # Load closures in that territory
#     closures = frappe.get_all("Closure", 
#         filters={"territory": territory},
#         fields=["name", "status", "project"]
#     )

#     project_matrix = {}

#     for closure in closures:
#         closure_id = closure.name
#         status = closure.status
#         project = closure.project
#         history = history_map.get(closure_id, [])

#         if not status or not project:
#             continue

#         tat_days = tat_map.get((status, territory)) or global_tat_map.get(status)
#         tat_crossed = False

#         if tat_days and history:
#             start_date_raw = history[1]["date"] if len(history) >= 2 else history[0]["date"]
#             try:
#                 start_date = start_date_raw.date() if isinstance(start_date_raw, datetime) else \
#                     datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S.%f").date()
#             except ValueError:
#                 start_date = datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S").date()

#             days_elapsed = (datetime.now().date() - start_date).days
#             tat_crossed = days_elapsed > tat_days

#         if project not in project_matrix:
#             project_matrix[project] = {}
#         if status not in project_matrix[project]:
#             project_matrix[project][status] = {
#                 "count": 0,
#                 "crossed_ids": []
#             }

#         project_matrix[project][status]["count"] += 1
#         if tat_crossed:
#             project_matrix[project][status]["crossed_ids"].append(closure_id)

#     # Load project names
#     project_names = frappe._dict({
#         row.name: row.project_name for row in frappe.get_all("Project", fields=["name", "project_name","customer"], order_by="customer ASC")
#     })

#     # Build output list
#     result = []
#     for project, statuses in project_matrix.items():
#         for status, data in statuses.items():
#             result.append({
#                 "project": project,
#                 "project_name": project_names.get(project, project),
#                 "status": status,
#                 "count": data["count"],
#                 "tat_crossed_count": len(data["crossed_ids"]),
#                 "closure_ids": data["crossed_ids"]
#             })

#     return result



@frappe.whitelist()
def get_project_details_for_territory(territory):
    from datetime import datetime
    from collections import defaultdict

    settings = frappe.get_single("Closure Settings")
    tat_map = {}
    global_tat_map = {}

    # Build territory-specific and global TAT maps
    for row in settings.closure_tat_days:
        if row.territory:
            tat_map[(row.status, row.territory)] = row.tat_days
        else:
            global_tat_map[row.status] = row.tat_days

    # Load closure history sorted by latest date
    history_map = defaultdict(list)
    history_data = frappe.db.sql("""
        SELECT parent, status, date
        FROM `tabClosure Status History`
        ORDER BY parent, date DESC
    """, as_dict=True)

    for row in history_data:
        history_map[row.parent].append(row)
        if len(history_map[row.parent]) > 2:
            history_map[row.parent] = history_map[row.parent][:2]

    # ✅ Load closures in that territory (include so_created)
    closures = frappe.get_all(
        "Closure",
        filters={"territory": territory,"nationality":"Indian"},
        fields=["name", "status", "project", "so_created"]
    )

    project_matrix = {}

    for closure in closures:
        closure_id = closure.name
        status = closure.status
        project = closure.project
        so_created = closure.so_created  or 0 # ✅ new field
        history = history_map.get(closure_id, [])

        if not status or not project:
            continue

        tat_days = tat_map.get((status, territory)) or global_tat_map.get(status)
        tat_crossed = False

        if tat_days and history:
            start_date_raw = history[1]["date"] if len(history) >= 2 else history[0]["date"]
            try:
                start_date = start_date_raw.date() if isinstance(start_date_raw, datetime) else \
                    datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S.%f").date()
            except ValueError:
                start_date = datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S").date()

            days_elapsed = (datetime.now().date() - start_date).days
            tat_crossed = days_elapsed > tat_days

        if project not in project_matrix:
            project_matrix[project] = {}
        if status not in project_matrix[project]:
            project_matrix[project][status] = {
                "count": 0,
                "crossed_ids": [],
                "so_not_created_count": 0  # ✅ initialize
            }

        project_matrix[project][status]["count"] += 1

        # ✅ Track TAT crossed
        if tat_crossed:
            project_matrix[project][status]["crossed_ids"].append(closure_id)

        # ✅ Track SO not created
        if not so_created:
            project_matrix[project][status]["so_not_created_count"] += 1
            
            

    # Load project names
    project_names = frappe._dict({
        row.name: row.project_name for row in frappe.get_all(
            "Project",
            fields=["name", "project_name", "customer"],
            order_by="customer ASC"
        )
    })

    # Build output list
    result = []
    for project, statuses in project_matrix.items():
        for status, data in statuses.items():
            result.append({
                "project": project,
                "project_name": project_names.get(project, project),
                "status": status,
                "count": data["count"],
                "tat_crossed_count": len(data["crossed_ids"]),
                "closure_ids": data["crossed_ids"],
                "so_not_created_count": data["so_not_created_count"]  # ✅ include in output
            })

    return result




@frappe.whitelist()
def get_candidates_tat_crossed_from_history():
    from datetime import datetime
    from frappe.utils import now_datetime

    target_statuses = ["Submit(SPOC)", "Submitted(Client)", "Interviewed"]

    closures = frappe.get_all("Candidate",
        filters={"pending_for": ["in", target_statuses]},
        fields=["name", "given_name", "pending_for", "project", "territory", "subject"]
    )

    project_map = {}

    for closure in closures:
        pending_for = closure.pending_for

        history = frappe.get_all("Candidate status",
            filters={"parent": closure.name, "status": pending_for},
            fields=["sourced_date"],
            order_by="sourced_date desc",
            limit=1
        )

        if history:
            start_date = history[0]["sourced_date"]
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")

            days_pending = (now_datetime().date() - start_date.date()).days
            if days_pending > 10:
                project_id = closure.project or "No Project"
                project_name = frappe.db.get_value("Project", project_id, "project_name") if project_id and project_id != "No Project" else "No Project"
                subject = closure.subject or "-"
                territory = closure.territory or ""

                if project_name not in project_map:
                    project_map[project_name] = {
                        "territory": territory,
                        "status_map": []
                    }

                # Group by status + subject
                status_list = project_map[project_name]["status_map"]
                status_entry = next(
                    (s for s in status_list if s["status"] == pending_for and s["subject"] == subject),
                    None
                )

                candidate_data = {
                    "closure_id": closure.name,
                }

                if status_entry:
                    status_entry["count"] += 1
                    status_entry["candidates"].append(candidate_data)
                else:
                    status_list.append({
                        "status": pending_for,
                        "subject": subject,
                        "count": 1,
                        "candidates": [candidate_data]
                    })

    return project_map

@frappe.whitelist()
def get_planned_work_monitor_filter(start_date=None, end_date=None, executive=None, week_plan=None):
    import frappe
    from frappe.utils import getdate, add_days, formatdate
    if week_plan:
        parent_docs = frappe.get_doc("REC Week Plan",week_plan)
    else:
        parent_docs = frappe.get_all("REC Week Plan", filters={
            "start_date": ["<=", end_date],
            "end_date": [">=", start_date],
        }, fields=["name"])
        parent_docs = [frappe.get_doc("REC Week Plan", d.name) for d in parent_docs]
    # Parse dates
    start_date = getdate(start_date)
    end_date = getdate(end_date)
   
    # Generate date range
    date_list = []
    current = start_date
    while current <= end_date:
        date_list.append(current)
        current = add_days(current, 1)

    result = {}

    def init_task_entry(subject, project,project_name=None):
        return {
            "subject": subject,
            "project": project,
            "project_name": project_name,
            "dates": {str(d): {"rt": 0, "ac": 0} for d in date_list}
        }

    # Process allocation (RT)
    for parent_doc in parent_docs if isinstance(parent_docs, list) else [parent_docs]:
        for row in parent_doc.allocation:
            if not row.exe:
                continue
            row_date = getdate(row.date)
            if start_date <= row_date <= end_date and (not executive or row.exe == executive):
                result.setdefault(row.exe, {})
                task_map = result[row.exe]
                if row.task not in task_map:
                    project_name = frappe.db.get_value("Project", row.project, "project_name") if row.project else ""
                    task_map[row.task] = init_task_entry(row.subject, row.project, project_name)
                    # task_map[row.task] = init_task_entry(row.subject, row.project)
                task_map[row.task]["dates"][str(row_date)]["rt"] = row.rc or 0
                task_map[row.task]["dates"][str(row_date)]["ac"] += row.ac or 0

    # Process dsr (AC)
    for parent_doc in parent_docs if isinstance(parent_docs, list) else [parent_docs]:
        for row in parent_doc.dsr:
            if not row.exe:
                continue
            row_date = getdate(row.date)
            if start_date <= row_date <= end_date and (not executive or row.exe == executive):
                result.setdefault(row.exe, {})
                task_map = result[row.exe]
                if row.task not in task_map:
                    project_name = frappe.db.get_value("Project", row.project, "project_name") if row.project else ""
                    task_map[row.task] = init_task_entry(row.subject, row.project, project_name)

                    # task_map[row.task] = init_task_entry(row.subject, row.project)
                task_map[row.task]["dates"][str(row_date)]["ac"] += row.ac or 0
                task_map[row.task]["dates"][str(row_date)]["rt"] += row.rc or 0
    executive_meta = {}
    for exe in result.keys():
        emp = frappe.db.get_value("Employee", {"user_id": exe}, ["employee_name", "image"], as_dict=True)
        if emp:
            executive_meta[exe] = {
                "name": emp.employee_name,
                "image": emp.image
            }

    return {
        "date_headers": [formatdate(d, "d MMM") for d in date_list],
        "raw_dates": [str(d) for d in date_list],
        "data": result,
        "executive_meta": executive_meta
    }
    

@frappe.whitelist()
def get_planned_work_monitor(start_date=None, end_date=None, executive=None, week_plan=None):
    import frappe
    from frappe.utils import getdate, add_days, formatdate
    if week_plan:
        parent_doc = frappe.get_doc("REC Week Plan",week_plan)
    else:
        parent_doc = frappe.get_doc("REC Week Plan", {"workflow_state": "Planned"})
    # Parse dates
    start_date = getdate(start_date)
    end_date = getdate(end_date)
   
    # Generate date range
    date_list = []
    current = start_date
    while current <= end_date:
        date_list.append(current)
        current = add_days(current, 1)

    result = {}

    def init_task_entry(subject, project,project_name=None):
        return {
            "subject": subject,
            "project": project,
            "project_name": project_name,
            "dates": {str(d): {"rt": 0, "ac": 0} for d in date_list}
        }

    # Process allocation (RT)
    for row in parent_doc.allocation:
        row_date = getdate(row.date)
        if start_date <= row_date <= end_date and (not executive or row.exe == executive):
            result.setdefault(row.exe, {})
            task_map = result[row.exe]
            if row.task not in task_map:
                project_name = frappe.db.get_value("Project", row.project, "project_name") if row.project else ""
                task_map[row.task] = init_task_entry(row.subject, row.project, project_name)
                # task_map[row.task] = init_task_entry(row.subject, row.project)
            task_map[row.task]["dates"][str(row_date)]["rt"] = row.rc or 0
            task_map[row.task]["dates"][str(row_date)]["ac"] += row.ac or 0

    # Process dsr (AC)
    for row in parent_doc.dsr:
        row_date = getdate(row.date)
        if start_date <= row_date <= end_date and (not executive or row.exe == executive):
            result.setdefault(row.exe, {})
            task_map = result[row.exe]
            if row.task not in task_map:
                project_name = frappe.db.get_value("Project", row.project, "project_name") if row.project else ""
                task_map[row.task] = init_task_entry(row.subject, row.project, project_name)

                # task_map[row.task] = init_task_entry(row.subject, row.project)
            task_map[row.task]["dates"][str(row_date)]["ac"] += row.ac or 0
            task_map[row.task]["dates"][str(row_date)]["rt"] += row.rc or 0
    # Get Employee Names and Images
    executive_meta = {}
    for exe in result.keys():
        emp = frappe.db.get_value("Employee", {"user_id": exe}, ["employee_name", "image"], as_dict=True)
        if emp:
            executive_meta[exe] = {
                "name": emp.employee_name,
                "image": emp.image
            }

    return {
        "date_headers": [formatdate(d, "d MMM") for d in date_list],
        "raw_dates": [str(d) for d in date_list],
        "data": result,
        "executive_meta": executive_meta
    }


# import frappe
# from frappe.utils import flt
# from collections import defaultdict

# @frappe.whitelist()
# def get_payment_collection_details(from_date=None, to_date=None):
#     if not from_date and not to_date:
#         today = frappe.utils.today()
#         fiscal_year = frappe.db.get_value(
#             "Fiscal Year", 
#             filters={
#                 "year_start_date": ["<=", today], 
#                 "year_end_date": [">=", today]
#             },
#             fieldname=["year_start_date", "year_end_date"],
#             as_dict=True
#         )
#         if fiscal_year:
#             from_date = fiscal_year["year_start_date"]
#             to_date = fiscal_year["year_end_date"]

#     payment_entries = frappe.db.sql("""
#         SELECT per.allocated_amount, per.reference_name AS sales_invoice, pe.name AS payment_entry
#         FROM `tabPayment Entry Reference` per
#         INNER JOIN `tabPayment Entry` pe ON pe.name = per.parent
#         WHERE pe.docstatus = 1
#           AND per.reference_doctype = 'Sales Invoice'
#           AND pe.payment_type = 'Receive'
#           AND pe.posting_date BETWEEN %s AND %s
#     """, (from_date, to_date), as_dict=True)

#     grouped_data = defaultdict(lambda: {
#         "employee_code": "",
#         "employee_name": "",
#         "department": "",
#         "role": "",
#         "collection_value": 0.0,
#         "net_collection": 0.0
#     })

#     for ref in payment_entries:
#         si = frappe.get_doc("Sales Invoice", ref.sales_invoice)
#         if not si:
#             continue

#         participants = []

#         if si.get("account_manager"):
#             participants.append({"employee": si.account_manager, "role": "AM"})

#         if si.get("delivery_manager"):
#             participants.append({"employee": si.delivery_manager, "role": "DM"})

#         if si.get("services") == "REC-I":
#             for item in si.items:
#                 if item.get("candidate_owner"):
#                     participants.append({"employee": item.candidate_owner, "role": "CO"})

#         if not participants:
#             continue

#         has_tax = 1 if si.get("taxes") else 0
#         allocated = flt(ref.allocated_amount)

#         # Precompute net after tax
#         net_total = allocated / 1.18 if has_tax else allocated

#         for p in participants:
#             emp = frappe.db.get_value("Employee", {"user_id": p["employee"], "status": "Active"},
#                                     ["name", "employee_name", "department"],
#                                     as_dict=True)
#             if not emp:
#                 continue

#             key = (emp.name, p["role"])

#             # Default to total allocated amount
#             share = allocated
#             net_share = net_total

#             if p["role"] == "CO":
#                 # Sum amounts from items owned by this candidate_owner
#                 item_share = 0.0
#                 net_item_share = 0.0
#                 for item in si.items:
#                     if item.candidate_owner == p["employee"]:
#                         item_share += flt(item.amount)
#                         net_item_share += item_share / 1.18 if has_tax else item_share
#                 share = item_share
#                 net_share = net_item_share

#             grouped_data[key]["employee_code"] = emp.name
#             grouped_data[key]["employee_name"] = emp.employee_name
#             grouped_data[key]["department"] = emp.department
#             grouped_data[key]["role"] = p["role"]
#             grouped_data[key]["collection_value"] += round(share, 2)
#             grouped_data[key]["net_collection"] += round(net_share, 2)


#     return sorted(
#     grouped_data.values(),
#     key=lambda x: (
#         {"AM": 0, "DM": 1, "CO": 2}.get(x["role"], 99),
#         x["employee_name"].lower() if x["employee_name"] else ""
#     )
# )

import frappe
from frappe.utils import flt
from collections import defaultdict

@frappe.whitelist()
def get_payment_collection_details(from_date=None, to_date=None):
    if not from_date and not to_date:
        today = frappe.utils.today()
        fiscal_year = frappe.db.get_value(
            "Fiscal Year",
            filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
            fieldname=["year_start_date", "year_end_date"],
            as_dict=True
        )
        if fiscal_year:
            from_date = fiscal_year["year_start_date"]
            to_date = fiscal_year["year_end_date"]

    payment_entries = frappe.db.sql("""
        SELECT per.allocated_amount, per.reference_name AS sales_invoice, pe.name AS payment_entry
        FROM `tabPayment Entry Reference` per
        INNER JOIN `tabPayment Entry` pe ON pe.name = per.parent
        WHERE pe.docstatus = 1
          AND per.reference_doctype = 'Sales Invoice'
          AND pe.payment_type = 'Receive'
          AND pe.posting_date BETWEEN %s AND %s
    """, (from_date, to_date), as_dict=True)

    grouped_data = defaultdict(lambda: {
        "employee_code": "",
        "employee_name": "",
        "department": "",
        "role": "",
        "collection_value": 0.0,
        "net_collection": 0.0
    })

    for ref in payment_entries:
        si = frappe.get_doc("Sales Invoice", ref.sales_invoice)
        if not si:
            continue

        allocated = flt(ref.allocated_amount)
        has_tax = 1 if si.get("taxes") else 0

        is_rec_i = si.services == "REC-I"
        candidate_owners = set()
        participants = []

        if is_rec_i:
            # Add candidate owners from items
            for item in si.items:
                if item.candidate_owner:
                    candidate_owners.add(item.candidate_owner)
                    participants.append({"employee": item.candidate_owner, "role": "CO"})

            # Add AM only if not in candidate_owners
            if si.account_manager and si.account_manager not in candidate_owners:
                participants.append({"employee": si.account_manager, "role": "AM"})

            # Add DM only if not in candidate_owners
            if si.delivery_manager and si.delivery_manager not in candidate_owners:
                participants.append({"employee": si.delivery_manager, "role": "DM"})
        else:
            # For non REC-I, include AM and DM directly
            if si.account_manager:
                participants.append({"employee": si.account_manager, "role": "AM"})
            if si.delivery_manager:
                participants.append({"employee": si.delivery_manager, "role": "DM"})

        if not participants:
            continue

        for p in participants:
            emp = frappe.db.get_value("Employee", {"user_id": p["employee"], "status": "Active"},
                                      ["name", "employee_name", "department"],
                                      as_dict=True)
            if not emp:
                continue

            key = (emp.name, p["role"])

            share = 0.0
            net_share = 0.0

            if is_rec_i:
                # Share per item if service is REC-I
                for item in si.items:
                    item_amount = flt(item.amount)
                    net_amount = item_amount / 1.18 if has_tax else item_amount

                    if p["role"] == "CO" and item.candidate_owner == p["employee"]:
                        share += item_amount
                        net_share += net_amount

                    elif p["role"] in ("AM", "DM"):
                        share += item_amount
                        net_share += net_amount
            else:
                # If not REC-I, full allocated goes to AM/DM
                share = allocated
                net_share = allocated / 1.18 if has_tax else allocated

            grouped_data[key]["employee_code"] = emp.name
            grouped_data[key]["employee_name"] = emp.employee_name
            grouped_data[key]["department"] = emp.department
            grouped_data[key]["role"] = p["role"]
            grouped_data[key]["collection_value"] += round(share, 2)
            grouped_data[key]["net_collection"] += round(net_share, 2)

    # Sort by role and then name
    return sorted(
        grouped_data.values(),
        key=lambda x: (
            {"AM": 0, "DM": 1, "CO": 2}.get(x["role"], 99),
            x["employee_name"].lower() if x["employee_name"] else ""
        )
    )


import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from frappe.utils import today
from io import BytesIO
from .rec_i_dashboard import get_payment_collection_details

@frappe.whitelist()
def download_payment_excel(from_date=None, to_date=None):
    if not from_date or not to_date:
        today_date = today()
        fiscal = frappe.db.get_value("Fiscal Year", {
            "year_start_date": ["<=", today_date],
            "year_end_date": [">=", today_date]
        }, ["year_start_date", "year_end_date"], as_dict=True)
        if fiscal:
            from_date = fiscal.year_start_date
            to_date = fiscal.year_end_date

    data = get_payment_collection_details(from_date=from_date, to_date=to_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Collection"

    headers = ["S.No", "Employee Code", "Employee Name", "Department", "Role", "Collection Value", "Net Collection"]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="FF1E0C6F", end_color="FF1E0C6F", fill_type="solid")  # Deep Blue
    header_font = Font(bold=True, color="FFFFFFFF")  # White text
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = u'₹#,##0.00'  # Indian currency format

    # Apply header styles
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Add data rows with S.No
    for idx, row_data in enumerate(data, start=1):
        row = [
            idx,
            row_data.get("employee_code"),
            row_data.get("employee_name"),
            row_data.get("department") or "-",
            row_data.get("role"),
            float(row_data.get("collection_value") or 0),
            float(row_data.get("net_collection") or 0)
        ]
        ws.append(row)

    # Apply border and currency format
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border

        # Format collection columns
        row[5].number_format = currency_format  # Collection Value
        row[6].number_format = currency_format  # Net Collection

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Return file to browser
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from frappe.utils import formatdate

    filename = f"Payment_Collection_{formatdate(from_date, 'dd-mm-yyyy')}_to_{formatdate(to_date, 'dd-mm-yyyy')}.xlsx"

    frappe.response['filename'] = filename
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'
    frappe.response['headers'] = {
        'Content-Disposition': f'attachment; filename={filename}',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }

# import frappe
# from frappe.utils import flt
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from io import BytesIO
# from frappe import _

# @frappe.whitelist()
# def download_payment_details_excel(from_date=None, to_date=None):
#     frappe.response['filename'] = 'Payment Collection Details.xlsx'
#     frappe.response['type'] = 'binary'
#     frappe.response['filecontent'] = generate_payment_excel(from_date, to_date)

# def generate_payment_excel(from_date=None, to_date=None):
#     if not from_date and not to_date:
#         today = frappe.utils.today()
#         fiscal_year = frappe.db.get_value(
#             "Fiscal Year",
#             filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
#             fieldname=["year_start_date", "year_end_date"],
#             as_dict=True
#         )
#         if fiscal_year:
#             from_date = fiscal_year["year_start_date"]
#             to_date = fiscal_year["year_end_date"]

#     payment_entries = frappe.db.sql("""
#         SELECT per.allocated_amount, per.reference_name AS sales_invoice, pe.name AS payment_entry
#         FROM `tabPayment Entry Reference` per
#         INNER JOIN `tabPayment Entry` pe ON pe.name = per.parent
#         WHERE pe.docstatus = 1
#           AND per.reference_doctype = 'Sales Invoice'
#           AND pe.payment_type = 'Receive'
#           AND pe.posting_date BETWEEN %s AND %s
#     """, (from_date, to_date), as_dict=True)

#     rows = []
#     for ref in payment_entries:
#         si = frappe.get_doc("Sales Invoice", ref.sales_invoice)
#         if not si:
#             continue

#         has_tax = 1 if si.get("taxes") else 0
#         allocated = flt(ref.allocated_amount)
#         net_total = allocated / 1.18 if has_tax else allocated

#         participants = []

#         if si.account_manager:
#             participants.append({"employee": si.account_manager, "role": "AM"})

#         if si.delivery_manager:
#             participants.append({"employee": si.delivery_manager, "role": "DM"})

#         if si.services == "REC-I":
#             for item in si.items:
#                 if item.candidate_owner:
#                     participants.append({"employee": item.candidate_owner, "role": "CO"})

#         if not participants:
#             continue

#         for p in participants:
#             emp = frappe.db.get_value("Employee", {"user_id": p["employee"], "status": "Active"},
#                                       ["name", "employee_name", "department"],
#                                       as_dict=True)
#             if not emp:
#                 continue

#             share = allocated
#             net_share = net_total

#             if p["role"] == "CO":
#                 item_share = 0.0
#                 net_item_share = 0.0
#                 for item in si.items:
#                     if item.candidate_owner == p["employee"]:
#                         amount = flt(item.amount)
#                         item_share += amount
#                         net_item_share += amount / 1.18 if has_tax else amount
#                 share = item_share
#                 net_share = net_item_share

#             rows.append({
#                 "payment_entry": ref.payment_entry,
#                 "sales_invoice": ref.sales_invoice,
#                 "allocated_amount": round(share, 2),
#                 "net_amount": round(net_share, 2),
#                 "employee_code": emp.name,
#                 "employee_name": emp.employee_name,
#                 "department": emp.department,
#                 "role": p["role"]
#             })

#     # Create Excel
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Payment Collection Details"

#     headers = [
#         "Payment Entry", "Sales Invoice", "Allocated Amount", "Net Amount",
#         "Employee Code", "Employee Name", "Department", "Role"
#     ]
#     ws.append(headers)

#     for row in rows:
#         ws.append([
#             row["payment_entry"],
#             row["sales_invoice"],
#             row["allocated_amount"],
#             row["net_amount"],
#             row["employee_code"],
#             row["employee_name"],
#             row["department"],
#             row["role"]
#         ])

#     # Formatting: Auto width
#     for col in ws.columns:
#         max_length = 0
#         column = get_column_letter(col[0].column)
#         for cell in col:
#             try:
#                 max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         ws.column_dimensions[column].width = max_length + 2

#     # Return binary
#     output = BytesIO()
#     wb.save(output)
#     return output.getvalue()
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import frappe
from frappe.utils import flt
from frappe import _

@frappe.whitelist()
def download_payment_details_excel(from_date=None, to_date=None):
    frappe.response['filename'] = 'Payment Collection Details.xlsx'
    frappe.response['type'] = 'binary'
    frappe.response['filecontent'] = generate_payment_excel(from_date, to_date)

def generate_payment_excel(from_date=None, to_date=None):
    if not from_date and not to_date:
        today = frappe.utils.today()
        fiscal_year = frappe.db.get_value(
            "Fiscal Year",
            filters={"year_start_date": ["<=", today], "year_end_date": [">=", today]},
            fieldname=["year_start_date", "year_end_date"],
            as_dict=True
        )
        if fiscal_year:
            from_date = fiscal_year["year_start_date"]
            to_date = fiscal_year["year_end_date"]

    payment_entries = frappe.db.sql("""
        SELECT per.allocated_amount, per.reference_name AS sales_invoice, pe.name AS payment_entry
        FROM `tabPayment Entry Reference` per
        INNER JOIN `tabPayment Entry` pe ON pe.name = per.parent
        WHERE pe.docstatus = 1
          AND per.reference_doctype = 'Sales Invoice'
          AND pe.payment_type = 'Receive'
          AND pe.posting_date BETWEEN %s AND %s
    """, (from_date, to_date), as_dict=True)

    rows = []

    for ref in payment_entries:
        si = frappe.get_doc("Sales Invoice", ref.sales_invoice)
        if not si:
            continue

        has_tax = 1 if si.get("taxes") else 0
        allocated_total = flt(ref.allocated_amount)
        net_total = allocated_total / 1.18 if has_tax else allocated_total

        candidate_owners = set()
        allocations = []

        if si.services == "REC-I":
            for item in si.items:
                item_amount = flt(item.amount)
                item_net = item_amount / 1.18 if has_tax else item_amount

                if item.candidate_owner:
                    candidate_owners.add(item.candidate_owner)
                    allocations.append({
                        "employee": item.candidate_owner,
                        "role": "CO",
                        "amount": item_amount,
                        "net_amount": item_net
                    })

                if si.account_manager and si.account_manager not in candidate_owners:
                    allocations.append({
                        "employee": si.account_manager,
                        "role": "AM",
                        "amount": item_amount,
                        "net_amount": item_net
                    })

                if si.delivery_manager and si.delivery_manager not in candidate_owners:
                    allocations.append({
                        "employee": si.delivery_manager,
                        "role": "DM",
                        "amount": item_amount,
                        "net_amount": item_net
                    })

        else:
            if si.account_manager:
                allocations.append({
                    "employee": si.account_manager,
                    "role": "AM",
                    "amount": allocated_total,
                    "net_amount": net_total
                })
            if si.delivery_manager:
                allocations.append({
                    "employee": si.delivery_manager,
                    "role": "DM",
                    "amount": allocated_total,
                    "net_amount": net_total
                })

        for alloc in allocations:
            emp = frappe.db.get_value("Employee", {"user_id": alloc["employee"], "status": "Active"},
                                      ["name", "employee_name", "department"], as_dict=True)
            if not emp:
                continue

            rows.append({
                "payment_entry": ref.payment_entry,
                "sales_invoice": ref.sales_invoice,
                "allocated_amount": round(alloc["amount"], 2),
                "net_amount": round(alloc["net_amount"], 2),
                "employee_code": emp.name,
                "employee_name": emp.employee_name,
                "department": emp.department,
                "role": alloc["role"]
            })

    # === Create Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Collection Details"

    # === Styles ===
    header_fill = PatternFill(start_color="FF1E0C6F", end_color="FF1E0C6F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = u'₹#,##0.00'

    # === Headers ===
    headers = [
        "S.No", "Payment Entry", "Sales Invoice","Employee Code", "Employee Name", "Department", "Role", "Allocated Amount", "Net Amount"
        
    ]
    ws.append(headers)

    # Apply styles to header
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # === Data Rows ===
    for idx, row in enumerate(rows, start=1):
        ws.append([
            idx,
            row["payment_entry"],
            row["sales_invoice"],
            row["employee_code"],
            row["employee_name"],
             row["department"],
            row["role"],
            row["allocated_amount"],
            row["net_amount"],
        ])

    # === Apply currency format and borders ===
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=9):
        for cell in row:
            cell.border = thin_border
        row[3].number_format = currency_format  # Allocated Amount
        row[4].number_format = currency_format  # Net Amount

    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@frappe.whitelist()
def get_rec_live_status():
    date = nowdate()
    rows = frappe.db.sql("""
        SELECT 
            recruiter,
            position,
            status,
            cnt
        FROM (
            SELECT 
                c.candidate_created_by AS recruiter,
                c.position,
                cs.status,
                COUNT(DISTINCT c.name) AS cnt,
                ROW_NUMBER() OVER (PARTITION BY c.name ORDER BY cs.sourced_date DESC) AS rn
            FROM `tabCandidate` c
            INNER JOIN `tabCandidate status` cs ON cs.parent = c.name
            WHERE c.candidate_created_by IS NOT NULL
            AND c.position IS NOT NULL
            AND DATE(cs.sourced_date) = %s
            GROUP BY c.name, c.candidate_created_by, c.position, cs.status, cs.sourced_date
        ) t
        WHERE rn = 1
        ORDER BY recruiter
    """, (date,), as_dict=True)

    data = {}
    for r in rows:
        recruiter = r.recruiter or "Unassigned"
        if recruiter not in data:
            data[recruiter] = {
                "tasks": [],
                "idb": 0,
                "source": 0,
                "qc": 0,
                "spoc": 0
            }

        task = next((t for t in data[recruiter]["tasks"] if t["position"] == r.position), None)
        if not task:
            task = {
                "position": r.position,
                "idb": 0,
                "source": 0,
                "qc": 0,
                "spoc": 0
            }
            data[recruiter]["tasks"].append(task)

        if r.status == "IDB":
            data[recruiter]["idb"] += r.cnt
            task["idb"] += r.cnt
        elif r.status == "Sourced":
            data[recruiter]["source"] += r.cnt
            task["source"] += r.cnt
        elif r.status == "Pending QC":
            data[recruiter]["qc"] += r.cnt
            task["qc"] += r.cnt
        elif r.status == "Submit(SPOC)":
            data[recruiter]["spoc"] += r.cnt
            task["spoc"] += r.cnt

    for rec in data:
        data[rec]["positions_count"] = len({t["position"] for t in data[rec]["tasks"]})

    return {"data": data}



# @frappe.whitelist()
# def get_current_status_counts(executive, project, task=None, date=None):
#     import frappe
#     from frappe.utils import getdate

#     date = getdate(date) if date else getdate()
#     current_statuses = ["IDB", "Sourced", "Pending QC", "Submit(SPOC)"]
#     counts = {}

#     task_condition = "AND cs.task = %s" if task else ""
#     task_value = (task,) if task else ()
#     for status in current_statuses:
#         if status == "Submit(SPOC)":
#             query = f"""
#                 SELECT DISTINCT c.name
#                 FROM `tabCandidate` c
#                 INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
#                 WHERE c.candidate_created_by = %s
#                 AND cs.project = %s
#                 AND cs.status = %s
#                 {task_condition}
#                 AND DATE(cs.sourced_date) = %s
#             """
#             params = (executive, project, status) + task_value + (date,)
#         else:
#             query = f"""
#                 SELECT DISTINCT c.name
#                 FROM `tabCandidate` c
#                 INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
#                 WHERE c.candidate_created_by = %s
#                 AND c.pending_for = %s
#                 {task_condition}
#                 AND DATE(cs.sourced_date) = %s
#             """
#             params = (executive, status) + task_value + (date,)

#         ppl = frappe.db.sql(query, params, as_dict=True)
#         counts[status] = len(ppl)

#     extra_query = f"""
#         SELECT DISTINCT c.name, cs.project, cs.task, cs.status
#         FROM `tabCandidate` c
#         INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
#         WHERE c.candidate_created_by = %s
#         {task_condition}
#         AND DATE(cs.sourced_date) = %s
#         AND cs.status NOT IN %s
#     """
#     extra_params = (executive,) + task_value + (date, tuple(current_statuses))
#     extra_candidates = frappe.db.sql(extra_query, extra_params, as_dict=True)

#     info_query = f"""
#         SELECT DISTINCT cs.project, cs.task
#         FROM `tabCandidate` c
#         INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
#         WHERE c.candidate_created_by = %s
#         {task_condition}
#         AND DATE(cs.sourced_date) = %s
#     """
#     info_params = (executive,) + task_value + (date,)
#     project_task_info = frappe.db.sql(info_query, info_params, as_dict=True)

#     return {
#         "counts": counts,
#         "extra_candidates": extra_candidates,
#         "project_task_info": project_task_info
#     }


import frappe
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from io import BytesIO
import base64

@frappe.whitelist()
def download_closure_matrix_with_projects():
    settings = frappe.get_single("Closure Settings")

    tat_map, global_tat_map = {}, {}
    for row in settings.closure_tat_days:
        if row.territory:
            tat_map[(row.status, row.territory)] = row.tat_days
        else:
            global_tat_map[row.status] = row.tat_days

    history_map = defaultdict(list)
    history_data = frappe.db.sql("""
        SELECT parent, status, date
        FROM `tabClosure Status History`
        ORDER BY parent, date DESC
    """, as_dict=True)
    for row in history_data:
        history_map[row.parent].append(row)
        if len(history_map[row.parent]) > 2:
            history_map[row.parent] = history_map[row.parent][:2]

    closures = frappe.get_all("Closure",filters={"nationality":"Indian"}, fields=["name", "territory", "status", "project","so_created"])

    project_names = frappe._dict({
        row.name: row.project_name for row in frappe.get_all("Project", fields=["name", "project_name"])
    })

    header_name = [
        "PSL", "WL", "SO", "COL", "SOL",
        "Visa", "PM", "PCC", "CA", "TT", "FM",
        "BIO", "VS", "POE", "TKT", "OB", "OBD"
    ]
    
    statuses = [
        "PSL", "Waitlisted", "Sales Order", "Client Offer Letter", "Signed Offer Letter",
        "Visa", "Premedical", "PCC", "Certificate Attestation", "Trade Test", "Final Medical",
        "Biometric", "Visa Stamping", "Emigration", "Ticket", "Onboarding", "Onboarded"
    ]

    territory_matrix = defaultdict(lambda: defaultdict(lambda: {"count": 0, "crossed_count": 0,"so_not_created_count": 0}))
    project_matrix = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {"count": 0, "crossed_count": 0, "so_not_created_count": 0})))

    for closure in closures:
        territory, status, project, closure_id = closure.territory, closure.status, closure.project, closure.name
        so_created = closure.get("so_created") or 0
        if not territory or not status:
            continue

        history = history_map.get(closure_id, [])
        tat_days = tat_map.get((status, territory)) or global_tat_map.get(status)
        tat_crossed = False

        if tat_days and history:
            start_date_raw = history[1]["date"] if len(history) >= 2 else history[0]["date"]
            try:
                start_date = start_date_raw.date() if isinstance(start_date_raw, datetime) else datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S.%f").date()
            except ValueError:
                start_date = datetime.strptime(str(start_date_raw), "%Y-%m-%d %H:%M:%S").date()
            tat_crossed = (datetime.now().date() - start_date).days > tat_days

        territory_matrix[territory][status]["count"] += 1
        territory_matrix[territory][status]["so_not_created_count"] += 0 if so_created else 1
        if tat_crossed:
            territory_matrix[territory][status]["crossed_count"] += 1

        if project:
            project_matrix[territory][project][status]["count"] += 1
            project_matrix[territory][project][status]["so_not_created_count"] += 0 if so_created else 1
            if tat_crossed:
                project_matrix[territory][project][status]["crossed_count"] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Closure Matrix"

    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    territory_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Entire territory row color
    project_fill = PatternFill(start_color="F1F9FF", end_color="F1F9FF", fill_type="solid")
    red_font = Font(color="FF0000")
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Territory/Project"] + header_name + ["Total"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    for territory, statuses_map in territory_matrix.items():
        terr_total = sum([statuses_map[st]["count"] for st in statuses])
        terr_so_not_created_total = sum([statuses_map[st]["so_not_created_count"] for st in statuses])
        

        if terr_total == 0:
            continue

        terr_row = [territory]
        for st in statuses:
            data = statuses_map[st]
            # count = statuses_map[st]["count"]
            # crossed = statuses_map[st]["crossed_count"]
            count = data["count"]
            crossed = data["crossed_count"]
            so_not_created = data["so_not_created_count"]
            if count == 0:
                terr_row.append("-")
            elif crossed > 0:
                terr_row.append(f"{count if count > 0 else '-'}/{crossed if crossed > 0 else '-'}({so_not_created if so_not_created > 0 else '-'})")
            else:
                terr_row.append(f"{count if count > 0 else '-'}({so_not_created if so_not_created > 0 else '-'})")
        # terr_row.append(terr_total)
        terr_row.append(f"{terr_total if terr_total>0 else '-'} ({terr_so_not_created_total if terr_so_not_created_total>0 else '-'})")
        ws.append(terr_row)

        for col_idx in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = border
            cell.fill = territory_fill
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = center_align
            if col_idx >= 2 and col_idx <= len(statuses)+1 and statuses_map[statuses[col_idx-2]]["crossed_count"] > 0:
                cell.font = red_font

        for project, proj_statuses in project_matrix[territory].items():
            proj_total = sum([proj_statuses[st]["count"] for st in statuses])
            proj_so_not_created_total = sum([proj_statuses[st]["so_not_created_count"] for st in statuses])
            

            if proj_total == 0:
                continue
            proj_row = [f"   ↳ {project_names.get(project, project)}"]
            for st in statuses:
                data = proj_statuses[st]
                # count = proj_statuses[st]["count"]
                # crossed = proj_statuses[st]["crossed_count"]
                count = data["count"]
                crossed = data["crossed_count"]
                so_not_created = data["so_not_created_count"]
                if count == 0:
                    proj_row.append("-")
                elif crossed > 0:
                    proj_row.append(f"{count if count > 0 else '-'}/{crossed if crossed > 0 else '-'}({so_not_created if so_not_created > 0 else '-'})")
                else:
                    proj_row.append(f"{count if count > 0 else '-'}({so_not_created if so_not_created > 0 else '-'})")
            # proj_row.append(proj_total)
            proj_row.append(f"{proj_total if proj_total>0 else '-'} ({proj_so_not_created_total if proj_so_not_created_total>0 else '-'})")
            ws.append(proj_row)
            for col_idx in range(1, len(headers)+1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.fill = project_fill
                cell.border = border
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = center_align
                if col_idx >= 2 and col_idx <= len(statuses)+1 and proj_statuses[statuses[col_idx-2]]["crossed_count"] > 0:
                    cell.font = red_font

    grand_row = ["Grand Total"]
    for st in statuses:
        total_count = sum(territory_matrix[territory][st]["count"] for territory in territory_matrix)
        so_not_created_total = sum(territory_matrix[territory][st]["so_not_created_count"] for territory in territory_matrix)
        if total_count == 0:
            grand_row.append("-")
        else:    
            grand_row.append(f"{total_count if total_count>0 else '-'} ({so_not_created_total if so_not_created_total>0 else '-'})")


        # grand_row.append(total_count)
    # grand_row.append(sum([cell for cell in grand_row[1:]])) 
    grand_total_count = sum([territory_matrix[territory][st]["count"] for territory in territory_matrix for st in statuses])
    grand_so_not_created_total = sum([territory_matrix[territory][st]["so_not_created_count"] for territory in territory_matrix for st in statuses])
    if grand_total_count == 0:
            grand_row.append("-")
    else:
        grand_row.append(f"{grand_total_count if grand_total_count>0 else '-'} ({grand_so_not_created_total if grand_so_not_created_total>0 else '-'})")

    ws.append(grand_row) 
    # ws.append(grand_row)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = bold_font
        cell.border = border
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.alignment = center_align

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return base64.b64encode(output.read()).decode("utf-8")

@frappe.whitelist()
def get_current_status_counts(executive, project=None, task=None, date=None):
    import frappe
    from frappe.utils import getdate

    date = getdate(date) if date else getdate()
    current_statuses = ["IDB", "Sourced", "Pending QC", "Submit(SPOC)"]
    counts = {}

    task_condition = "AND cs.task = %s" if task else ""
    task_value = (task,) if task else ()

    # 🔹 Project/Task-specific counts
    for status in current_statuses:
        if status == "IDB":
            query = f"""
                SELECT DISTINCT c.name
                FROM `tabCandidate` c
                INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
                WHERE c.candidate_created_by = %s
                AND cs.project = %s
                AND cs.status = 'IDB'
                AND c.pending_for = 'IDB'
                {task_condition}
                AND DATE(cs.sourced_date) = %s
            """
            params = (executive, project) + task_value + (date,)
        elif status == "Submit(SPOC)":
            query = f"""
                SELECT DISTINCT c.name
                FROM `tabCandidate` c
                INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
                WHERE c.candidate_created_by = %s
                AND cs.project = %s
                AND cs.status = %s
                {task_condition}
                AND DATE(cs.sourced_date) = %s
            """
            params = (executive, project, status) + task_value + (date,)
        else:
            query = f"""
                SELECT DISTINCT c.name
                FROM `tabCandidate` c
                INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
                WHERE c.candidate_created_by = %s
                AND c.pending_for = %s
                {task_condition}
                AND DATE(cs.sourced_date) = %s
            """
            params = (executive, status) + task_value + (date,)

        ppl = frappe.db.sql(query, params, as_dict=True)
        counts[status] = len(ppl)

    # 🔹 Extra candidates (outside current_statuses)
    extra_query = f"""
        SELECT DISTINCT c.name, cs.project, cs.task, cs.status
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.candidate_created_by = %s
        {task_condition}
        AND DATE(cs.sourced_date) = %s
        AND cs.status NOT IN %s
    """
    extra_params = (executive,) + task_value + (date, tuple(current_statuses))
    extra_candidates = frappe.db.sql(extra_query, extra_params, as_dict=True)

    # 🔹 Project/Task info
    info_query = f"""
        SELECT DISTINCT cs.project, cs.task
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.candidate_created_by = %s
        {task_condition}
        AND DATE(cs.sourced_date) = %s
    """
    info_params = (executive,) + task_value + (date,)
    project_task_info = frappe.db.sql(info_query, info_params, as_dict=True)

    # 🔹 Executive-only counts
    task_candidates = frappe.db.sql(f"""
        SELECT DISTINCT c.name
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.candidate_created_by = %s
        AND cs.project = %s
    """, (executive, project), as_dict=True) if project else []

    task_candidate_names = [d.name for d in task_candidates]

    executive_only_counts = {}
    executive_idb_only = 0

    for status in current_statuses:
        if status == "IDB":
            # Executive-only IDB
            q = """
                SELECT COUNT(DISTINCT c.name) AS cnt
                FROM `tabCandidate` c
                INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
                WHERE c.candidate_created_by = %s
                AND DATE(cs.sourced_date) = %s
                AND cs.status = 'IDB'
                AND c.pending_for = 'IDB'
                AND DATE(c.creation) BETWEEN %s AND %s
            """
            params = [executive, date,date,date]
            if task_candidate_names:
                q += " AND c.name NOT IN %s"
                params.append(tuple(task_candidate_names))

            row = frappe.db.sql(q, params, as_dict=True)
            executive_idb_only = row[0].cnt if row else 0
        else:
            # Executive-only for other statuses
            q = """
                SELECT COUNT(DISTINCT c.name) AS cnt
                FROM `tabCandidate` c
                INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
                WHERE c.candidate_created_by = %s
                AND DATE(cs.sourced_date) = %s
                AND cs.status = %s
            """
            params = [executive, date, status]
            if task_candidate_names:
                q += " AND c.name NOT IN %s"
                params.append(tuple(task_candidate_names))

            row = frappe.db.sql(q, params, as_dict=True)
            executive_only_counts[status] = row[0].cnt if row else 0

    return {
        "counts": counts,                      # project/task-based
        "extra_candidates": extra_candidates,  # statuses outside list
        "project_task_info": project_task_info,
        "executive_only_counts": executive_only_counts,  # other statuses
        "executive_idb_only": executive_idb_only        # separate IDB count
    }

import frappe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import textwrap

def show_value(val):
    return val if val and val != 0 else "-"

# @frappe.whitelist()
# def download_candidate_summary_full():

#     data = frappe.db.sql("""
#         SELECT
#             p.customer,
#             p.name AS project,
#             p.project_name,
#             IFNULL(p.tvac,0) AS vao,
#             IFNULL(p.tsp,0) AS sp,
#             IFNULL(p.tfp,0) AS fp,
#             IFNULL(p.tsl,0) AS sl,
#             IFNULL(p.custom_t_lp,0) AS lp,
#             p.custom_spoc_remark AS remark,
#             c.task,
#             t.subject,
#             c.pending_for,
#             COUNT(c.name) AS count
#         FROM `tabProject` p
#         JOIN `tabCandidate` c ON c.project = p.name
#         JOIN `tabTask` t ON t.name = c.task
#         WHERE p.status = 'Open'
#         AND p.service IN ('REC-I','REC-D')
#         AND c.pending_for != 'IDB'
#         AND t.status NOT IN ('Hold','Completed','Cancelled')
#         GROUP BY p.customer, p.name, c.task, c.pending_for
#         ORDER BY p.customer, p.project_name, t.subject
#     """, as_dict=True)

#     tree = {}
#     project_level_totals = {}
#     counted_projects = set()

#     for row in data:
#         cust = row.customer
#         proj = row.project
#         task = row.task

#         tree.setdefault(cust, {})
#         tree[cust].setdefault(proj, {
#             "project_name": row.project_name,
#             "vao": row.vao,
#             "sp": row.sp,
#             "fp": row.fp,
#             "sl": row.sl,
#             "lp": row.lp,
#             "remark": row.remark or "",
#             "tasks": {}
#         })

#         tree[cust][proj]["tasks"].setdefault(task, {
#             "subject": row.subject,
#             "counts": {}
#         })

#         tree[cust][proj]["tasks"][task]["counts"][row.pending_for] = row.count

#         key = (cust, proj)
#         if key not in counted_projects:
#             project_level_totals.setdefault(cust, {"VAC":0,"SP":0,"FP":0,"SL":0,"LP":0})

#             project_level_totals[cust]["VAC"] += int(row.vao or 0)
#             project_level_totals[cust]["SP"] += int(row.sp or 0)
#             project_level_totals[cust]["FP"] += int(row.fp or 0)
#             project_level_totals[cust]["SL"] += int(row.sl or 0)
#             project_level_totals[cust]["LP"] += int(row.lp or 0)

#             counted_projects.add(key)

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Candidate Summary"

#     header_fill = PatternFill("solid", fgColor="1F4E78")
#     header_font = Font(bold=True, color="FFFFFF")
#     customer_fill = PatternFill("solid", fgColor="85819e")
#     project_fill = PatternFill("solid", fgColor="e7e6ec")
#     row1_fill = PatternFill("solid", fgColor="FFFFFF")
#     row2_fill = PatternFill("solid", fgColor="F2F2F2")

#     center = Alignment(horizontal="center", vertical="center")
#     left = Alignment(horizontal="left", vertical="center")

#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )

#     row_num = 1
#     sno = 1

#     headers = ["S.No","Customer / Project","","","","","","","VAC","SP","FP","SL","LP","SPOC Remarks"]

#     for col, h in enumerate(headers, 1):
#         cell = ws.cell(row=row_num, column=col, value=h)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border

#     ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=8)

#     row_num += 1

#     for cust, projects in tree.items():

#         totals = project_level_totals[cust]

#         ws.cell(row=row_num, column=1, value=sno)

#         ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=8)
#         ws.cell(row=row_num, column=2, value="[-] " + cust)

#         ws.cell(row=row_num, column=9, value=show_value(totals["VAC"]))
#         ws.cell(row=row_num, column=10, value=show_value(totals["SP"]))
#         ws.cell(row=row_num, column=11, value=show_value(totals["FP"]))
#         ws.cell(row=row_num, column=12, value=show_value(totals["SL"]))
#         ws.cell(row=row_num, column=13, value=show_value(totals["LP"]))

#         for col in range(1,15):
#             cell = ws.cell(row=row_num, column=col)
#             cell.fill = customer_fill
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.alignment = left if col == 2 else center
#             cell.border = border

#         row_num += 1
#         sno += 1

#         for proj, pdata in projects.items():

#             ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=8)
#             ws.cell(row=row_num, column=2, value="   " + pdata["project_name"])

#             ws.cell(row=row_num, column=9, value=show_value(pdata["vao"]))
#             ws.cell(row=row_num, column=10, value=show_value(pdata["sp"]))
#             ws.cell(row=row_num, column=11, value=show_value(pdata["fp"]))
#             ws.cell(row=row_num, column=12, value=show_value(pdata["sl"]))
#             ws.cell(row=row_num, column=13, value=show_value(pdata["lp"]))
#             ws.cell(row=row_num, column=14, value=pdata.get("remark", ""))
#             remark = pdata.get("remark", "")
#             remark = "\n".join(textwrap.wrap(remark, 50))

#             cell = ws.cell(row=row_num, column=14, value=remark)
#             cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

#             ws.row_dimensions[row_num].height = max(20, remark.count("\n") * 15)

#             for col in range(1,15):
#                 cell = ws.cell(row=row_num, column=col)
#                 cell.fill = project_fill
#                 cell.border = border

#                 if col == 2:
#                     cell.alignment = left
#                 elif col == 14:
#                     # ✅ keep wrap text
#                     cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
#                 else:
#                     cell.alignment = center

#             row_num += 1

#             task_headers = ["S.No","Position","SRC - IP","PQ - IP","SUB - IP","SUBC - CP","SL - CP",
#                             "LP - IP","LPC - IP","RPT - CP","IVD - CP","RP - CP","PSL", ""]

#             for col, h in enumerate(task_headers,1):
#                 cell = ws.cell(row=row_num, column=col, value=h)
#                 cell.fill = header_fill
#                 cell.font = header_font
#                 cell.alignment = center
#                 cell.border = border

#             row_num += 1

#             mapping = {
#                 3: "Sourced",
#                 4: "Pending QC",
#                 5: "Submit(SPOC)",
#                 6: "Submitted(Client)",
#                 7: "Shortlisted",
#                 8: "Linedup",
#                 9: "Linedup Confirmed",
#                 10: "Reported",
#                 11: "Interviewed",
#                 12: "Result Pending",
#                 13: "Proposed PSL"
#             }

#             task_sno = 1
#             i = 0

#             for task, tdata in pdata["tasks"].items():
#                 i += 1
#                 counts = tdata["counts"]
#                 fill = row1_fill if i % 2 else row2_fill

#                 ws.cell(row=row_num, column=1, value=task_sno)
#                 ws.cell(row=row_num, column=2, value=tdata["subject"])

#                 for col, key in mapping.items():
#                     ws.cell(row=row_num, column=col, value=show_value(counts.get(key, 0)))

#                 for col in range(1,15):
#                     cell = ws.cell(row=row_num, column=col)
#                     cell.fill = fill
#                     cell.alignment = center
#                     cell.border = border

#                 row_num += 1
#                 task_sno += 1

#     widths = [6,40,10,10,10,10,10,10,10,10,10,10,10,80]
#     for i,w in enumerate(widths,1):
#         ws.column_dimensions[get_column_letter(i)].width = w

#     file_path = "/tmp/Candidate_Project_Summary.xlsx"
#     wb.save(file_path)

#     with open(file_path, "rb") as f:
#         frappe.response.filename = "Candidate_Project_Summary.xlsx"
#         frappe.response.filecontent = f.read()
#         frappe.response.type = "download"

@frappe.whitelist()
def download_task_summary_full():

    data = frappe.db.sql("""
        SELECT
            p.customer,
            p.name AS project,
            p.project_name,
            IFNULL(p.tvac,0) AS vao,
            IFNULL(p.tsp,0) AS sp,
            IFNULL(p.tfp,0) AS fp,
            IFNULL(p.tsl,0) AS sl,
            IFNULL(p.custom_t_lp,0) AS lp,
            p.custom_spoc_remark AS remark,
            c.task,
            t.subject,
            c.pending_for,
            COUNT(c.name) AS count
        FROM `tabProject` p
        JOIN `tabCandidate` c ON c.project = p.name
        JOIN `tabTask` t ON t.name = c.task
        WHERE p.status = 'Open'
        AND p.service IN ('REC-I','REC-D')
        AND c.pending_for != 'IDB'
        AND t.status NOT IN ('Hold','Completed','Cancelled')
        GROUP BY p.customer, p.name, c.task, c.pending_for
        ORDER BY p.customer, p.project_name, t.subject
    """, as_dict=True)

    tree = {}
    project_level_totals = {}
    counted_projects = set()

    for row in data:
        cust = row.customer
        proj = row.project
        task = row.task

        tree.setdefault(cust, {})
        tree[cust].setdefault(proj, {
            "project_name": row.project_name,
            "vao": row.vao,
            "sp": row.sp,
            "fp": row.fp,
            "sl": row.sl,
            "lp": row.lp,
            "remark": row.remark or "",
            "tasks": {}
        })

        tree[cust][proj]["tasks"].setdefault(task, {
            "subject": row.subject,
            "counts": {}
        })

        tree[cust][proj]["tasks"][task]["counts"][row.pending_for] = row.count

        key = (cust, proj)
        if key not in counted_projects:
            project_level_totals.setdefault(cust, {"VAC":0,"SP":0,"FP":0,"SL":0,"LP":0})

            project_level_totals[cust]["VAC"] += int(row.vao or 0)
            project_level_totals[cust]["SP"] += int(row.sp or 0)
            project_level_totals[cust]["FP"] += int(row.fp or 0)
            project_level_totals[cust]["SL"] += int(row.sl or 0)
            project_level_totals[cust]["LP"] += int(row.lp or 0)

            counted_projects.add(key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Summary"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    pos_fill = PatternFill("solid", fgColor="2483cc")
    header_font = Font(bold=True, color="FFFFFF")
    customer_fill = PatternFill("solid", fgColor="85819e")
    project_fill = PatternFill("solid", fgColor="e7e6ec")
    row1_fill = PatternFill("solid", fgColor="e6f0ff")
    row2_fill = PatternFill("solid", fgColor="FFFFFF")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_num = 1
    sno = 1

    headers = ["S.No","Customer / Project","","","","","","","VAC","SP","FP","SL","LP","SPOC Remarks"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=8)

    row_num += 1
    task_headers = ["S.No","Position","SRC - IP","PQ - IP","SUB - IP","SUBC - CP","SL - CP",
                    "LP - IP","LPC - IP","RPT - CP","IVD - CP","RP - CP","PSL", ""]

    for col, h in enumerate(task_headers,1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.fill = pos_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    row_num += 1

    for cust, projects in tree.items():

        totals = project_level_totals[cust]

        ws.cell(row=row_num, column=1, value=sno)

        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=8)
        ws.cell(row=row_num, column=2, value="[-] " + cust)

        ws.cell(row=row_num, column=9, value=show_value(totals["VAC"]))
        ws.cell(row=row_num, column=10, value=show_value(totals["SP"]))
        ws.cell(row=row_num, column=11, value=show_value(totals["FP"]))
        ws.cell(row=row_num, column=12, value=show_value(totals["SL"]))
        ws.cell(row=row_num, column=13, value=show_value(totals["LP"]))

        for col in range(1,15):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = customer_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = left if col == 2 else center
            cell.border = border

        row_num += 1
        sno += 1

        for proj, pdata in projects.items():

            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=8)
            ws.cell(row=row_num, column=2, value="   " + pdata["project_name"])

            ws.cell(row=row_num, column=9, value=show_value(pdata["vao"]))
            ws.cell(row=row_num, column=10, value=show_value(pdata["sp"]))
            ws.cell(row=row_num, column=11, value=show_value(pdata["fp"]))
            ws.cell(row=row_num, column=12, value=show_value(pdata["sl"]))
            ws.cell(row=row_num, column=13, value=show_value(pdata["lp"]))
            ws.cell(row=row_num, column=14, value=pdata.get("remark", ""))
            remark = pdata.get("remark", "")
            remark = "\n".join(textwrap.wrap(remark, 50))

            cell = ws.cell(row=row_num, column=14, value=remark)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            ws.row_dimensions[row_num].height = max(20, remark.count("\n") * 15)

            for col in range(1,15):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = project_fill
                cell.border = border

                if col == 2:
                    cell.alignment = left
                elif col == 14:
                    # ✅ keep wrap text
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = center

            row_num += 1


            mapping = {
                3: "Sourced",
                4: "Pending QC",
                5: "Submit(SPOC)",
                6: "Submitted(Client)",
                7: "Shortlisted",
                8: "Linedup",
                9: "Linedup Confirmed",
                10: "Reported",
                11: "Interviewed",
                12: "Result Pending",
                13: "Proposed PSL"
            }

            task_sno = 1
            i = 0

            for task, tdata in pdata["tasks"].items():
                i += 1
                counts = tdata["counts"]
                fill = row1_fill if i % 2 else row2_fill

                ws.cell(row=row_num, column=1, value=task_sno)
                ws.cell(row=row_num, column=2, value=tdata["subject"])

                for col, key in mapping.items():
                    ws.cell(row=row_num, column=col, value=show_value(counts.get(key, 0)))

                for col in range(1,15):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = fill
                    cell.border = border
                    if col == 2:
                        cell.alignment = right
                    else:
                        cell.alignment = center

                row_num += 1
                task_sno += 1

    widths = [6,60,10,10,10,10,10,10,10,10,10,10,10,80]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

    file_path = "/tmp/Candidate_Project_Summary.xlsx"
    wb.save(file_path)

    with open(file_path, "rb") as f:
        frappe.response.filename = "Candidate_Project_Summary.xlsx"
        frappe.response.filecontent = f.read()
        frappe.response.type = "download"



import frappe
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import textwrap
from frappe.utils import now_datetime, get_datetime

#customer crt code , candidate wrong

# @frappe.whitelist()
# def download_candidate_summary_full():

#     data = frappe.db.sql("""
#         SELECT
#             p.customer,
#             p.name AS project,
#             p.project_name,
#             IFNULL(p.tvac,0) AS vao,
#             IFNULL(p.tsp,0) AS sp,
#             IFNULL(p.tfp,0) AS fp,
#             IFNULL(p.tsl,0) AS sl,
#             IFNULL(p.custom_t_lp,0) AS lp,
#             p.custom_spoc_remark AS remark,
#             c.task,
#             t.subject,
#             c.pending_for,        
#             c.name AS candidate_id,
#             c.given_name,
#             c.passport_number,
#             c.position,
#             c.mobile_number,
#             COUNT(c.name) AS count
#         FROM `tabProject` p
#         JOIN `tabCandidate` c ON c.project = p.name
#         JOIN `tabTask` t ON t.name = c.task
#         WHERE p.status = 'Open'
#         AND p.service IN ('REC-I','REC-D')
#         AND c.pending_for IN (
#             'Submit(SPOC)',
#             'Submitted(Client)',
#             'Shortlisted',
#             'Linedup',
#             'Linedup Confirmed',
#             'Reported',
#             'Interviewed',
#             'Result Pending'
#         )
#         GROUP BY p.customer, p.name, c.task, c.pending_for
#         ORDER BY p.customer, p.project_name, t.subject
#     """, as_dict=True)
#     #  2. TREE BUILD (same structure as report 1)
#     tree = {}

#     for row in data:
#         cust = row.customer
#         proj = row.project

#         tree.setdefault(cust, {})
#         tree[cust].setdefault(proj, {
#             "project_name": row.project_name,
#             "tvac": row.tvac or 0,
#             "tsp": row.tsp or 0,
#             "tfp": row.tfp or 0,
#             "tsl": row.tsl or 0,
#             "lp": row.custom_t_lp or 0,
#             "remark": row.custom_spoc_remark or "",
#             "candidates": []
#         })

#         if row.candidate_id:
#             tree[cust][proj]["candidates"].append(row)

#     # 3. EXCEL SETUP
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Candidate Summary"

#     header_fill = PatternFill("solid", fgColor="1F4E78")
#     pos_fill = PatternFill("solid", fgColor="2483cc")
#     customer_fill = PatternFill("solid", fgColor="85819e")
#     project_fill = PatternFill("solid", fgColor="e7e6ec")
#     row1_fill = PatternFill("solid", fgColor="e6f0ff")
#     row2_fill = PatternFill("solid", fgColor="FFFFFF")

#     header_font = Font(bold=True, color="FFFFFF")

#     center = Alignment(horizontal="center", vertical="center")
#     left = Alignment(horizontal="left", vertical="center")

#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )

#     row_num = 1
#     sno = 1

#     #  HEADER
#     headers = [
#         "S.No", "Customer / Project", "",
#         "VAC", "SP", "FP", "SL", "LP",
#         "Remarks"
#     ]

#     for col, h in enumerate(headers, 1):
#         cell = ws.cell(row=row_num, column=col, value=h)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border

#     ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
#     row_num += 1

#     # SUB HEADER
#     sub_headers = [
#         "S.No", "Candidate ID", "Name",
#         "Passport", "Position", "Status",
#         "Age", "Mobile", ""
#     ]

#     for col, h in enumerate(sub_headers, 1):
#         cell = ws.cell(row=row_num, column=col, value=h)
#         cell.fill = pos_fill
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border

#     row_num += 1

#     # MAIN LOOP
#     for cust, projects in tree.items():

#         ws.cell(row=row_num, column=1, value=sno)
#         ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
#         ws.cell(row=row_num, column=2, value="[-] " + cust)

#         for col in range(1, 10):
#             cell = ws.cell(row=row_num, column=col)
#             cell.fill = customer_fill
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.alignment = center
#             cell.border = border

#         row_num += 1
#         sno += 1

#         # PROJECT LOOP
#         for proj_id, pdata in projects.items():

#             ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
#             ws.cell(row=row_num, column=2, value="   " + pdata["project_name"])

#             ws.cell(row=row_num, column=4, value=pdata["tvac"])
#             ws.cell(row=row_num, column=5, value=pdata["tsp"])
#             ws.cell(row=row_num, column=6, value=pdata["tfp"])
#             ws.cell(row=row_num, column=7, value=pdata["tsl"])
#             ws.cell(row=row_num, column=8, value=pdata["lp"])

#             remark = "\n".join(textwrap.wrap(pdata["remark"], 50))
#             ws.cell(row=row_num, column=9, value=remark)

#             for col in range(1, 10):
#                 cell = ws.cell(row=row_num, column=col)
#                 cell.fill = project_fill
#                 cell.border = border
#                 cell.alignment = left if col == 2 else center

#             row_num += 1

#             # CANDIDATES (IMPORTANT FIX)
#             i = 0
#             for c in pdata["candidates"]:

#                 i += 1
#                 fill = row1_fill if i % 2 else row2_fill

#                 ws.cell(row=row_num, column=1, value=i)
#                 ws.cell(row=row_num, column=2, value="      " + (c.candidate_id or ""))
#                 ws.cell(row=row_num, column=3, value=c.given_name)
#                 ws.cell(row=row_num, column=4, value=c.passport_number)
#                 ws.cell(row=row_num, column=5, value=c.position)
#                 ws.cell(row=row_num, column=6, value=c.pending_for)
#                 ws.cell(row=row_num, column=7, value=" ")
#                 ws.cell(row=row_num, column=8, value=c.mobile_number)

#                 for col in range(1, 10):
#                     cell = ws.cell(row=row_num, column=col)
#                     cell.fill = fill
#                     cell.border = border
#                     cell.alignment = center

#                 row_num += 1

#     # COLUMN WIDTH
#     widths = [6, 25, 30, 15, 15, 15, 10, 15, 40]
#     for i, w in enumerate(widths, 1):
#         ws.column_dimensions[get_column_letter(i)].width = w

#     # DOWNLOAD
#     file_path = "/tmp/Candidate_Project_Summary.xlsx"
#     wb.save(file_path)

#     with open(file_path, "rb") as f:
#         frappe.response.filename = "Candidate_Project_Summary.xlsx"
#         frappe.response.filecontent = f.read()
#         frappe.response.type = "download"

# customer wrong , candidate crt code.

# @frappe.whitelist()
# def download_candidate_summary_full():

#     # 🔹 1. GET SAME DATA AS DASHBOARD FUNCTION
#     data = get_candidate_data_candidate_dashboard()
#     candidates = data.get("candidate", [])

#     candidates = [c for c in candidates if c.get("pending_for") != "IDB"]

#     # 🔹 2. GET PROJECTS
#     projects = frappe.db.get_all(
#         "Project",
#         filters={
#             "status": "Open",
#             "service": ["in", ["REC-I", "REC-D"]]
#         },
#         fields=[
#             "name", "customer", "project_name",
#             "tvac", "tsp", "tfp", "tsl", "custom_t_lp",
#             "custom_spoc_remark"
#         ],
#         order_by="customer, project_name"
#     )

#     valid_projects = [p.name for p in projects]

#     # 🔹 Filter candidates only valid projects
#     candidates = [c for c in candidates if c.get("project") in valid_projects]

#     # 🔹 3. GROUP CANDIDATES BY PROJECT
#     candidate_map = {}
#     for c in candidates:
#         candidate_map.setdefault(c.get("project"), []).append(c)

#     # 🔹 4. GROUP PROJECTS BY CUSTOMER
#     tree = {}
#     project_map = {}

#     for p in projects:
#         project_map[p.name] = p
#         tree.setdefault(p.customer, []).append(p)

#     # 🔹 5. EXCEL SETUP
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Candidate Summary"

#     header_fill = PatternFill("solid", fgColor="1F4E78")
#     pos_fill = PatternFill("solid", fgColor="2483cc")
#     customer_fill = PatternFill("solid", fgColor="85819e")
#     project_fill = PatternFill("solid", fgColor="e7e6ec")
#     row1_fill = PatternFill("solid", fgColor="e6f0ff")
#     row2_fill = PatternFill("solid", fgColor="FFFFFF")

#     header_font = Font(bold=True, color="FFFFFF")

#     center = Alignment(horizontal="center", vertical="center")
#     left = Alignment(horizontal="left", vertical="center")
#     right = Alignment(horizontal="right", vertical="center")

#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )

#     row_num = 1
#     sno = 1

#     # 🔹 MAIN HEADER
#     headers = ["S.No", "Customer / Project", "", "VAC", "SP", "FP", "SL", "LP", "Remarks"]

#     for col, h in enumerate(headers, 1):
#         cell = ws.cell(row=row_num, column=col, value=h)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border

#     ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
#     row_num += 1

#     # 🔹 SUB HEADER
#     task_headers = ["S.No", "Candidate ID", "Name", "Passport", "Position", "Status", "Age", "Mobile", ""]

#     for col, h in enumerate(task_headers, 1):
#         cell = ws.cell(row=row_num, column=col, value=h)
#         cell.fill = pos_fill
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border

#     row_num += 1

#     # 🔹 LOOP CUSTOMERS
    
#     for cust, proj_list in tree.items():

#         ws.cell(row=row_num, column=1, value=sno)

#         ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
#         ws.cell(row=row_num, column=2, value="[-] " + cust)

#         for col in range(1, 10):
#             cell = ws.cell(row=row_num, column=col)
#             cell.fill = customer_fill
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.alignment = left if col == 2 else center
#             cell.border = border

#         row_num += 1
#         sno += 1

#         # 🔥 PROJECT LOOP
#         for p in proj_list:

#             proj = project_map.get(p.name)

#             # ===== PROJECT SUMMARY ROW =====
#             ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
#             ws.cell(row=row_num, column=2, value="   " + p.project_name)

#             ws.cell(row=row_num, column=4, value=proj.tvac if proj else 0)
#             ws.cell(row=row_num, column=5, value=proj.tsp if proj else 0)
#             ws.cell(row=row_num, column=6, value=proj.tfp if proj else 0)
#             ws.cell(row=row_num, column=7, value=proj.tsl if proj else 0)
#             ws.cell(row=row_num, column=8, value=proj.custom_t_lp if proj else 0)

#             remark = (proj.custom_spoc_remark if proj else "") or ""
#             remark = "\n".join(textwrap.wrap(remark, 50))
#             ws.cell(row=row_num, column=9, value=remark)

#             for col in range(1, 10):
#                 cell = ws.cell(row=row_num, column=col)
#                 cell.fill = project_fill
#                 cell.border = border
#                 cell.alignment = left if col == 2 else center

#             row_num += 1

#             # ===== CANDIDATES UNDER PROJECT (FIXED ALIGNMENT) =====
#             candidate_list = candidate_map.get(p.name, [])

#             i = 0
#             for c in candidate_list:

#                 i += 1
#                 fill = row1_fill if i % 2 else row2_fill

#                 ws.cell(row=row_num, column=1, value=i)
#                 ws.cell(row=row_num, column=2, value="      " + (c.get("name") or ""))
#                 ws.cell(row=row_num, column=3, value=c.get("given_name"))
#                 ws.cell(row=row_num, column=4, value=c.get("passport_number"))
#                 ws.cell(row=row_num, column=5, value=c.get("position"))
#                 ws.cell(row=row_num, column=6, value=c.get("pending_for"))
#                 ws.cell(row=row_num, column=7, value=c.get("status_age"))
#                 ws.cell(row=row_num, column=8, value=c.get("mobile_number"))

#                 for col in range(1, 10):
#                     cell = ws.cell(row=row_num, column=col)
#                     cell.fill = fill
#                     cell.border = border
#                     cell.alignment = center

#                 row_num += 1

#     # 🔹 COLUMN WIDTH
#     widths = [6, 20, 30, 15, 20, 20, 10, 20, 40]

#     for i, w in enumerate(widths, 1):
#         ws.column_dimensions[get_column_letter(i)].width = w

#     # 🔹 DOWNLOAD
#     file_path = "/tmp/Candidate_Project_Summary.xlsx"
#     wb.save(file_path)

#     with open(file_path, "rb") as f:
#         frappe.response.filename = "Candidate_Project_Summary.xlsx"
#         frappe.response.filecontent = f.read()
#         frappe.response.type = "download"




@frappe.whitelist()
def download_candidate_summary_full():

    #  1. GET DATA FROM DASHBOARD FUNCTION
    data = get_candidate_data_candidate_dashboard()
    candidates = data.get("candidate", [])


    # 🔹 2. GET PROJECTS
    projects = frappe.db.get_all(
        "Project",
        filters={
            "status": "Open",
            "service": ["in", ["REC-I", "REC-D"]]
        },
        fields=[
            "name", "customer", "project_name",
            "tvac", "tsp", "tfp", "tsl", "custom_t_lp",
            "custom_spoc_remark"
        ],
        order_by="customer, project_name"
    )
    candidates = [c for c in candidates if c.get("pending_for") != "IDB"]
    candidate_project_set = set(c.get("project") for c in candidates if c.get("project"))
    projects = [p for p in projects if p.name in candidate_project_set]
    valid_projects = set(p.name for p in projects)
    candidates = [c for c in candidates if c.get("project") in valid_projects]
    project_map = {p.name: p for p in projects}
    
    # candidates = [c for c in candidates if c.get("pending_for") != "IDB"]
    # candidate_project_set = set(c.get("project") for c in candidates if c.get("project"))
    # # projects = [p for p in projects if p.name in candidate_project_set]
    # valid_projects = set(p.name for p in projects)
    # candidates = [c for c in candidates if c.get("project") in valid_projects]
    # project_map = {p.name: p for p in projects}

    tree = {}

    for p in projects:
        tree.setdefault(p.customer, {})
        tree[p.customer].setdefault(p.name, {
            "project": p,
            "candidates": []
        })

    # attach candidates properly
    for c in candidates:
        proj = c.get("project")
        if proj in project_map:
            customer = project_map[proj].customer
            if customer in tree and proj in tree[customer]:
                tree[customer][proj]["candidates"].append(c)

    # 🔹 5. EXCEL SETUP
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Summary"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    pos_fill = PatternFill("solid", fgColor="2483cc")
    customer_fill = PatternFill("solid", fgColor="85819e")
    project_fill = PatternFill("solid", fgColor="e7e6ec")
    row1_fill = PatternFill("solid", fgColor="e6f0ff")
    row2_fill = PatternFill("solid", fgColor="FFFFFF")

    header_font = Font(bold=True, color="FFFFFF")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_num = 1
    sno = 1

    # 🔹 MAIN HEADER
    headers = ["S.No", "Customer / Project", "", "VAC", "SP", "FP", "SL", "LP", "Remarks"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    row_num += 1

    # 🔹 SUB HEADER
    sub_headers = ["S.No", "Candidate ID", "Name", "Passport", "Position", "Status", "Age", "Mobile", ""]

    for col, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.fill = pos_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    row_num += 1

    # 🔹 LOOP CUSTOMERS
    for cust, projects_dict in tree.items():

        ws.cell(row=row_num, column=1, value=sno)
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=2, value="[-] " + cust)

        for col in range(1, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = customer_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = left if col == 2 else center
            cell.border = border

        row_num += 1
        sno += 1

        # 🔹 PROJECT LOOP
        for proj_id, pdata in projects_dict.items():

            proj = pdata["project"]

            # PROJECT ROW
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
            ws.cell(row=row_num, column=2, value="   " + proj.project_name)

            ws.cell(row=row_num, column=4, value=clean_value(proj.tvac))
            ws.cell(row=row_num, column=5, value=clean_value(proj.tsp))
            ws.cell(row=row_num, column=6, value=clean_value(proj.tfp))
            ws.cell(row=row_num, column=7, value=clean_value(proj.tsl))
            ws.cell(row=row_num, column=8, value=clean_value(proj.custom_t_lp))

            remark = proj.custom_spoc_remark or ""
            remark = "\n".join(textwrap.wrap(remark, 50))
            ws.cell(row=row_num, column=9, value=remark)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = project_fill
                cell.border = border
                cell.alignment = left if col == 2 else center

            row_num += 1

            # 🔹 CANDIDATES
            candidate_list = pdata.get("candidates", [])

            i = 0
            for c in candidate_list:

                i += 1
                fill = row1_fill if i % 2 else row2_fill

                ws.cell(row=row_num, column=1, value=i)
                ws.cell(row=row_num, column=2, value="      " + (c.get("name") or ""))
                ws.cell(row=row_num, column=3, value=c.get("given_name"))
                ws.cell(row=row_num, column=4, value=c.get("passport_number"))
                ws.cell(row=row_num, column=5, value=c.get("position"))
                ws.cell(row=row_num, column=6, value=c.get("pending_for"))
                ws.cell(row=row_num, column=7, value=clean_value(c.get("status_age")))
                ws.cell(row=row_num, column=8, value=clean_value(c.get("mobile_number")))

                for col in range(1, 10):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = fill
                    cell.border = border
                    if col in [2,3,5,6,9]:
                        cell.alignment = left   
                    else:
                        cell.alignment = center

                row_num += 1

    # COLUMN WIDTH
    widths = [6, 20, 30, 15, 60, 20, 10, 20, 40]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # DOWNLOAD FILE
    file_path = "/tmp/Candidate_Project_Summary.xlsx"
    wb.save(file_path)

    with open(file_path, "rb") as f:
        frappe.response.filename = "Candidate_Project_Summary.xlsx"
        frappe.response.filecontent = f.read()
        frappe.response.type = "download"


def clean_value(val):
    if val is None or val == "" or val == 0 or val == "0":
        return "-"
    return val
from frappe.utils import now_datetime, get_datetime


@frappe.whitelist()
def get_candidate_data_candidate_dashboard(project=None):

    filters = {}


    if project:
        filters["project"] = project
    else:
        filters["project"] = ["is", "set"]

    filters["pending_for"] = [
        "in",
        [
            "Submit(SPOC)", "Submitted(Client)", "Shortlisted",
            "Linedup", "Linedup Confirmed", "Reported",
            "Interviewed", "Result Pending"
        ]
    ]

    candidates = frappe.db.get_all(
        "Candidate",
        filters=filters,
        fields=[
            "name","given_name","passport_number","project","position",
            "pending_for","mobile_number","whatsapp_number","customer"
        ],
        order_by="modified desc",
        # limit_page_length=10
    )

    if not candidates:
        return {"candidate": []}

    candidate_names = [c.name for c in candidates]

    status_rows = frappe.db.sql("""
        SELECT parent, sourced_date
        FROM `tabCandidate status`
        WHERE parent IN %(names)s
    """, {"names": tuple(candidate_names)}, as_dict=True)

    status_map = {s.parent: s.sourced_date for s in status_rows}

    project_names = list(set([c.project for c in candidates if c.project]))

    project_rows = frappe.db.get_all(
        "Project",
        filters={"name": ["in", project_names]},
        fields=["name","tvac","tsp","tfp","tsl","custom_t_lp"]
    )

    project_map = {p.name: p for p in project_rows}

    from frappe.utils import now_datetime, get_datetime

    for c in candidates:

        sourced_date = status_map.get(c.name)

        if sourced_date:
            age_days = (now_datetime() - get_datetime(sourced_date)).days
            c["status_age"] = age_days
        else:
            c["status_age"] = 0

        # if c.project and c.project in project_map:
        #     c.update(project_map[c.project])
    return {"candidate": candidates}



# @frappe.whitelist()
# def get_dropped_closure_active_so():

#     data = frappe.db.sql("""
#         SELECT 
#             c.name AS closure_id,
#             c.passport_no AS passport_number,
#             c.given_name AS name,
#             c.customer AS client
#         FROM `tabClosure` c
#         INNER JOIN `tabSales Order` so
#             ON so.passport_number = c.passport_no
#         WHERE 
#             c.status = 'Dropped'
#             AND c.so_created = 1
#             AND so.status NOT IN ('Closed', 'Cancelled')
#     """, as_dict=True)

#     return data



@frappe.whitelist()
def get_dropped_closure_active_so():
    direct_data = frappe.db.sql("""
        SELECT DISTINCT
            c.name AS closure_id,
            c.passport_no AS passport_number,
            c.given_name AS name,
            c.customer AS client,
            c.status AS closure_status,
            so.name AS so_name,
            so.custom_closure_status,
            so.status AS so_status
        FROM `tabClosure` c
        INNER JOIN `tabSales Order` so
            ON so.passport_number = c.passport_no
        WHERE 
            c.status = 'Dropped'
            AND c.so_created = 1
            AND so.status NOT IN ('Closed', 'Cancelled', 'Completed')
    """, as_dict=True)


    item_data = frappe.db.sql("""
        SELECT DISTINCT
            c.name AS closure_id,
            c.passport_no AS passport_number,
            c.given_name AS name,
            c.customer AS client,
            c.status AS closure_status,
            so.name AS so_name,
            so.custom_closure_status,
            so.status AS so_status
        FROM `tabClosure` c
        INNER JOIN `tabSales Order Item` soi
            ON soi.item_code IN (c.name, c.passport_no)
        INNER JOIN `tabSales Order` so
            ON so.name = soi.parent
        WHERE 
            c.status = 'Dropped'
            AND c.so_created = 1
            AND so.status NOT IN ('Closed', 'Cancelled', 'Completed')
    """, as_dict=True)

    unique_map = {}

    for row in direct_data + item_data:
        key = row["so_name"]
        unique_map[key] = row  

    result = list(unique_map.values())

    return result



import frappe
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@frappe.whitelist()
def download_dropped_closure_active_so_excel():

    data = get_dropped_closure_active_so()

    wb = Workbook()
    ws = wb.active
    ws.title = "Dropped Closure SO Report"

    # Columns
    columns = [
        "Closure ID",
        "Passport Number",
        "Name",
        "Client",
        "Sales Order",
        "SO Closure Status",
        "SO Status"
    ]

    # Header Style
    header_fill = PatternFill(
        start_color="0F1568",
        end_color="0F1568",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    # Row Colors
    even_fill = PatternFill(
        start_color="E7E6EC",
        end_color="E7E6EC",
        fill_type="solid"
    )

    odd_fill = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    # Border
    thin = Side(style="thin", color="CCCCCC")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # Alignment
    alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Add Header
    for col_num, column_title in enumerate(columns, 1):

        cell = ws.cell(row=1, column=col_num)

        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    # Add Data
    for row_num, row in enumerate(data, 2):

        fill = odd_fill if row_num % 2 != 0 else even_fill

        row_values = [
            row.get("closure_id", ""),
            row.get("passport_number", ""),
            row.get("name", ""),
            row.get("client", ""),
            row.get("so_name", ""),
            row.get("custom_closure_status", ""),
            row.get("so_status", "")
        ]

        for col_num, value in enumerate(row_values, 1):

            cell = ws.cell(row=row_num, column=col_num)

            cell.value = value
            cell.fill = fill
            cell.border = border

            if col_num == 7:
                cell.alignment = center_alignment
            else:
                cell.alignment = alignment

    # Column Width
    column_widths = {
        1: 22,
        2: 22,
        3: 28,
        4: 28,
        5: 24,
        6: 24,
        7: 18
    }

    for col_num, width in column_widths.items():

        ws.column_dimensions[
            get_column_letter(col_num)
        ].width = width

    # Generate File
    output = BytesIO()

    wb.save(output)

    frappe.response["filename"] = "Dropped_Closure_SO_Report.xlsx"

    frappe.response["filecontent"] = output.getvalue()

    frappe.response["type"] = "binary"



# @frappe.whitelist()
# def get_arrived_closure_active_so():

#     data = frappe.db.sql("""
#         SELECT 
#             c.name AS closure_id,
#             c.passport_no AS passport_number,
#             c.given_name AS name,
#             c.customer AS client
#         FROM `tabClosure` c
#         INNER JOIN `tabSales Order` so
#             ON so.passport_number = c.passport_no
#         WHERE 
#             c.status = 'Arrived'
#             AND c.so_created = 1
#             AND so.status NOT IN ('Closed', 'Cancelled')
#     """, as_dict=True)

#     return data



@frappe.whitelist()
def get_arrived_closure_active_so():

    direct_data = frappe.db.sql("""
        SELECT DISTINCT
            c.name AS closure_id,
            c.passport_no AS passport_number,
            c.given_name AS name,
            c.customer AS client,
            c.status AS closure_status,
            so.name AS so_name,
            so.custom_closure_status,
            so.status AS so_status
        FROM `tabClosure` c
        INNER JOIN `tabSales Order` so
            ON so.passport_number = c.passport_no
        WHERE 
            c.status = 'Arrived'
            AND c.so_created = 1
            AND so.status NOT IN ('Closed', 'Cancelled', 'Completed')
    """, as_dict=True)

    item_data = frappe.db.sql("""
        SELECT DISTINCT
            c.name AS closure_id,
            c.passport_no AS passport_number,
            c.given_name AS name,
            c.customer AS client,
            c.status AS closure_status,
            so.name AS so_name,
            so.custom_closure_status,
            so.status AS so_status
        FROM `tabClosure` c
        INNER JOIN `tabSales Order Item` soi
            ON (soi.item_code = c.name OR soi.item_code = c.passport_no)
        INNER JOIN `tabSales Order` so
            ON so.name = soi.parent
        WHERE 
            c.status = 'Arrived'
            AND c.so_created = 1
            AND so.status NOT IN ('Closed', 'Cancelled', 'Completed')
    """, as_dict=True)

    unique_map = {}

    for row in direct_data + item_data:
        key = row["closure_id"]   
        unique_map[key] = row   

    return list(unique_map.values())




@frappe.whitelist()
def get_case_status_report_html():

    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'

    data += '<tr style="background-color: #002060; color: white; position: sticky; top: 0; z-index: 1;">' \
        '<td style="text-align:center; font-weight:bold; color:white;">Customer</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">0-5</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">6-10</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">11-15</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">>15</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Entry Grand Total</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Entry-Insuff</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Execution-Insuff</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Grand Total</td>' \
        '</tr>'

    # Fetch all batches with batch_status not Completed
    batches = frappe.db.get_all(
        "Batch",
        {"batch_status": ("!=", "Completed")},
        ["name", "customer"]
    )

    customer_data = {}

    grand_totals = {
        "0-5": 0,
        "6-10": 0,
        "11-15": 0,
        ">15": 0,
        "Entry-Insuff": 0,
        "Execution-Insuff": 0
    }

    for batch in batches:

        customer = batch.customer

        if customer not in customer_data:
            customer_data[customer] = {
                "0-5": 0,
                "6-10": 0,
                "11-15": 0,
                ">15": 0,
                "Entry-Insuff": 0,
                "Execution-Insuff": 0
            }

        cases = frappe.db.get_all(
            "Case",
            {"batch": batch.name},
            ["name", "case_status", "actual_tat"]
        )

        for case in cases:

            if case.case_status in ["Draft", "Entry Completed", "Entry-QC", "Execution"]:

                if case.actual_tat and 0 <= case.actual_tat <= 5:
                    customer_data[customer]["0-5"] += 1

                elif case.actual_tat and 6 <= case.actual_tat <= 10:
                    customer_data[customer]["6-10"] += 1

                elif case.actual_tat and 11 <= case.actual_tat <= 15:
                    customer_data[customer]["11-15"] += 1

                elif case.actual_tat and case.actual_tat > 15:
                    customer_data[customer][">15"] += 1

            if case.case_status == "Entry-Insuff":
                customer_data[customer]["Entry-Insuff"] += 1

            if case.case_status == "Execution-Insuff":
                customer_data[customer]["Execution-Insuff"] += 1

    # Append rows
    for customer, counts in customer_data.items():

        entry_grand_total = (
            counts["0-5"] +
            counts["6-10"] +
            counts["11-15"] +
            counts[">15"]
        )

        grand_total = (
            counts["Entry-Insuff"] +
            counts["Execution-Insuff"]
        )

        for key in grand_totals:
            grand_totals[key] += counts[key]

        data += f'''
            <tr>
                <td style="text-align:center;">{customer}</td>
                <td style="text-align:center;">{counts["0-5"] or ""}</td>
                <td style="text-align:center;">{counts["6-10"] or ""}</td>
                <td style="text-align:center;">{counts["11-15"] or ""}</td>
                <td style="text-align:center;">{counts[">15"] or ""}</td>
                <td style="text-align:center;">{entry_grand_total or ""}</td>
                <td style="text-align:center;">{counts["Entry-Insuff"] or ""}</td>
                <td style="text-align:center;">{counts["Execution-Insuff"] or ""}</td>
                <td style="text-align:center;">{grand_total or ""}</td>
            </tr>
        '''

    overall_grand_total = (
        grand_totals["Entry-Insuff"] +
        grand_totals["Execution-Insuff"]
    )

    entry_grand_total_sum = (
        grand_totals["0-5"] +
        grand_totals["6-10"] +
        grand_totals["11-15"] +
        grand_totals[">15"]
    )

    data += f'''
        <tr style="font-weight: bold; background-color: #f2f2f2;">
            <td style="text-align:center;">Grand Total</td>
            <td style="text-align:center;">{grand_totals["0-5"] or ""}</td>
            <td style="text-align:center;">{grand_totals["6-10"] or ""}</td>
            <td style="text-align:center;">{grand_totals["11-15"] or ""}</td>
            <td style="text-align:center;">{grand_totals[">15"] or ""}</td>
            <td style="text-align:center;">{entry_grand_total_sum or ""}</td>
            <td style="text-align:center;">{grand_totals["Entry-Insuff"] or ""}</td>
            <td style="text-align:center;">{grand_totals["Execution-Insuff"] or ""}</td>
            <td style="text-align:center;">{overall_grand_total or ""}</td>
        </tr>
    '''

    data += '</table>'

    return data





@frappe.whitelist()
def download_case_status_report_excel():

    import openpyxl
    from io import BytesIO
    from frappe.utils.xlsxutils import make_xlsx
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case Status Report"

    headers = [
        "Customer",
        "0-5",
        "6-10",
        "11-15",
        ">15",
        "Entry Grand Total",
        "Entry-Insuff",
        "Execution-Insuff",
        "Grand Total"
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    batches = frappe.db.get_all(
        "Batch",
        {"batch_status": ("!=", "Completed")},
        ["name", "customer"]
    )

    customer_data = {}

    for batch in batches:

        customer = batch.customer

        if customer not in customer_data:
            customer_data[customer] = {
                "0-5": 0,
                "6-10": 0,
                "11-15": 0,
                ">15": 0,
                "Entry-Insuff": 0,
                "Execution-Insuff": 0
            }

        cases = frappe.db.get_all(
            "Case",
            {"batch": batch.name},
            ["case_status", "actual_tat"]
        )

        for case in cases:

            if case.case_status in ["Draft", "Entry Completed", "Entry-QC", "Execution"]:

                if case.actual_tat and 0 <= case.actual_tat <= 5:
                    customer_data[customer]["0-5"] += 1

                elif case.actual_tat and 6 <= case.actual_tat <= 10:
                    customer_data[customer]["6-10"] += 1

                elif case.actual_tat and 11 <= case.actual_tat <= 15:
                    customer_data[customer]["11-15"] += 1

                elif case.actual_tat and case.actual_tat > 15:
                    customer_data[customer][">15"] += 1

            if case.case_status == "Entry-Insuff":
                customer_data[customer]["Entry-Insuff"] += 1

            if case.case_status == "Execution-Insuff":
                customer_data[customer]["Execution-Insuff"] += 1
    grand_totals = {
        "0-5": 0,
        "6-10": 0,
        "11-15": 0,
        ">15": 0,
        "Entry-Insuff": 0,
        "Execution-Insuff": 0
    }
    row_no = 2

    for customer, counts in customer_data.items():

        entry_grand_total = (
            counts["0-5"] +
            counts["6-10"] +
            counts["11-15"] +
            counts[">15"]
        )

        grand_total = (
            counts["Entry-Insuff"] +
            counts["Execution-Insuff"]
        )

        ws.append([
            customer,
            counts["0-5"],
            counts["6-10"],
            counts["11-15"],
            counts[">15"],
            entry_grand_total,
            counts["Entry-Insuff"],
            counts["Execution-Insuff"],
            grand_total
        ])

        row_fill = "FFFFFF" if row_no % 2 == 0 else "E7E6EC"

        for cell in ws[row_no]:
            cell.fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_no += 1
        for key in grand_totals:
            grand_totals[key] += counts[key]


    overall_grand_total = (
        grand_totals["Entry-Insuff"] +
        grand_totals["Execution-Insuff"]
    )

    entry_grand_total_sum = (
        grand_totals["0-5"] +
        grand_totals["6-10"] +
        grand_totals["11-15"] +
        grand_totals[">15"]
    )

    ws.append([
        "Grand Total",
        grand_totals["0-5"],
        grand_totals["6-10"],
        grand_totals["11-15"],
        grand_totals[">15"],
        entry_grand_total_sum,
        grand_totals["Entry-Insuff"],
        grand_totals["Execution-Insuff"],
        overall_grand_total
    ])

    for cell in ws[row_no]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D9D9D9",
            end_color="D9D9D9",
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response['filename'] = 'Case Status Report.xlsx'
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'

import frappe
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@frappe.whitelist()
def download_arrived_closure_active_so_excel():

    data = get_arrived_closure_active_so()

    wb = Workbook()
    ws = wb.active

    ws.title = "Arrived Closure SO Report"

    # COLUMN HEADERS
    columns = [
        "Closure ID",
        "Passport Number",
        "Name",
        "Client",
        "Sales Order",
        "SO Closure Status",
        "SO Status"
    ]

    # HEADER COLOR
    header_fill = PatternFill(
        start_color="0F1568",
        end_color="0F1568",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    # ROW COLORS
    odd_fill = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    even_fill = PatternFill(
        start_color="E7E6EC",
        end_color="E7E6EC",
        fill_type="solid"
    )

    # BORDER
    thin = Side(style="thin", color="CCCCCC")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # ALIGNMENT
    left_align = Alignment(
        horizontal="left",
        vertical="center"
    )

    center_align = Alignment(
        horizontal="center",
        vertical="center"
    )

    # ADD HEADER
    for col_num, column_title in enumerate(columns, 1):

        cell = ws.cell(row=1, column=col_num)

        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align

    # ADD DATA
    for row_num, row in enumerate(data, 2):

        fill = odd_fill if row_num % 2 != 0 else even_fill

        values = [
            row.get("closure_id", ""),
            row.get("passport_number", ""),
            row.get("name", ""),
            row.get("client", ""),
            row.get("so_name", ""),
            row.get("custom_closure_status", ""),
            row.get("so_status", "")
        ]

        for col_num, value in enumerate(values, 1):

            cell = ws.cell(
                row=row_num,
                column=col_num
            )

            cell.value = value
            cell.fill = fill
            cell.border = border

            if col_num in [1, 2, 5, 7]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # COLUMN WIDTH
    widths = {
        1: 22,
        2: 22,
        3: 28,
        4: 28,
        5: 24,
        6: 24,
        7: 18
    }

    for col_num, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col_num)
        ].width = width

    # GENERATE EXCEL
    output = BytesIO()

    wb.save(output)

    frappe.response["filename"] = "Arrived_Closure_SO_Report.xlsx"

    frappe.response["filecontent"] = output.getvalue()

    frappe.response["type"] = "binary"




@frappe.whitelist()
def get_ptsr_data_closure_wise_all():

    groups = {
        "internal": {
            "status": ["PSL", "Emigration", "Ticket", "Onboarding"],
            "nationality": "Indian",
            "sa_name": None
        },

        "candidate": {
            "status": ["Signed Offer Letter", "Premedical", "PCC", "Final Medical"],
            "nationality": "Indian",
            "sa_name": ["is", "not set"]
        },

        "agent": {
            "status": ["Signed Offer Letter", "Premedical", "PCC", "Final Medical"],
            "nationality": "Indian",
            "sa_name": ["is", "set"]
        },

        "supplier": {
            "status": ["Certificate Attestation", "Biometric", "Trade Test", "Visa Stamping"],
            "nationality": "Indian",
            "sa_name": None
        },

        "client": {
            "status": ["Client Offer Letter", "Visa"],
            "nationality": "Indian",
            "sa_name": None
        },

        "nepal": {
            "status": [
                "PSL", "Emigration", "Ticket", "Onboarding",
                "Signed Offer Letter", "Premedical", "PCC", "Final Medical",
                "Certificate Attestation", "Biometric", "Trade Test", "Visa Stamping"
            ],
            "nationality": "Nepali",
            "sa_name": None
        },

        "srilanka": {
            "status": [
                "PSL", "Emigration", "Ticket", "Onboarding",
                "Signed Offer Letter", "Premedical", "PCC", "Final Medical",
                "Certificate Attestation", "Biometric", "Trade Test", "Visa Stamping"
            ],
            "nationality": "Srilankan",
            "sa_name": None
        }
    }

    result = {}
    project_cache = {}

    def process_rows(rows):

        for c in rows:

            c["custom_history"] = frappe.get_all(
                "Closure Status History",
                filters={
                    "parent": c["name"],
                    "parenttype": "Closure",
                    "parentfield": "custom_history"
                },
                fields=["date"]
            )

            project = c.get("project")

            if project and project not in project_cache:

                project_cache[project] = frappe.db.get_value(
                    "Project",
                    project,
                    [
                        "name",
                        "project_name",
                        "tvac",
                        "tsp",
                        "tfp",
                        "tsl",
                        "custom_t_lp",
                        "custom_spoc_remark"
                    ],
                    as_dict=True
                )

            p = project_cache.get(project) or {}

            c["project_id"] = p.get("name") or project
            c["project_name"] = p.get("project_name") or project

            c["tvac"] = p.get("tvac") or 0
            c["tsp"] = p.get("tsp") or 0
            c["tfp"] = p.get("tfp") or 0
            c["tsl"] = p.get("tsl") or 0
            c["custom_t_lp"] = p.get("custom_t_lp") or 0
            c["custom_spoc_remark"] = p.get("custom_spoc_remark") or "-"

            c["task_subject"] = frappe.db.get_value(
                "Task",
                c.get("task"),
                "subject"
            ) or "No Position"

        return rows

    for key, cfg in groups.items():

        filters = {
            "status": ["in", cfg["status"]],
            "nationality": cfg["nationality"]
        }

        if cfg.get("sa_name"):
            filters["sa_name"] = cfg["sa_name"]

        data = frappe.db.get_all(
            "Closure",
            filters=filters,
            fields=[
                "name","given_name","passport_no","customer",
                "territory","status","remark","last_updated_on",
                "sa_name","sa_mobile_number","associate","mobile",
                "std_remarks","custom_next_follow_up_on",
                "task","project"
            ],
            order_by="last_updated_on desc"
        )

        result[key] = process_rows(data)

    return result




@frappe.whitelist()
def download_task_summary():

    import frappe
    import openpyxl

    from io import BytesIO
    from collections import defaultdict

    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )

    # =====================================================
    # GET DATA
    # =====================================================

    all_data = get_ptsr_data_closure_wise_all()

    data = []

    for key, value in all_data.items():

        data.extend(value)


    grouped = defaultdict(list)

    for d in data:

        customer = d.get("customer") or "Unknown"

        grouped[customer].append(d)


    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Task Summary"


    dark_blue = PatternFill(
        start_color="0F1568",
        end_color="0F1568",
        fill_type="solid"
    )

    light_blue = PatternFill(
        start_color="12A9E3",
        end_color="12A9E3",
        fill_type="solid"
    )

    customer_fill = PatternFill(
        start_color="85819E",
        end_color="85819E",
        fill_type="solid"
    )

    odd_fill = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    even_fill = PatternFill(
        start_color="E7E6EC",
        end_color="E7E6EC",
        fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    thin = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # =====================================================
    # MERGE
    # =====================================================

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:N1")

    ws.merge_cells("O1:O1")
    ws.merge_cells("P1:P1")
    ws.merge_cells("Q1:Q1")
    ws.merge_cells("R1:R1")
    ws.merge_cells("S1:S1")

    # =====================================================
    # HEADER VALUES
    # =====================================================

    ws["A1"] = "S.No"
    ws["B1"] = "Customer"

    ws["O1"] = "VAC"
    ws["P1"] = "SP"
    ws["Q1"] = "FP"
    ws["R1"] = "SL"
    ws["S1"] = "LP"

    headers = [
        "Project",
        "Position",
        "PSL",
        "COL",
        "SOL",
        "VISA",
        "PM",
        "PCC",
        "CA",
        "FM",
        "BIO",
        "QVP",
        "TT",
        "VS",
        "POE",
        "TKT",
        "OB",
        "OD"
    ]

    col = 2

    for h in headers:

        ws.cell(
            row=2,
            column=col,
            value=h
        )

        col += 1

    # =====================================================
    # HEADER STYLE
    # =====================================================

    for row in ws.iter_rows(
        min_row=1,
        max_row=2,
        min_col=1,
        max_col=19
    ):

        for cell in row:

            cell.font = white_font
            cell.alignment = center
            cell.border = border

            if cell.row == 1:
                cell.fill = dark_blue
            else:
                cell.fill = light_blue

    # =====================================================
    # WIDTH
    # =====================================================

    widths = {
        "A": 8,
        "B": 35,
        "C": 28,
        "D": 12,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 10,
        "L": 10,
        "M": 10,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10,
        "R": 10,
        "S": 10
    }

    for col_name, width in widths.items():

        ws.column_dimensions[col_name].width = width

    # =====================================================
    # ROW HEIGHT
    # =====================================================

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # =====================================================
    # START ROW
    # =====================================================

    current_row = 3

    customer_index = 1

    # =====================================================
    # CUSTOMER LOOP
    # =====================================================

    for customer in sorted(grouped.keys()):

        rows = grouped[customer]

        # =================================================
        # TOTALS
        # =================================================

        totals = {
            "tvac": 0,
            "tsp": 0,
            "tfp": 0,
            "tsl": 0,
            "tlp": 0
        }

        project_map = {}

        for r in rows:

            pid = (
                r.get("project_id")
                or r.get("project_name")
            )

            if pid not in project_map:

                project_map[pid] = r

        for p in project_map.values():

            totals["tvac"] += int(p.get("tvac") or 0)
            totals["tsp"] += int(p.get("tsp") or 0)
            totals["tfp"] += int(p.get("tfp") or 0)
            totals["tsl"] += int(p.get("tsl") or 0)
            totals["tlp"] += int(p.get("custom_t_lp") or 0)

        # =================================================
        # CUSTOMER ROW
        # =================================================

        # customer_values = [

        #     customer_index,
        #     customer,
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     "",
        #     totals["tvac"],
        #     totals["tsp"],
        #     totals["tfp"],
        #     totals["tsl"],
        #     totals["tlp"]
        # ]

        # =========================================
        # MERGE CUSTOMER NAME
        # =========================================

        ws.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=14
        )

        # S.NO
        cell = ws.cell(
            row=current_row,
            column=1,
            value=customer_index
        )

        cell.fill = customer_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = center

        # CUSTOMER NAME
        cell = ws.cell(
            row=current_row,
            column=2,
            value=customer
        )

        cell.fill = customer_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = left

        # APPLY STYLE FOR MERGED AREA
        for col_num in range(2, 15):

            c = ws.cell(
                row=current_row,
                column=col_num
            )

            c.fill = customer_fill
            c.border = border

        # VAC TO LP
        summary_values = [
            totals["tvac"],
            totals["tsp"],
            totals["tfp"],
            totals["tsl"],
            totals["tlp"]
        ]

        start_col = 15

        for value in summary_values:

            cell = ws.cell(
                row=current_row,
                column=start_col,
                value=value
            )

            cell.fill = customer_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = center

            start_col += 1


        current_row += 1

        customer_index += 1

        # =================================================
        # TASK GROUP
        # =================================================

        task_grouped = {}

        for r in rows:

            key = (
                str(r.get("project_name") or "")
                + "##" +
                str(r.get("task_subject") or "")
            )

            if key not in task_grouped:

                task_grouped[key] = []

            task_grouped[key].append(r)

        task_index = 1

        # =================================================
        # TASK ROWS
        # =================================================

        for idx, task_rows in enumerate(
            task_grouped.values(),
            start=1
        ):

            first = task_rows[0]

            def get_count(status_name):

                return len([
                    x for x in task_rows
                    if (
                        x.get("status") or ""
                    ).strip() == status_name
                ])

            fill = odd_fill if idx % 2 else even_fill

            row_data = [

                task_index,
                first.get("project_name") or "-",
                first.get("task_subject") or "-",

                get_count("PSL"),
                get_count("Client Offer Letter"),
                get_count("Signed Offer Letter"),
                get_count("Visa"),
                get_count("Premedical"),
                get_count("PCC"),
                get_count("Certificate Attestation"),
                get_count("Final Medical"),
                get_count("Biometric"),
                get_count("QVP"),
                get_count("Trade Test"),
                get_count("Visa Stamping"),
                get_count("Emigration"),
                get_count("Ticket"),
                get_count("Onboarding"),
                get_count("Onboarded")
            ]

            for col_num, value in enumerate(
                row_data,
                start=1
            ):

                display_value = value

                if (
                    col_num >= 4
                    and col_num <= 19
                    and (
                        value == 0
                        or value == ""
                        or value is None
                    )
                ):

                    display_value = "-"

                cell = ws.cell(
                    row=current_row,
                    column=col_num,
                    value=display_value
                )

                cell.fill = fill
                cell.border = border
                cell.alignment = center

                if col_num in [2, 3]:

                    cell.alignment = left

            current_row += 1

            task_index += 1

    # =====================================================
    # FREEZE
    # =====================================================

    ws.freeze_panes = "A3"

    # =====================================================
    # DOWNLOAD
    # =====================================================

    output = BytesIO()

    wb.save(output)

    frappe.response["filename"] = "Task_Summary.xlsx"

    frappe.response["filecontent"] = output.getvalue()

    frappe.response["type"] = "binary"




@frappe.whitelist()
def download_candidate_summary():

    import frappe
    import openpyxl

    from io import BytesIO
    from collections import defaultdict

    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )

    # =====================================================
    # GET DATA
    # =====================================================

    all_data = get_ptsr_data_closure_wise_all()

    data = []

    for key, value in all_data.items():

        data.extend(value)

    # =====================================================
    # GROUP CUSTOMER
    # =====================================================

    grouped = defaultdict(list)

    for d in data:

        customer = d.get("customer") or "Unknown"

        grouped[customer].append(d)

    # =====================================================
    # WORKBOOK
    # =====================================================

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Candidate Summary"

    # =====================================================
    # STYLES
    # =====================================================

    dark_blue = PatternFill(
        start_color="0F1568",
        end_color="0F1568",
        fill_type="solid"
    )

    light_blue = PatternFill(
        start_color="12A9E3",
        end_color="12A9E3",
        fill_type="solid"
    )

    customer_fill = PatternFill(
        start_color="85819E",
        end_color="85819E",
        fill_type="solid"
    )

    odd_fill = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    even_fill = PatternFill(
        start_color="E7E6EC",
        end_color="E7E6EC",
        fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    thin = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # =====================================================
    # MERGE
    # =====================================================

    ws.merge_cells("A1:F1")

    # =====================================================
    # TOP HEADER
    # =====================================================

    ws["A1"] = "Customer"

    # =====================================================
    # SECOND HEADER
    # =====================================================

    headers = [

        "S.No",
        "Project",
        "CLID",
        "Name",
        "PP Number",
        "Contact",
        "Status",
        "Age",
        "Last Update",
        "Next Action",
        "Next Action On",
        "Remark"
    ]

    col = 1

    for h in headers:

        ws.cell(
            row=2,
            column=col,
            value=h
        )

        col += 1

    # =====================================================
    # HEADER STYLE
    # =====================================================

    for row in ws.iter_rows(
        min_row=1,
        max_row=2,
        min_col=1,
        max_col=12
    ):

        for cell in row:

            cell.font = white_font
            cell.alignment = center
            cell.border = border

            if cell.row == 1:
                cell.fill = dark_blue
            else:
                cell.fill = light_blue

    # =====================================================
    # WIDTH
    # =====================================================

    widths = {

        "A": 8,
        "B": 28,
        "C": 18,
        "D": 28,
        "E": 20,
        "F": 18,
        "G": 18,
        "H": 12,
        "I": 20,
        "J": 25,
        "K": 20,
        "L": 35
    }

    for col_name, width in widths.items():

        ws.column_dimensions[col_name].width = width

    # =====================================================
    # ROW HEIGHT
    # =====================================================

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # =====================================================
    # START ROW
    # =====================================================

    current_row = 3

    customer_index = 1

    # =====================================================
    # CUSTOMER LOOP
    # =====================================================

    for customer in sorted(grouped.keys()):

        rows = grouped[customer]

        # =================================================
        # CUSTOMER ROW
        # =================================================

        ws.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=7
        )

        # S.NO
        cell = ws.cell(
            row=current_row,
            column=1,
            value=customer_index
        )

        cell.fill = customer_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = center

        # CUSTOMER NAME
        cell = ws.cell(
            row=current_row,
            column=2,
            value=customer
        )

        cell.fill = customer_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = left

        # MERGED STYLE
        for col_num in range(2, 8):

            c = ws.cell(
                row=current_row,
                column=col_num
            )

            c.fill = customer_fill
            c.border = border

        # REMAINING EMPTY CELLS
        for col_num in range(8, 13):

            c = ws.cell(
                row=current_row,
                column=col_num
            )

            c.fill = customer_fill
            c.border = border

        current_row += 1

        customer_index += 1

        # =================================================
        # CANDIDATE ROWS
        # =================================================

        sub_index = 1

        for idx, row in enumerate(rows, start=1):

            fill = odd_fill if idx % 2 else even_fill

            row_data = [

                sub_index,
                row.get("project_name") or "-",
                row.get("name") or "-",
                row.get("given_name") or "-",
                row.get("passport_no") or "-",
                row.get("mobile") or "-",
                row.get("status") or "-",
                "-",
                row.get("last_updated_on") or "-",
                row.get("std_remarks") or "-",
                row.get("custom_next_follow_up_on") or "-",
                row.get("remark") or "-"
            ]

            for col_num, value in enumerate(
                row_data,
                start=1
            ):

                cell = ws.cell(
                    row=current_row,
                    column=col_num,
                    value=value
                )

                cell.fill = fill
                cell.border = border
                cell.alignment = center

                if col_num in [2, 3, 4, 10, 12]:

                    cell.alignment = left

            current_row += 1

            sub_index += 1

    # =====================================================
    # FREEZE
    # =====================================================

    ws.freeze_panes = "A3"

    # =====================================================
    # DOWNLOAD
    # =====================================================

    output = BytesIO()

    wb.save(output)

    frappe.response["filename"] = "Candidate_Summary.xlsx"

    frappe.response["filecontent"] = output.getvalue()

    frappe.response["type"] = "binary"


