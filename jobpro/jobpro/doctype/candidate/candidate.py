# -*- coding: utf-8 -*-
# Copyright (c) 2020, teamPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe.utils import today, add_years,nowdate
import json
from datetime import date,datetime
import qrcode
import base64
from PIL import Image
from io import BytesIO
from erpnext.setup.utils import get_exchange_rate
from frappe.utils.data import date_diff, now_datetime, nowdate, today, add_days
from frappe import throw, msgprint
from frappe import throw, _


class Candidate(Document):
    def before_insert(self):
        if self.task:
            task = frappe.get_doc('Task', self.task)
            if task.get("custom_criteria_table"):
                for row in task.custom_criteria_table:
                    self.append("custom_criteria", {
                        "scheduling_parameter": row.scheduling_parameter,
                        "scheduling_criteria": row.scheduling_criteria
                    })
        self.append('custom_status_transition',{
            'status':self.pending_for,
            'sourced_date':now_datetime(),
            'sourced_by':self.candidate_created_by,
            'project':self.project,
            'task':self.task,
            'position':self.position
        })
        if not self.candidate_created_by:
            self.candidate_created_by=frappe.session.user
        
    def after_insert(self):
        
        candidate_name = self.name
        
        
        
        frappe.db.after_commit(lambda: frappe.enqueue(
        method="jobpro.jobpro.doctype.candidate.candidate.create_user_for_candidate",
        queue='short',
        candidate_name=candidate_name
        ))


        # frappe.db.after_commit(lambda: create_user_for_candidate(candidate_name))
        
        # avail_mail = None
        # avail_mob = None
        
        # if self.mail_id or self.mobile_number:
        
        #     if self.mail_id:
                
        #         avail_mail = frappe.db.sql(
        #         "SELECT name FROM `tabUser` WHERE email = %s ORDER BY creation",
        #         (self.mail_id,),
        #         as_dict=True
        #     )
                
        #     if self.mobile_number:
        #         avail_mob = frappe.db.sql(
        #             "SELECT name FROM `tabUser` WHERE mobile_no = %s ORDER BY creation",
        #             (self.mobile_number,),
        #             as_dict=True
        #         )
            
        #     if avail_mob:
        #         return
                
        #     elif avail_mail:
        #         return
             
        #     else:
        #         if self.mail_id and self.mail_id.strip():
        #             user = frappe.new_doc("User")
        #             user.first_name = self.given_name
        #             user.email = self.mail_id
        #             user.mobile_no = self.mobile_number
        #             user.role_profile_name = "JOBPRO"
        #             user.insert(ignore_permissions=True)
                    
        #             reset_link = f"http://139.5.190.19:8080/frontend/reset-password?user={self.mail_id}"
        #             message = f"""
        #             <p style="color: #05264e; font-weight: 700; font-size: 15px;">Dear {self.given_name},</p>
        #             <p>Please click the link below to Set password for your account:</p>
        #             <button style="border: 0px solid black; border-radius: 8px; color: white; padding: 3px; width: 140px; background-color: #0070cc;"><a href="{reset_link}" target="_blank" style="color: white; text-align: center; text-decoration: none;">Reset Password</a></button>
        #             <p>If you did not request to set a password, please ignore this email.</p>
        #             <p style="color: #0070cc;">Best regards,<br>Jobpro Team</p>
        #             """
        #             frappe.sendmail(
        #                 recipients=[self.mail_id],
        #                 subject="Set Your Password",
        #                 message=message,
        #                 reference_doctype="User",
        #                 reference_name=self.given_name,
        #                 delayed=False,
        #             )    
            

    def validate(self):
        if self.task:
            frappe.db.set_value("JOBPRO Candidate", {"parent": self.task}, "status", self.pending_for)
            
        
        client_si = self.custom_client_si or 0
        candidate_si = self.custom_candidate_si or 0
        associate_si = self.custom_associate_si or 0

        if client_si > 0 and self.custom_payment_from:
            if self.custom_client_billing_currency == "INR":
                self.custom_client_payment_company_currency = client_si
            else:
                conversion = get_exchange_rate(self.custom_client_billing_currency, "INR")
                self.custom_client_payment_company_currency = conversion * client_si

        if candidate_si > 0 and self.custom_payment_from:
            if self.custom_billing_currency == "INR":
                self.custom_candidate_payment_company_currency = candidate_si
            else:
                conversion = get_exchange_rate(self.custom_billing_currency, "INR")
                self.custom_candidate_payment_company_currency = conversion * candidate_si

        if associate_si > 0 and self.custom_payment_from:
            if self.custom_associate_paymentinr == "INR":
                self.custom_associate_payment_company_currency = associate_si
            else:
                conversion = get_exchange_rate(self.custom_associate_paymentinr, "INR")
                self.custom_associate_payment_company_currency = conversion * associate_si
        input_str = 'QR data'
        qr = qrcode.make(input_str)
        temp = BytesIO()
        qr.save(temp, "PNG")
        temp.seek(0)
        b64 = base64.b64encode(temp.read())
        qr_img =  "<img src='data:image/png;base64,{0}'/>".format(b64.decode("utf-8"))
        self.qr = "<h1>Test</h1>"

        # if self.territory == "India":
        #     # frappe.set_value('Candidate',"service","REC-D")
        #     self.service = 'REC-D'
        # else:
        #     # frappe.set_value('Candidate',"service","REC-I")
        #     self.service = 'REC-I'
    #     if self.dob and datetime.strptime(self.dob,'%Y-%m-%d') > date.today():
    #         frappe.throw("Date Of Birth can't be Future Date")
    #     if self.issued_date and datetime.strptime(self.issued_date,'%Y-%m-%d') > nowdate():
    #         frappe.throw("Issued Date can't be Future Date")
    #     if self.expected_doj and atetime.strptime(self.expected_doj,'%Y-%m-%d') < nowdate():
    #         frappe.throw("Expected Date of Joining can't be Past Date")

    # def validate_date(self):        
    #     if self.issued_date and datetime.strptime(self.issued_date,'%Y-%m-%d') > nowdate():
    #         return "Issued Date can't be Future Date"
            
    # def validate_dob(self):
    #     if self.dob and self.dob > nowdate():
    #         return "DOB can't be Future Date"

@frappe.whitelist()
def update_criteria_table(task_id,name):
    project = frappe.get_doc("Task", task_id)
    if not project.custom_criteria_table:
        return
    task = frappe.get_doc("Candidate", name)
    for row in project.custom_criteria_table:
        task.append("custom_criteria", {
            
            "scheduling_criteria": row.scheduling_criteria,  
            "scheduling_parameter": row.scheduling_parameter         
        })
    task.save()

@frappe.whitelist()
def Specialization(q_data):
    category = frappe.get_value("Qualification", {"name": q_data},['category'])
    # specialization = frappe.get_all("Specialization", {"category": category},['name'])
    return category
@frappe.whitelist()
def check_territory(territory):
    site = frappe.db.get_value("Territory",{"territory_name":territory},["parent_territory"])
    return site
# @frappe.whitelist()
# def restrict_candidate():
#     if frappe.db.exists('Candidate',name):



@frappe.whitelist()
def create_closure(doc,method):
    if doc.pending_for == 'Proposed PSL':
        
        closure_id = frappe.db.exists("Closure", {"candidate": doc.name})
        if closure_id:
            closure = frappe.get_doc("Closure", closure_id)
        else:    
            closure = frappe.new_doc("Closure")
        closure.update({
            "candidate": doc.name,
            "given_name": doc.given_name,
            "territory":doc.territory,
            "mobile": doc.mobile_number,
            "mail_id": doc.mail_id,
            "basic": doc.basic,
            "food_allowance": doc.food,
            "other_allowance": doc.other_allowances,
            "customer": doc.customer,
            "nationality": doc.nationality,
            "task": doc.task,
            "project": doc.project,
            "candidate_owner":doc.candidate_created_by,
            "sa_id":doc.sa_agent,
            "sa_name":doc.sa_agent_name,
            "sams_name":doc.sa_agent_name,
            "passport_no": doc.passport_number,
            "passport_number": doc.passport_number,
            "date_of_birth":doc.dob,
            "ecr_status":doc.ecr_status,
            "issued_date":doc.issued_date,
            "expiry_date":doc.passport_expiry_date,
            "expected_doj":doc.expected_doj or '',
            "place_of_issue":doc.place_of_issue,
            "selection_date":doc.interview_date,
            "interview_location":doc.interview_location,
            "service":doc.service,
            # "vaccination_certificate":doc.certificate_attach,
            # "vaccination_status":doc.vaccination_status
            })
        if doc.irf:
            closure.update({"irf": doc.irf})
        if doc.candidate_image:
            closure.update({"photo": doc.candidate_image})
        if doc.passport:
            closure.update({"passport": doc.passport})
        if doc.offer_letter:
            closure.update({"offer_letter": doc.offer_letter})
        closure.flags.ignore_mandatory = True
        closure.save(ignore_permissions=True)   
        frappe.db.commit()

def validate(self):
    if self.passport_number:
        existing = frappe.db.exists('Candidate', {'passport_number': self.passport_number, 'name': ['!=', self.name]})
        if existing:
            frappe.throw(('Passport Number must be unique'), frappe.DuplicateEntryError)

import frappe

@frappe.whitelist()
def send_interview_email(candidate_email, candidate_name, candidate_created_by, name):
   
    subject = f"Interview Invitation for {candidate_name}"
    
    message = f"""
    <p><b>Dear {candidate_name},</p>
    <p>You have been invited for an interview. Kindly find the attached documents for your reference.</p><b><br><br><br>
    <p><strong>Best Regards,<br>TEAMPRO HR & IT Services Pvt. Ltd.</strong></p>
    """
    
    frappe.sendmail(
        sender=candidate_created_by,
        recipients=[candidate_email],
        cc=[candidate_created_by,"sangeetha.a@groupteampro.com"],
        subject=subject,
        message=message,
        now=True,
        attachments=[frappe.attach_print("Candidate", name,
            file_name="Interview Call Letter", print_format="Interview Call Letter")]
    )
    
@frappe.whitelist()
def create_closure_with_payment(doc,method):
    if doc.pending_for == 'Proposed PSL':
        
        closure_id = frappe.db.exists("Closure", {"candidate": doc.name})
        if closure_id:
            closure = frappe.get_doc("Closure", closure_id)
        else:    
            closure = frappe.new_doc("Closure")
        closure.update({
            "candidate": doc.name,
            "given_name": doc.given_name,
            "territory":doc.territory,
            "mobile": doc.mobile_number,
            "mail_id": doc.mail_id,
            "basic": doc.basic,
            "food_allowance": doc.food,
            "other_allowance": doc.other_allowances,
            "customer": doc.customer,
            "nationality": doc.nationality,
            "task": doc.task,
            "project": doc.project,
            "candidate_owner":doc.candidate_created_by,
            "sa_id":doc.sa_agent,
            "sa_name":doc.sa_agent_name,
            "sams_name":doc.sa_agent_name,
            "passport_no": doc.passport_number,
            "passport_number": doc.passport_number,
            "date_of_birth":doc.dob,
            "ecr_status":doc.ecr_status,
            "issued_date":doc.issued_date,
            "expiry_date":doc.passport_expiry_date,
            "expected_doj":doc.expected_doj or '',
            "place_of_issue":doc.place_of_issue,
            "selection_date":doc.interview_date,
            "interview_location":doc.interview_location,
            "service":doc.service,
            })
        if doc.irf:
            closure.update({"irf": doc.irf})
        if doc.candidate_image:
            closure.update({"photo": doc.candidate_image})
        if doc.passport:
            closure.update({"passport": doc.passport})
        if doc.offer_letter:
            closure.update({"offer_letter": doc.offer_letter})
        if not doc.custom_payment_from:
           
            frappe.throw("Kindly enter the payment from before move the status to PSL")
        closure.payment=doc.custom_payment_from
        if doc.custom_payment_from=="Candidate":
            if not doc.custom_candidate_si:
               
                frappe.throw("Kindly enter the Candidate SI in Payment Details")
                    
            if not doc.custom_billing_currency:
                
                frappe.throw("Kindly enter the Candidate Billing Currency in Payment Details")
            closure.candidate_si=doc.custom_candidate_si
            closure.candidate_payment_company_currenc=doc.custom_candidate_payment_company_currency
            closure.candidate_service_charge=doc.custom_candidate_si
            closure.billing_currency=doc.custom_billing_currency
        if doc.custom_payment_from=="Client":
            if not doc.custom_client_si:
                
                frappe.throw("Kindly enter the Client SI in Payment Details")
            if not doc.custom_client_billing_currency:
                frappe.throw("Kindly enter the Client Billing Currency in Payment Details")
            closure.client_si=doc.custom_client_si
            closure.client_payment_company_currency=doc.custom_client_payment_company_currency
            closure.custom_client_billing_currency=doc.custom_client_billing_currency
        if doc.custom_payment_from=="Associate":
            if not doc.custom_associate_si:
                
                frappe.throw("Kindly enter the Associate SI in Payment Details")
            if not doc.custom_associate_paymentinr:
                
                frappe.throw("Kindly enter the Associate Billing Currency in Payment Details")
            closure.associate_si=doc.custom_associate_si
            closure.associate_service_charge=doc.custom_associate_si
            closure.custom_associate_billing_currency=doc.custom_associate_paymentinr
        if doc.custom_payment_from=="Both":
            if not doc.custom_client_si:
                frappe.throw("Kindly enter the Client SI in Payment Details")
            if not doc.custom_candidate_si:
                
                frappe.throw("Kindly enter the Candidate SI in Payment Details")
            if not doc.custom_billing_currency:
                
                frappe.throw("Kindly enter the Candidate Billing Currency in Payment Details")
            if not doc.custom_client_billing_currency:
               
                frappe.throw("Kindly enter the Client Billing Currency in Payment Details")
            closure.client_si=doc.custom_client_si
            closure.candidate_si=doc.custom_candidate_si
            closure.billing_currency=doc.custom_billing_currency
            closure.custom_client_billing_currency=doc.custom_client_billing_currency
        
        closure.flags.ignore_mandatory = True
        closure.save(ignore_permissions=True)   
        frappe.db.commit()

@frappe.whitelist()
def update_sa_details_task(task=None, sa_agent=None):
    if sa_agent and task:
        if frappe.db.exists("Task", task):
            task_doc = frappe.get_doc("Task", task)
            found = False
            if task_doc.sa_detail:
                for i in task_doc.sa_detail:
                    if i.sa_id == sa_agent:
                        i.received_count += 1
                        found = True
                        break  
            if not found:
                task_doc.append("sa_detail", {
                    "sa_id": sa_agent,
                    "received_count": 1
                })
            task_doc.save(ignore_permissions=True)
            frappe.db.commit()
            return "Updated successfully"



import requests
from urllib.parse import urljoin

@frappe.whitelist()
def create_user_for_candidate(candidate_name):
    try:
        candidate = frappe.get_doc("Candidate", candidate_name)

        params = {
            'email': candidate.mail_id,
            'first_name': candidate.given_name,
            'mobile_no': candidate.mobile_number
        }

        url = "https://teamprohr.com/api/method/jobpro.jobpro_web.user_creation_from_candidate"
        headers = {
            'Content-Type': 'application/json',
            'Authorization': 'token 7ccea99fadbdcc7:6ac2db28fa141a2'
        }

        response = requests.post(url, headers=headers, json=params, verify=True)
        response.raise_for_status()

        res = response.json()
        return res

    except requests.exceptions.RequestException as e:
        frappe.throw(f"HTTP error: {str(e)}")
        return
    except json.JSONDecodeError:
        frappe.throw("Failed to decode JSON response from server")
        return
    except Exception as e:
        frappe.throw(f"Unexpected error: {str(e)}")
        return
    

@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_specialization(doctype, txt, searchfield, start, page_len, filters):
    degree = filters.get("degree")
    if not degree:
        return []

    return frappe.db.sql("""
        SELECT DISTINCT app.specialization
        FROM `tabQualification` sp
        INNER JOIN `tabCandidate Specialization` app
            ON app.parent = sp.name
        WHERE sp.name = %s
          AND app.specialization LIKE %s
        ORDER BY app.specialization
        LIMIT %s, %s
    """, (degree, "%" + txt + "%", start, page_len))

import frappe

@frappe.whitelist()
def bulk_update_candidate_status(doc_names, status, idb_remarks=None, any_other_reason=None):
    doc_names = frappe.parse_json(doc_names)

    for name in doc_names:
        doc = frappe.get_doc("Candidate", name)
        frappe.errprint(doc.name)
        # doc.pending_for = status

        # doc.append("custom_status_transition", {
        #     "status": status,
        #     "sourced_date": frappe.utils.now_datetime(),
        #     "sourced_by": frappe.session.user,
        #     "project": doc.project,
        #     "task": doc.task,
        # })

        # if status == "IDB":
        #     doc.custom_idbremarks = idb_remarks
        #     if idb_remarks == "Any other":
        #         doc.custom_any_other_reason = any_other_reason

        # doc.save(ignore_permissions=True)

    frappe.db.commit()
    return "OK"


from frappe.utils.background_jobs import enqueue
import frappe

@frappe.whitelist()
def enqueue_candidate_owner_update():
    # Enqueue the job to update candidates in background
    enqueue(
        method=update_candidate_owner, 
        queue="long", 
        timeout=86400  
    )

def update_candidate_owner():
    # Fetch candidates where pending_for = "IDB"
    candidates = frappe.get_all(
        "Candidate", 
        filters={"pending_for": "IDB"}, 
        fields=["name", "candidate_created_by"]
    )
    
    for candidate in candidates:
        # Update owner field
        frappe.db.set_value("Candidate", candidate.name, "candidate_created_by", "cv@groupteampro.com")

import os
import fitz  # PyMuPDF
from io import BytesIO
from PyPDF2 import PdfReader, PdfMerger
from docx import Document  # For Word (.docx)
import pandas as pd  # For Excel (.xlsx)
import pypdf
from PIL import Image
from io import BytesIO
import fitz  # PyMuPDF
from PyPDF2 import PdfReader, PdfMerger
import pypdf


import os
from PIL import Image
import fitz  # PyMuPDF
from io import BytesIO
from PyPDF2 import PdfReader, PdfMerger
import frappe
import pypdf
from frappe.utils.pdf import get_pdf



@frappe.whitelist()
def merge_and_download_pdf_masked_new(docname):
    doc = frappe.get_doc("Candidate", docname)
    merger = PdfMerger()

    try:
        
        print_format_pdf = get_pdf(frappe.get_print("Candidate", docname, "Interview form Masked"))

        
        print_pdf_file = BytesIO(print_format_pdf)
        print_pdf_file.seek(0)

        
        reader1 = PdfReader(print_pdf_file, strict=False)
        if len(reader1.pages) == 0:
            frappe.throw(_("No pages found in the print format PDF!"))
        
        
        pages1 = remove_blank_pages_new(reader1)

        
        print_pdf_path = f"/tmp/clean_print_{docname}.pdf"
        with open(print_pdf_path, "wb") as temp_pdf:
            writer = PdfWriter()
            for page in pages1:
                writer.add_page(page)
            writer.write(temp_pdf)

        
        merger.append(print_pdf_path)

    except Exception as e:
       
        frappe.log_error(message=str(e), title="Error in Print Format")
        frappe.throw(_("Error generating print format: {0}").format(str(e)))

    
    for idx, attachment in enumerate(doc.custom_attachments, start=1):
        pdf_path = None
        if attachment.attachment_name !="Un Masked CV" and attachment.file:
            
            
            file_url = attachment.file
            file_path = download_external_file_new(file_url)

            
            if is_pdf_new(file_path):
                pdf_path = file_path  
            elif is_word_new(file_path):
                pdf_path = convert_word_to_pdf_new(file_path)
            elif is_excel_new(file_path):
                pdf_path = convert_excel_to_pdf_new(file_path)
            elif is_csv_new(file_path):
                pdf_path = convert_csv_to_pdf_with_setup_new(file_path)    
            elif is_image_new(file_path):
                pdf_path = convert_image_to_pdf_new(file_path)
            else:
                frappe.throw(_("Unsupported file format: {0}").format(file_path))

        if pdf_path is not None:
        
            merger.append(pdf_path)

    
    merged_pdf_path = f"/tmp/merged_{docname}.pdf"
    with open(merged_pdf_path, "wb") as output_file:
        merger.write(output_file)
    merger.close()

    
    compressed_pdf_path = f"/tmp/compressed_{docname}.pdf"
    compress_pdf_with_fitz_new(merged_pdf_path, compressed_pdf_path)

    
    logo_file_url = "https://erp.teamproit.com/file/32a83e99e6/cfde950a7aTeampro -logo.png"
    logo_path = download_external_file_new(logo_file_url)
    final_pdf_path = f"/tmp/final_with_logo_{docname}.pdf"
    add_logo_to_pdf_with_fitz_new(compressed_pdf_path, final_pdf_path, logo_path)

    
    with open(final_pdf_path, "rb") as file:
        frappe.local.response.filename = f"Merged_{docname}.pdf"
        frappe.local.response.filecontent = file.read()
        frappe.local.response.type = "download"

@frappe.whitelist()
def merge_and_download_pdf_unmasked_new(docname):
    doc = frappe.get_doc("Candidate", docname)
    merger = PdfMerger()

    try:
        
        print_format_pdf = get_pdf(frappe.get_print("Candidate", docname, "Interview Application Candidate"))

        
        print_pdf_file = BytesIO(print_format_pdf)
        print_pdf_file.seek(0)

        
        reader1 = PdfReader(print_pdf_file, strict=False)
        if len(reader1.pages) == 0:
            frappe.throw(_("No pages found in the print format PDF!"))
        
        
        pages1 = remove_blank_pages_new(reader1)

        
        print_pdf_path = f"/tmp/clean_print_{docname}.pdf"
        with open(print_pdf_path, "wb") as temp_pdf:
            writer = PdfWriter()
            for page in pages1:
                writer.add_page(page)
            writer.write(temp_pdf)

        
        merger.append(print_pdf_path)

    except Exception as e:
       
        frappe.log_error(message=str(e), title="Error in Print Format")
        frappe.throw(_("Error generating print format: {0}").format(str(e)))

    
    for idx, attachment in enumerate(doc.custom_attachments, start=1):
        pdf_path = None
        if attachment.attachment_name !="Masked CV" and attachment.file:
            
            
            file_url = attachment.file
            file_path = download_external_file_new(file_url)

            
            if is_pdf_new(file_path):
                pdf_path = file_path  
            elif is_word_new(file_path):
                pdf_path = convert_word_to_pdf_new(file_path)
            elif is_excel_new(file_path):
                pdf_path = convert_excel_to_pdf_new(file_path)
            elif is_csv_new(file_path):
                pdf_path = convert_csv_to_pdf_with_setup_new(file_path)    
            elif is_image_new(file_path):
                pdf_path = convert_image_to_pdf_new(file_path)
            else:
                frappe.throw(_("Unsupported file format: {0}").format(file_path))

        if pdf_path is not None:
        
            merger.append(pdf_path)

    
    merged_pdf_path = f"/tmp/merged_{docname}.pdf"
    with open(merged_pdf_path, "wb") as output_file:
        merger.write(output_file)
    merger.close()

    
    compressed_pdf_path = f"/tmp/compressed_{docname}.pdf"
    compress_pdf_with_fitz_new(merged_pdf_path, compressed_pdf_path)

    
    logo_file_url = "https://erp.teamproit.com/file/32a83e99e6/cfde950a7aTeampro -logo.png"
    logo_path = download_external_file_new(logo_file_url)
    final_pdf_path = f"/tmp/final_with_logo_{docname}.pdf"
    add_logo_to_pdf_with_fitz_new(compressed_pdf_path, final_pdf_path, logo_path)

    
    with open(final_pdf_path, "rb") as file:
        frappe.local.response.filename = f"Merged_{docname}.pdf"
        frappe.local.response.filecontent = file.read()
        frappe.local.response.type = "download"


from PyPDF2 import PdfReader, PdfWriter

def convert_pdf_to_temp_new(file_path, idx, name):
    """Clean PDF and save a temporary valid PDF using PyPDF2 only"""
    reader = PdfReader(file_path)
    writer = PdfWriter()

    for page in reader.pages:
        writer.add_page(page)

    temp_pdf_path = f"/tmp/{name}_{idx}.pdf"
    with open(temp_pdf_path, "wb") as f:
        writer.write(f)

    return temp_pdf_path




def compress_pdf_with_fitz_new(input_path, output_path, quality=60):
    """Compress PDF using PyMuPDF."""
    doc = fitz.open(input_path)
    doc.save(output_path, garbage=4, deflate=True, clean=True)
    doc.close()

def add_logo_to_pdf_with_fitz_new(input_pdf_path, output_pdf_path, logo_path):
    doc = fitz.open(input_pdf_path)
    logo = fitz.open(logo_path)

    logo_width = 100
    logo_height = 70  
    top_margin = 1    

    for page in doc:
        page_width = page.rect.width
        logo_rect = fitz.Rect(
            page_width - logo_width - 30,
            top_margin,
            page_width - 30,
            top_margin + logo_height
        )
        page.insert_image(logo_rect, filename=logo_path)

    doc.save(output_pdf_path)
    doc.close()

def is_image_new(file_path):
    """Check if a file is an image (JPG, PNG, etc.)."""
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
    return any(file_path.lower().endswith(ext) for ext in image_extensions)


def convert_image_to_pdf_new(image_path):
    """Convert an image (JPG, PNG) to a proper PDF."""
    img = Image.open(image_path)
    pdf_path = f"/tmp/{os.path.basename(image_path)}.pdf"

    
    img = img.convert("RGB")

    
    img.save(pdf_path, "PDF", quality=75)
    return pdf_path



def remove_blank_pages_new(pdf_reader):
    valid_pages = []
    for page in pdf_reader.pages:
        try:
            text = page.extract_text()
            if not text or text.isspace():
                
                valid_pages.append(page)
            else:
                valid_pages.append(page)
        except:
            valid_pages.append(page)
    return valid_pages



import os
import requests
import frappe
from urllib.parse import urlparse, unquote
import uuid

# def download_external_file_new(file_url):
#     """Download file from external storage (e.g., S3, local storage) and return a local path."""
    
    
#     if not file_url.startswith("http"):
#         file_url = frappe.utils.get_url() + file_url

    
#     response = requests.get(file_url, stream=True)
#     if response.status_code != 200:
#         frappe.throw("Failed to download file from external storage.")

    
#     parsed_url = urlparse(file_url)
#     filename = unquote(os.path.basename(parsed_url.path))  
#     filename = filename.replace(" ", "_")  
#     filename = f"{uuid.uuid4().hex}_{filename}"  

#     # Save to /tmp
#     file_path = os.path.join("/tmp", filename)
#     with open(file_path, "wb") as f:
#         f.write(response.content)

#     return file_path

def download_external_file_new(file_url):
    """Download file from external storage (e.g., S3, local storage) and return a local path."""

    try:
        if not file_url.startswith("http"):
            file_url = frappe.utils.get_url() + file_url

        frappe.logger().info(f"Downloading File URL: {file_url}")

        response = requests.get(
            file_url,
            stream=True,
            timeout=30,
            verify=False
        )

        frappe.logger().info(f"Response Status: {response.status_code}")

        if response.status_code != 200:
            frappe.throw(
                f"Failed to download file. Status Code: {response.status_code} URL: {file_url}"
            )

        parsed_url = urlparse(file_url)
        filename = unquote(os.path.basename(parsed_url.path))

        if not filename:
            filename = "temp_file"

        filename = filename.replace(" ", "_")
        filename = f"{uuid.uuid4().hex}_{filename}"

        file_path = os.path.join("/tmp", filename)

        with open(file_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        if not os.path.exists(file_path):
            frappe.throw("File was not saved properly.")

        return file_path

    except Exception as e:
        frappe.log_error(
            title="External File Download Error",
            message=frappe.get_traceback()
        )
        frappe.throw(f"Error downloading file: {str(e)}")

def is_pdf_new(file_path):
    """Check if a file is a valid PDF."""
    try:
        with open(file_path, "rb") as f:
            PdfReader(f, strict=False)
        return True
    except Exception:
        return False


def is_word_new(file_path):
    """Check if a file is a valid Word document (.docx)."""
    # return file_path.lower().endswith('.docx')
    return file_path.lower().endswith((".docx", ".doc"))

def is_excel_new(file_path):
    """Check if a file is a valid Excel document (.xlsx)."""
    return file_path.lower().endswith('.xlsx')

def is_csv_new(file_path):
    return file_path.lower().endswith('.csv')


import subprocess
import os

def convert_word_to_pdf_new(word_file_path):
    base, ext = os.path.splitext(word_file_path)
    pdf_file_path = base + ".pdf"

    subprocess.run(['unoconv', '-f', 'pdf', word_file_path], check=True)

    return pdf_file_path

from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins, PrintPageSetup

def set_excel_page_setup_new(excel_file_path):
    wb = load_workbook(excel_file_path)
    for ws in wb.worksheets:
       
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        
        
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  
        
        
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

    wb.save(excel_file_path)


import subprocess
import os

def convert_excel_to_pdf_new(excel_file_path):
    set_excel_page_setup_new(excel_file_path)
    pdf_file_path = excel_file_path.replace(".xlsx", ".pdf")

    subprocess.run(
        ["unoconv", "-f", "pdf", "-o", pdf_file_path, excel_file_path],
        check=True
    )

    if not os.path.exists(pdf_file_path):
        raise Exception("PDF conversion failed")

    return pdf_file_path


import pandas as pd

def csv_to_excel_new(csv_path, excel_path):
    df = pd.read_csv(csv_path)
    df.to_excel(excel_path, index=False)

def convert_csv_to_pdf_with_setup_new(csv_file_path):
    excel_file_path = csv_file_path.replace(".csv", ".xlsx")
    pdf_file_path = csv_file_path.replace(".csv", ".pdf")

    
    csv_to_excel_new(csv_file_path, excel_file_path)

    
    set_excel_page_setup_new(excel_file_path)

    
    subprocess.run(
        ["unoconv", "-f", "pdf", "-o", pdf_file_path, excel_file_path],
        check=True
    )

    if not os.path.exists(pdf_file_path):
        raise Exception("CSV to PDF conversion failed")

    return pdf_file_path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
@frappe.whitelist()
def test_download():
    filename = "Candidate Details"
    build_xlsx_response_test(filename)

def build_xlsx_response_test(filename):
    xlsx_file = make_xlsx_test(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
 

def make_xlsx_test(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    # Set column widths
    # for col in range(ord('A'), ord('N')):  # Columns A to M
    #     ws.column_dimensions[chr(col)].width = 20
    
    column_widths = {
        
        "A": 5,  
        "B": 15,  
        "C": 15,  
        "D": 40,  
        "E": 15,  
        "F": 15,  
        "G": 10,  
        "H": 10,  
        "I":10 , 
        "J": 36,  
        "K": 18,  
        "L":18 ,  
        "M": 18,
        "N":15,
        "O":70   
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


    # Define headers
    # headers = [
    # "S.No", "CD ID", "PP Number", "Candidate Name", "Qualification",
    # "Specialization", "Experience", "", "", "Current Employer",
    # "Salary (SAR)", "", "Current Location", "Notice Period", "Remarks"
    # ]

    sub_headers = [
        "", "", "", "", "", "",
        "India", "Gulf", "Total",
        "",
        "Current", "Expected",
        "", "", ""
    ]


    # Define styles
    position_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    position_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="00b1f0", end_color="00b1f0", fill_type="solid")
    header_font = Font(bold=True)
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    even_row_fill = PatternFill(
    start_color="e6f2f1",
    end_color="e6f2f1",
    fill_type="solid"
    )
    # Fetch candidate data grouped by positions
    position_candidates = get_data_grouped_by_position(args)
    
    
        

    for position, position_data in position_candidates.items():
        
        candidates = position_data["rows"]
        currency = position_data["currency"]

        headers = [
            "S.No", "CD ID", "PP Number", "Candidate Name", "Qualification",
            "Specialization", "Experience", "", "",
            "Current Employer",
            f"Salary ({currency})", "",
            "Current Location", "Notice Period", "Remarks"
        ]
        
        # Add position row
        position_row = ws.max_row + 1
        ws.merge_cells(start_row=position_row, start_column=1, end_row=position_row, end_column=15)
        cell = ws.cell(row=position_row, column=1)
        cell.value = f"{position}"
        cell.fill = position_fill
        cell.font = position_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

        # Add headers
        # header_row = ws.max_row + 1
        # for col_num, header in enumerate(headers, start=1):
        #     cell = ws.cell(row=header_row, column=col_num)
        #     cell.value = header
        #     cell.fill = header_fill
        #     cell.font = header_font
        #     cell.alignment = Alignment(horizontal="center", vertical="center")
        #     cell.border = black_border
        
        
        header_row_1 = ws.max_row + 1
        header_row_2 = header_row_1 + 1

        # First header row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row_1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = black_border

        # Second header row (sub headers)
        for col_num, header in enumerate(sub_headers, start=1):
            cell = ws.cell(row=header_row_2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = black_border
            
        rowspan_columns = [1,2,3,4,5,6,10,13,14,15]
        for col in rowspan_columns:
            ws.merge_cells(
                start_row=header_row_1,
                start_column=col,
                end_row=header_row_2,
                end_column=col
            )
    
        ws.merge_cells(
        start_row=header_row_1,
        start_column=7,
        end_row=header_row_1,
        end_column=9
        )
        
        ws.merge_cells(
        start_row=header_row_1,
        start_column=11,
        end_row=header_row_1,
        end_column=12
        )

        row_num = ws.max_row + 1

        
        
        

        # Add details for the position
        for idx, candidate in enumerate(candidates, start=1):
            # ws.append(candidate)
            row_num = ws.max_row + 1
            is_even_row = idx % 2 == 0
            
            cell = ws.cell(row=row_num, column=1)
            cell.value = idx
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = black_border
            if is_even_row:
                cell.fill = even_row_fill
            
            
            
            for col_num, value in enumerate(candidate, start=2):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = black_border
                
                if col_num in[1, 2,3,  5, 6,7,8,9,10,14]:
                    
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_num in[11,12]:
                    
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                if col_num == 15:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")    
                    
                if is_even_row:
                    cell.fill = even_row_fill    
                    

        # Add an empty row for separation
        ws.append([])

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_grouped_by_position(args):
    data = {}
    candidates = frappe.get_all(
        "Candidate",
        filters={'submitted_date': nowdate(), 'candidate_created_by': args.candidate_created_by},
        fields=["name", "passport_number", "given_name", "highest_degree","specialization","india_experience",
                "overseas_experience", "total_experience",  "current_employer",
                "current_ctc", "expected_ctc", "location", "notice_period_months",
                "remarks_1", "position","currency_ctc",]
    )
    
    for candidate in candidates:
        position = candidate.get("position", "")
        currency=candidate.currency_ctc
        formatted_ctc = f"{currency} {candidate.current_ctc}" if candidate.current_ctc else "0"
        formatted_expected_ctc = f"{currency} {candidate.expected_ctc}" if candidate.expected_ctc else "0"
        if position not in data:
            data[position] = {
                "currency": currency,
                "rows": []
            }
        data[position]["rows"].append([
            candidate.name, candidate.passport_number, candidate.given_name,
            candidate.highest_degree, candidate.specialization, candidate.india_experience, candidate.overseas_experience, 
            candidate.total_experience, candidate.current_employer,
            formatted_ctc,formatted_expected_ctc , candidate.location, 
            candidate.notice_period_months, candidate.remarks_1, 
        ])

    return data

@frappe.whitelist()
def download_excel():
    filename = "Candidate Details"
    build_xlsx_response_new(filename)

def build_xlsx_response_new(filename):
    xlsx_file = make_xlsx_new(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_new(data, sheet_name="Candidates", wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    
    if wb is None:
        wb = Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    # Set column widths
    for col in range(ord('A'), ord('M')):  # Columns A to L
        ws.column_dimensions[chr(col)].width = 20

    # Define headers
    headers = ["Candidate ID", "PP Number", "Candidate Name", "Qualification", 
               "Total Yrs of Exp", "Overseas Exp", "Current Employer", 
               "Current Salary", "Exp. Salary", "Current Location", 
               "Notice Period", "Remarks"]

    # Define styles
    position_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    position_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="98D7F5", end_color="98D7F5", fill_type="solid")
    header_font = Font(bold=True)
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    # Fetch candidate data grouped by positions
    position_candidates = get_data_new(args)

    for position, candidates in position_candidates.items():
        # Add position row
        position_row = ws.max_row + 1
        ws.merge_cells(start_row=position_row, start_column=1, end_row=position_row, end_column=12)
        cell = ws.cell(row=position_row, column=1)
        cell.value = f"{position}"
        cell.fill = position_fill
        cell.font = position_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

        # Add headers
        header_row = ws.max_row + 1
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = black_border

        # Add details for the position
        for candidate in candidates:
            # ws.append(candidate)
            row_num = ws.max_row + 1
            for col_num, value in enumerate(candidate, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = black_border  # Apply border to each cell

        # Add an empty row for separation
        ws.append([])

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

import json
import frappe

def get_data_new(args):
    if args is None:
        args = {}

    if isinstance(args.get('args'), str):
        try:
            args['filters'] = json.loads(args['args'])
        except json.JSONDecodeError as e:
            frappe.log_error(title='JSON Decode Error', message=str(e))
            return []
    
    args['filters'] = args.get('filters', {})

    data = []
    date_filter = args['filters'].get('custom_date_filter')
    candidate_created_by_filter = args['filters'].get('custom_candidate_status_filter')
    candidate_status = args['filters'].get('custom_status_filter')
    # print(candidate_created_by_filter)
    # print(date_filter)
    if not date_filter or not candidate_created_by_filter:
        frappe.log_error(title='Missing filters', message='Date filter or candidate_created_by filter is missing.')
        return data

    filters = {}

    candidate_condition = candidate_created_by_filter.get('condition')
    candidate_value = candidate_created_by_filter.get('value')

    if candidate_condition in ('!=', '=',):
        filters['candidate_created_by'] = [candidate_condition, candidate_value]
    elif candidate_condition in ('like', 'not like') and isinstance(candidate_value, str):
        filters['candidate_created_by'] = [candidate_condition, candidate_value]
    elif candidate_condition == 'is' and candidate_value == 'set':
        filters['candidate_created_by'] = ["is", "set"]
    elif candidate_condition == 'is' and candidate_value == 'not set':
        filters['candidate_created_by'] = ["is", "not set"]
    elif candidate_condition in ('in', 'not in') and isinstance(candidate_value, list):
        filters['candidate_created_by'] = [candidate_condition, candidate_value]
    else:
        return data

    date_condition = date_filter.get('condition')
    date_value = date_filter.get('value')

    if date_condition == 'Between' and isinstance(date_value, list) and len(date_value) == 2:
        filters['submitted_date'] = ['between', [date_value[0], date_value[1]]]
    elif date_condition == 'in' and isinstance(date_value, list):
        filters['submitted_date'] = ['in', date_value]
    elif date_condition == 'not in' and isinstance(date_value, list):
        filters['submitted_date'] = ['not in', date_value]
    elif date_condition == 'is' and date_value == 'set':
        filters['submitted_date'] = ['is', 'set'] 
    elif date_condition == 'is' and date_value == 'not set':
        filters['submitted_date'] = ['is', 'not set'] 
    elif date_condition in ('<', '<=', '>', '>=') and isinstance(date_value, str):
        filters['submitted_date'] = [date_condition, date_value]
    elif date_condition in ('=', '!=') and isinstance(date_value, str):
        filters['submitted_date'] = [date_condition, date_value]  
    elif date_condition == 'Timespan' and isinstance(date_value, str):
        # print(date_value)
        from_date, to_date = get_timespan_custom(date_value)
        filters['submitted_date'] = ['between', [from_date, to_date]]
    elif date_condition == 'fiscal year' and isinstance(date_value, str):
        fiscal_year_start, fiscal_year_end = get_fiscal_year_custom(date_value)
        filters['submitted_date'] = ['between', [fiscal_year_start, fiscal_year_end]]
    else:
        # frappe.log_error(title='Invalid Date Filter', message='Date filter is not set properly.')
        return data
    
    if candidate_status:
        status_condition = candidate_status.get('condition')
        status_value = candidate_status.get('value')
        if status_condition and status_value:
            if status_value =="Submit(SPOC)" or status_value == "Submitted(Client)":
                filters['pending_for'] = [status_condition, status_value]
    
    data = {}
    candidates = frappe.get_all(
        "Candidate",
        filters=filters,
        fields=["name", "passport_number", "given_name", "highest_degree",
                "total_experience", "overseas_experience", "current_employer",
                "current_ctc", "expected_ctc", "location", "notice_period_months",
                "remarks_1", "position","currency_ctc"]
    )
    for candidate in candidates:
        position = candidate.get("position", "")
        currency=candidate.currency_ctc
        formatted_ctc = f"{currency} {candidate.current_ctc}" if candidate.current_ctc else "0"
        if position not in data:
            data[position] = []
        data[position].append([
            candidate.name, candidate.passport_number, candidate.given_name,
            candidate.highest_degree, candidate.total_experience, 
            candidate.overseas_experience, candidate.current_employer,
            formatted_ctc, candidate.expected_ctc, candidate.location, 
            candidate.notice_period_months, candidate.remarks_1
        ])

    return data


def get_timespan_custom(timespan):
    print(nowdate())
    today = nowdate()
    if timespan == "last week":
        start_date = add_days(today, -7)
        end_date = today
    elif timespan == "last month":
        start_date = add_months(today, -1)
        end_date = today
    elif timespan == "last quarter":
        start_date = add_months(today, -3)
        end_date = today
    elif timespan == "last year":
        start_date = add_months(today, -12)
        end_date = today
    elif timespan == "last 6 months":
        start_date = add_months(today, -6)
        end_date = today
    elif timespan == "today":
        start_date = end_date = today
    elif timespan == "yesterday":
        start_date = end_date = add_days(today, -1)
    elif timespan == "tomorrow":
        start_date = end_date = add_days(today, 1)
    elif timespan == "next month":
        start_date = add_months(today, 1)
        end_date = add_days(add_months(today, 1), -1)
    elif timespan == "next week":
        start_date = today
        end_date = add_days(today, 7)
    elif timespan == "next quarter":
        start_date = today
        end_date = add_months(today, 3)
    elif timespan == "next year":
        start_date = today
        end_date = add_months(today, 12)
    elif timespan == "next 6 months":
        start_date = today
        end_date = add_months(today, 6)
    elif timespan == "this week":
        start_date = get_first_day(today, "week")  
        end_date = add_days(start_date, 6) 
    elif timespan == "this month":
        start_date = get_first_day(today, "month")  
        end_date = get_last_day(today, "month")  
    elif timespan == "last month":
        start_date = add_months(today, -1)
        end_date = today
    elif timespan == "this quarter":
        start_date = get_first_day(today, "quarter")  
        end_date = get_last_day(today, "quarter")  
    elif timespan == "this year":
        start_date = get_first_day(today, "year")  
        end_date = get_last_day(today, "year")
    else:
        raise ValueError(f"Unsupported timespan: {timespan}")
    
    return start_date, end_date

def get_fiscal_year_custom(fiscal_year):
    fiscal_year_split = fiscal_year.split('-')
    start_year = fiscal_year_split[0]
    end_year = fiscal_year_split[1]

    start_date = date(int(start_year), 1, 1) 
    end_date = date(int(end_year), 12, 31)    

    return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
