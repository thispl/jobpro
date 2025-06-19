import frappe
import json
from frappe.utils.csvutils import read_csv_content
from six.moves import range
from six import string_types
import frappe
import json
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,datetime,get_first_day,get_last_day,today,fmt_money)
from datetime import datetime
from calendar import monthrange
from frappe import _, msgprint
from frappe.utils import flt
from frappe.utils import cstr, cint, getdate
from datetime import date
import requests


def bulk_update_from_csv(filename):
    #below is the method to get file from Frappe File manager
    from frappe.utils.file_manager import get_file
    #Method to fetch file using get_doc and stored as _file
    _file = frappe.get_doc("File", {"file_name": filename})
    #Path in the system
    filepath = get_file(filename)
    #CSV Content stored as pps

    pps = read_csv_content(filepath[1])
    count = 0
    for pp in pps:
        ld = frappe.db.exists("Lead",{'name':pp[0]})
        if ld:
            # items = frappe.get_all("Lead",{'name':pp[0]})
            # for item in items:
            i = frappe.get_doc('Lead',pp[0])
            if not i.contac%t_by:
                i.contact_date = pp[1]
                print(pp[1])
                # i.append("supplier_items",{
                #     'supplier' : pp[1]
                # })
                # i.save(ignore_permissions=True)
                # frappe.db.commit()


def update_timesheet():
    t_list = frappe.get_all("Timesheet",{"status":"Submitted"})
    for t in t_list:
        print(t)
        doc = frappe.get_doc("Timesheet",t.name)
        for d in doc.time_logs:
            if d.task:
                status = frappe.db.get_value("Task",d.task,"status")
                ewh = frappe.db.get_value("Task",d.task,"expected_time")
                print(status)
                print(ewh)
                doc.update({
                    "task_status":status,
                    "task_ewh":ewh
                })
                doc.save(ignore_permissions=True)
                frappe.db.commit()
                


@frappe.whitelist()
def update_candidate_list(candidate,project,customer,task):
    can = json.loads(candidate)
    for c in can:
        cand = frappe.get_doc("Candidate",(c["candidate_id"]))
        cand.update({
            "pending_for": c["candidate_status"],
            "degree" : c.get("degree"),
            # "specialization" : c.get("specialization"),
            # "current_ctc" :c.get("current_ctc"),
            # "current_ctc" :c.get("current_ctc"),
            "indian_experience" : c.get("indian_experience"),
            "gulf_experience" : c.get("gulf_experience"),
            "currency_type" : c.get("currency_type"),
            "expected_ctc" : c.get("expected_ctc"),
            "passport_no" : c.get("passport_no"),
            "expiry_date" : c.get("expiry_date"),
            "ecr_status" : c.get("ecr_status"),
            "current_location" : c.get("current_location"),
            "mobile" : c.get("mobile"),
            "associate_name" : c.get("associate"),
            "user" : c.get("user"),
        })
        # child = frappe.get_all('Candidate Task',{'parent':c["candidate_id"]},['customer','project','task','pending_for','territory'])
        # for ch in child:
        #     if ch.customer == customer and ch.project == project:
        #         ch.update({
        #             "pending_for":c["candidate_status"],
        #             })
        #         ch.db_update()
        cand.db_update()
        frappe.db.commit()

@frappe.whitelist()
def update_candidates(candidate):
    can = json.loads(candidate)
    for c in can:
        cand = frappe.get_doc("Candidate",(c["candidate_id"]))
        cand.update({
            "pending_for": c["candidate_status"],
            "current_location" : c.get("current_location"),
            "address_line_1" : c.get("address_line_1"),
            "mobile" : c.get("mobile"),
            "sa_agent_name" : c.get("sa_name"),
            "user" : c.get("user"),
        })
        cand.db_update()
        frappe.db.commit()

@frappe.whitelist()
def load_candidates(task):
    candidates = frappe.get_all("Candidate", {"task": task}, ["*"], order_by="given_name asc")
    # candidates = frappe.db.sql("""select `tabCandidate`.name as candidate_id,`tabCandidate`.pending_for as candidate_status,`tabCandidate`.given_name as given_name,
    # `tabCandidate`.mobile as mobile,`tabCandidate`.sa_name as sa_name,`tabCandidate`.candidate_created_by as candidate_created_by,`tabCandidate Task`.task
    # FROM `tabCandidate`
    # LEFT JOIN `tabCandidate Task` ON `tabCandidate`.name = `tabCandidate Task`.parent
    # WHERE `tabCandidate Task`.task = '%s' """%(task),as_dict=True)
    return candidates
@frappe.whitelist()
def update_task():
    candidate=frappe.get_all("Candidate")
    for cand in candidate:
        doc = frappe.get_doc("Candidate",cand.name)
        if doc.candidate_task:
            for d in doc.candidate_task:
                print(doc.name)
                if d.customer:
                    frappe.db.set_value('Candidate',doc.name,'customer',d.customer)
                if d.project:
                    frappe.db.set_value('Candidate',doc.name,'project',d.project)
                if d.task:
                    frappe.db.set_value('Candidate',doc.name,'task',d.task)
                if d.pending_for:
                    frappe.db.set_value('Candidate',doc.name,'pending_for',d.pending_for)
                if d.interview_date:
                    frappe.db.set_value('Candidate',doc.name,'interview_date',d.interview_date)
                if d.offer_letter:
                    frappe.db.set_value('Candidate',doc.name,'offer_letter',d.offer_letter)
                if d.territory:
                    frappe.db.set_value('Candidate',doc.name,'territory',d.territory)
                if d.basic:
                    frappe.db.set_value('Candidate',doc.name,'basic',d.basic)
                if d.food:
                    frappe.db.set_value('Candidate',doc.name,'food',d.food)
                if d.other_allowance:
                    frappe.db.set_value('Candidate',doc.name,'other_allowances',d.other_allowance)
                if d.interview_location:
                    frappe.db.set_value('Candidate',doc.name,'interview_location',d.interview_location)
                if d.expected_doj:
                    frappe.db.set_value('Candidate',doc.name,'expected_doj',d.expected_doj)
                if d.subject:
                    frappe.db.set_value('Candidate',doc.name,'position',d.subject)

@frappe.whitelist()
def sa_candidate(task,project):
    allocated = frappe.db.sql("""
        SELECT sa_agent,sa_agent_name,sa_mobile_number,count(sa_agent) as achieved_count,
        (SELECT COUNT(pending_for) as count1 FROM `tabCandidate` cc WHERE cc.task= '%s' and cc.pending_for = 'IDB' AND 
        cc.sa_agent = sa.name) as selected,
        (SELECT COUNT(pending_for) as count2 FROM `tabCandidate` cfp WHERE cfp.task= '%s' AND cfp.pending_for IN ('Submitted','Interviewed') AND
        cfp.sa_agent = sa.name) as fp,
        (SELECT COUNT(pending_for) as count3 FROM `tabCandidate` csl WHERE csl.task= '%s' AND csl.pending_for IN ('Linedup','Shortlisted') AND
        csl.sa_agent = sa.name) as sl,
        (SELECT COUNT(pending_for) as count4 FROM `tabCandidate` cpsl WHERE cpsl.task= '%s' AND  cpsl.pending_for ='Proposed PSL' AND 
        cpsl.sa_agent = sa.name) as psl,
        (SELECT COUNT(pending_for) as count5 FROM `tabCandidate` csp WHERE csp.task= '%s' AND csp.pending_for ='Sourced' AND 
        csp.sa_agent = sa.name) as sp,
        (SELECT COUNT(pending_for) as count6 FROM `tabCandidate` tsa WHERE tsa.task= '%s' AND tsa.pending_for IN ('Submitted','Interviewed','Linedup','Shortlisted','IDB','Proposed PSL','Sourced') AND 
        tsa.sa_agent = sa.name) as tsa
        FROM `tabCandidate` c 
        JOIN `tabSAMS` sa ON  sa.name = c.sa_agent 
        WHERE c.task='%s' AND c.sa_agent IS NOT NULL AND c.task IS NOT NULL GROUP BY sa.name
        """ %(task,task,task,task,task,task,task),as_dict=1)
    return(allocated)    


@frappe.whitelist()
def project_sa_candidate(project):
    allocated = frappe.db.sql("""
        SELECT sa_agent,sa_agent_name,sa_mobile_number,count(sa_agent) as achieved_count,
        (SELECT COUNT(pending_for) as count1 FROM `tabCandidate` cc WHERE cc.project= '%s' AND cc.pending_for = 'IDB' AND 
        cc.sa_agent = sa.name) as selected,
        (SELECT COUNT(pending_for) as count2 FROM `tabCandidate` cfp WHERE cfp.project= '%s' AND cfp.pending_for IN ('Submitted','Interviewed') AND
        cfp.sa_agent = sa.name) as fp,
        (SELECT COUNT(pending_for) as count3 FROM `tabCandidate` csl WHERE csl.project= '%s' AND csl.pending_for IN ('Linedup','Shortlisted') AND
        csl.sa_agent = sa.name) as sl,
        (SELECT COUNT(pending_for) as count4 FROM `tabCandidate` cpsl WHERE cpsl.project= '%s' AND  cpsl.pending_for ='Proposed PSL' AND 
        cpsl.sa_agent = sa.name) as psl,
        (SELECT COUNT(pending_for) as count5 FROM `tabCandidate` csp WHERE csp.project= '%s' AND csp.pending_for ='Sourced' AND 
        csp.sa_agent = sa.name) as sp,
        (SELECT COUNT(pending_for) as count6 FROM `tabCandidate` tsa WHERE tsa.project= '%s' AND tsa.pending_for IN ('Submitted','Interviewed','Linedup','Shortlisted','IDB','Proposed PSL','Sourced') AND 
        tsa.sa_agent = sa.name) as tsa
        FROM `tabCandidate` c 
        JOIN `tabSAMS` sa ON  sa.name = c.sa_agent 
        WHERE c.project='%s'AND c.sa_agent IS NOT NULL AND c.project IS NOT NULL GROUP BY sa.name
        """ %(project,project,project,project,project,project,project),as_dict=1)
    return(allocated)

@frappe.whitelist()
def count_task(project):
    task = frappe.db.count('Task',{'project':project,'status':('in',["Open","Working","Overdue","Pending Review"])})
    count_vac = frappe.db.sql("""
        SELECT SUM(vac) AS total_vac
        FROM `tabTask`
        WHERE project = %s
        AND status IN ('Open', 'Working', 'Pending Review', 'Overdue')
    """, (project,), as_dict=True)
    count = frappe.db.sql("""
        SELECT SUM(sp) AS total_sp, SUM(fp) AS total_fp, SUM(sl) AS total_sl, SUM(psl) AS total_psl
        FROM `tabTask`
        WHERE project = %s
        AND status IN ('Open', 'Working', 'Pending Review', 'Overdue', 'Completed')
    """, (project,), as_dict=True)
    return task, count_vac, count

@frappe.whitelist()
def task_count(project):
    task_it = frappe.db.count('Task',{'project':project})
    work_it = frappe.db.count('Task',{'project':project,"status":"Working"})
    pending_it = frappe.db.count('Task',{'project':project,"status":"Pending Review"})
    comp_it = frappe.db.count('Task',{'project':project,"status":"Completed"})
    open_it = frappe.db.count('Task',{'project':project,"status":"Open"})
    overdue_it = frappe.db.count('Task',{'project':project,"status":"Overdue"})
    return task_it,work_it,pending_it,comp_it,open_it,overdue_it

@frappe.whitelist()
def case_count(batch):  
    
    pend_bcs = frappe.db.count('Case',{'batch':batch,"case_status":("not in",["Case Report Completed","Entry-Insuff","Execution-Insuff"])})
    com_bcs = frappe.db.count('Case',{'batch':batch,"case_status":"Case Report Completed"})
    insuff_ent = frappe.db.count('Case',{'batch':batch,"case_status":"Entry-Insuff"})
    insuff_exe = frappe.db.count('Case',{'batch':batch,"case_status":"Execution-Insuff"})
    insuff=insuff_exe+insuff_ent
    return pend_bcs,com_bcs,insuff
    

# def bulk_update_from_csv(filename):
#     #below is the method to get file from Frappe File manager
#     from frappe.utils.file_manager import get_file
#     #Method to fetch file using get_doc and stored as _file
#     _file = frappe.get_doc("File", {"file_name": filename})
#     #Path in the system
#     filepath = get_file(filename)
#     #CSV Content stored as pps

#     pps = read_csv_content(filepath[1])
#     count = 0
#     for pp in pps:
#         ld = frappe.db.exists("Lead",{'name':pp[0]})
#         if ld:
            # items = frappe.get_all("Lead",{'name':pp[0]})
            # for item in items:
            # i = frappe.get_doc('Lead',pp[0])
            # if not i.temp_mobile_no:
            #     frappe.db.set_value("Lead",pp[0],"temp_mobile_no",pp[1])
            #     print(pp[0])
                # frappe.db.commit()

@frappe.whitelist()
def leave_allocation():
    employee = frappe.get_all("Employee" ,{"status":"Active","employment_type":"Full-time"}, ['name','date_of_joining'])
    for emp in employee:
        now_date = frappe.utils.datetime.datetime.now().date()
        age = now_date.year - emp.date_of_joining.year - ((now_date.month, now_date.day) < (emp.date_of_joining.month, emp.date_of_joining.day))
        if age:
            start_date = date(now_date.year, 1, 1)
            end_date = date(now_date.year, 12, 31)
            months = now_date.month
            leave_allocation = frappe.get_all("Leave Allocation",{"employee":emp.name,"from_date":start_date,"to_date":end_date},["name"])
            if not leave_allocation:
                leave_balance = 13 - int(months)
                allocation = frappe.new_doc("Leave Allocation")
                allocation.update({
                    "employee":emp.name,
                    "leave_type":"Casual Leave",
                    "from_date":now_date,
                    "to_date":end_date,
                    "new_leaves_allocated":leave_balance
                }).save(ignore_permissions=True)
                allocation.submit()
                frappe.db.commit()

@frappe.whitelist()
def auto_mail_to_sams(project,territory):
    task =frappe.get_all("Task",filters = {"project":project}, fields = ["*"])
    for t in task:
        sams =frappe.db.sql("""select email_address,person_name from`tabSAMS` where NOT sa_status= "Do Not Contact" """,as_dict=True)
        for s in sams:
            vacancy =t.vac * t.prop
            frappe.sendmail(
                recipients=["sarumathy.d@groupteampro.com"],
                subject='Regarding Vacancy' ,
                message="""<p>Dear %s,</p>
                <p>Greetings From TEAMPRO !!! <br>
                    We are always happy to be associated with you, and appreciate your sincere efforts to be support us in recruitment.
                    We have a manpower requirement������������������������������������������������������������������������������������������������������������������������������������������������������������������for once of our reputed client in %s������������������������������������������������������������������������������������������������������������������������������������������������������������������; The details for project is as below:
                    <table class='table table-bordered'>
                    <tr>
                    <th>Position</th>  <th>Vacancy</th>  <th> Valid till</th>
                    </tr>
                    <tr>
                     <td> %s </td> <td> %s </td> <td> %s </td>
                    </tr>
                     </table>
                     <br>
                     <table class='table table-bordered'>
                     <tr>
                    <th> Job Requirement</th> 
                    </tr>
                    <tr>
                    <th> Qualification Type </th>   <th> Temp Qualification </th> <th>Minimum Experience</th>  <th>Total Experience</th>
                    </tr>
                    <tr>
                    <td> %s </td>   <td> %s </td> <td> %s </td> <td> %s </td>
                    </tr>
                    <tr>
                    <th> Salary Type </th>  <th>Specialization</th> <th> Maximum Experience </th> <th>Currency</th>
                    </tr>
                    <tr>
                    <td> %s </td>  <td>%s</td> <td> %s </td> <td> %s </td>
                    </tr>
                    <tr>
                    <th> Category </th>  <th>Gulf Experience</th> <th> Amount </th> <th>Driving Licence (If any)</th>
                    </tr>
                    <tr>
                    <td> %s </td>  <td>%s</td> <td> %s </td> <td> %s </td>
                    </tr>
                     </table>
                     <br>
                     <table class='table table-bordered'>
                     <tr>
                    <th>Job Description </th> 
                    </tr>
                    <tr>
                    <td> %s </td>
                    </tr>
                     </table>
                     <br>
                     <table class='table table-bordered'>
                    <tr>
                    <th>Allowance and Benefits  </th> 
                    </tr>
                    <tr>
                    <th> Working Days/HRS </th>  <th> Food </th> <th>Visa Type </th> <th>Transportation </th> 
                    </tr>
                    <tr>
                    <td> %s </td>  <td>%s</td> <td> %s </td> <td> %s </td>
                    </tr>
                     <tr>
                    <th> Accommodation </th>  <th> Nationality </th> <th>  Contract Period Year</th> <th> Contract Period  Month</th>
                    </tr>
                    <tr>
                    <td> %s </td>  <td>%s</td> <td> %s </td> <td> %s </td>
                    </tr>
                    <tr>
                    <th>  Over Time </th>  <th> Joining Ticket </th> <th> Leave </th> <th> Any Other Allowance</th>
                    </tr>
                    <tr>
                    <td> %s </td>  <td>%s</td> <td> %s </td> <td> %s </td>
                    </tr>
                    </table>
                    <br>
                    Please contact us at������������������������������������������������������������������������������������������������������������������������������������������������������������������+91 73050 56202������������������������������������������������������������������������������������������������������������������������������������������������������������������/������������������������������������������������������������������������������������������������������������������������������������������������������������������+91 73050 56203������������������������������������������������������������������������������������������������������������������������������������������������������������������/������������������������������������������������������������������������������������������������������������������������������������������������������������������+91 73050 56201������������������������������������������������������������������������������������������������������������������������������������������������������������������/������������������������������������������������������������������������������������������������������������������������������������������������������������������+91 7550224400
                    or write to us at������������������������������������������������������������������������������������������������������������������������������������������������������������������cv@groupteampro.com������������������������������������������������������������������������������������������������������������������������������������������������������������������/������������������������������������������������������������������������������������������������������������������������������������������������������������������hr@groupteampro.com
                    </p> 
                    <style>
                    th {
                        background-color:989898
                    }
                    </style>
                    """ % (s.person_name,territory,t.subject,vacancy,t.exp_end_date ,t.qualification_type,t.temp_qualification,
                    t.minimum_experience,t.total_experience, t.salary_type,t.specialization,t.maximum_experience,t.currency,t.category,t.gulf_experience,
                    t.amount ,t.driving_licence ,t.description,t.working_days,t.food,t.visa_type,t.transportation,t.accommodation,t.nationality,
                    t.contract_period_year,t.contract_period__month,t.over_time,t.joining_ticket,t.leave,t.any_other_allowance,))

@frappe.whitelist()
def update_sams():
    sams = frappe.get_all("SAMS",{"sa_is_an_organization":1},["name","private_limited_company","proprietary_company"])
    for sm in sams:
        print(sm.name)
        if sm.private_limited_company:
            sam = frappe.get_doc("SAMS",sm.name)
            print(sam.name)
            print(sam.organization_name)
            sam.update({
                "company":"Pvt.Ltd.Company",
                "type":"Agent"
            }).save(ignore_permissions=True)
            frappe.db.commit()
        if sm.proprietary_company:
            sam = frappe.get_doc("SAMS",sm.name)
            print(sam.name)
            print(sam.organization_name)
            sam.update({
                "company":"Proprietary Company",
                "type":"Agent"
            }).save(ignore_permissions=True)
            frappe.db.commit()

# @frappe.whitelist()
# def late_entry():
#     employee = frappe.get_all("Employee",{"status":"Active"},["name"])
#     import calendar
#     now_date = frappe.utils.datetime.datetime.now().date()
#     month = calendar.monthrange(now_date.year, now_date.month)
#     start_date = date(now_date.year, now_date.month, 1)
#     end_date = date(now_date.year, now_date.month, month[1])
#     for emp in employee:
#         attendance = frappe.db.sql("""select name,employee,employee_name,shift,late_entry from `tabAttendance` where late_entry = 1 and employee = %s and attendance_date between %s and %s""",(emp.name,start_date,end_date),as_dict = 1)
#         count = len(attendance)

@frappe.whitelist()
def update_task(project,status):
    # for p in project:
    task = frappe.get_all("Task",{"project":project})
    for t in task:
        if(status == "Completed"):
            t.status = frappe.db.set_value("Task",t.name,"status","Completed")
        elif(status=="Cancelled"):
            t.status = frappe.db.set_value("Task",t.name,"status","Cancelled")
           
# @frappe.whitelist()
# def check_package_list(customer):    
#     cpb =frappe.db.sql("""select name from `tabCheck Package` where customer ='%s' """%(customer),as_dict=1)[0]
#     return cpb['name'] or ''

@frappe.whitelist()
def file_list():
    pdf = frappe.db.get_value('File', {'attached_to_doctype': 'Candidate'}['attached_to_name'])
    for i in pdf:
        print (i)

@frappe.whitelist()
def update_report():
    report = frappe.db.sql(""" update `tabEmployment` set workflow_state = "Draft" where name= "Employment-815" """)
    # frappe.db.set_value("Case","CS-BT-MF-2023-08-01-2257-2260","mobile_no","9747712411")

@frappe.whitelist()
def delete():
    report = frappe.db.sql(""" delete from `tabBatch` where name= "BT-KSL-2023-09-23-3910" """)
    # report1 = frappe.db.sql(""" delete from `tabEducation Checks` where name= "Education Checks-8392" """)
    # report2 = frappe.db.sql(""" delete from `tabEducation Checks` where name= "Education Checks-8391" """)


@frappe.whitelist()
def update_checkin():
    report = frappe.db.sql(""" update `tabEmployee Checkin` set attendance = "" where name= "EMP-CKIN-10-2023-000393" """)

@frappe.whitelist(allow_guest=True)
def create_task(**args):
    mattermost_data = json.loads(args)
    frappe.log_error(title='args',message=mattermost_data)
    return "ok" 

# @frappe.whitelist()
# def it_dpr():
#     date = today()
#     webhook_url = "https://pm.teamproit.com/hooks/htnwah6ne7ydicmqe9gw5ac6no"
#     task_data = frappe.db.sql("""SELECT * FROM tabTask WHERE status = 'Working' AND service = 'IT-SW' ORDER BY cb ASC""", as_dict=True)

#     row = f"###### @all DPR {date} \n"
#     row += "| Sr. |  ID  | Project | Subject | CB | Status | Revisions | AT | ET | RT | Priority |  Allocated On  |\n"
#     row += "| --- | ---- | ------- | ------- | -- | ------ | --------- | -- | -- | -- | -------- | -------------- |\n"
#     idx = 1
#     for data in task_data:
#         row += f"| {idx} | {data.name} | {data.project} | {data.subject} | {data.cb} | {data.status} | {data.revisions} | {data.actual_time} | {data.expected_time} | {data.rt} | {data.priority} | {data.custom_allocated_on} |\n"
#         idx = idx + 1
#     payload = {"text": row}

#     response = requests.post(webhook_url, json=payload)
#     print(response.status_code)
#     # print(task_data)

# def auto_dpr():
#     job = frappe.db.exists('Scheduled Job Type', 'it_dpr')
#     if not job:
#         sjt = frappe.new_doc("Scheduled Job Type")
#         sjt.update({
#             "method": 'jobpro.custom.it_dpr',
#             "frequency": 'Cron',
#             "cron_format": '00 1 * * *'
#         })
#         sjt.save(ignore_permissions=True)
@frappe.whitelist()
def update_sams_by(doc,method):
    if doc.sa_agent:
        frappe.db.set_value("Candidate",doc.name,'custom_sourced_by',"SAMS")
    else:
            frappe.db.set_value("Candidate",doc.name,'custom_sourced_by',"Normal")
       
    
@frappe.whitelist()
def update_task_counts(doc,method):
    if doc.task:
        submitted = frappe.db.count(
            'Candidate', {'task':doc.task, 'pending_for': 'Submitted(Client)'}) or 0
        psl = frappe.db.count('Candidate', {'task': doc.task, 'pending_for': (
            'in', ('Submitted(Client)', 'Proposed PSL'))}) or 0
        shortlisted = frappe.db.count(
            'Candidate', {'task': doc.task, 'pending_for':'Shortlisted'}) or 0
        linedup = frappe.db.count(
            'Candidate', {'task':doc.task, 'pending_for': 'Linedup'}) or 0
        interviewed = frappe.db.count(
            'Candidate', {'task':doc.task, 'pending_for': 'Interviewed'}) or 0
        
        frappe.db.set_value('Task', doc.task, 'psl', psl)
        frappe.db.set_value('Task', doc.task, 'fp', submitted + interviewed)
        frappe.db.set_value('Task', doc.task, 'sl', shortlisted)
        frappe.db.set_value('Task', doc.task, 'custom_lp',linedup)

        task_status = frappe.db.get_value('Task',doc.task, 'status')
        if task_status in ('Completed', 'Cancelled'):
            # if pps == 0:
            frappe.db.set_value('Task',doc.task, 'sp', 0)
        else:
            vac = frappe.db.get_value('Task', doc.task, 'vac')
            prop = frappe.db.get_value('Task',doc.task, 'prop')
            pps = (vac - psl) * prop - (submitted +
                                        interviewed + shortlisted +linedup)
            frappe.db.set_value('Task',doc.task, 'sp', pps)

# @frappe.whitelist()
# def passport_num_validate(passport_number,c_name):
#     if passport_number and c_name:
#         existing = frappe.db.exists('Candidate', {'passport_number':passport_number, 'given_name': ['!=',c_name]})
#         if existing :
#             return "Duplicate"

@frappe.whitelist()
def check_candidate_for_sourced(candidate_id):
    candidate = frappe.get_doc("Candidate", candidate_id)
    for i in candidate.custom_interview_history:
        if i.client_name == candidate.customer:
            interviewed_date_str = candidate.interviewed_date.strftime('%d-%m-%Y')
            return_msg= (
                f"This candidate has already attended an interview with this client on {interviewed_date_str} "
                f"at {candidate.interview_location} location, please contact the TM / HOD to move this candidate to sourced"
            )
            return return_msg


@frappe.whitelist()
def check_availability(name):
    project=frappe.db.get_all("Task",{'project':name},'name')
    for i in project:
        task=frappe.get_doc("Task",i.name)
        if not task.custom_project_plan_:
            return_ms=frappe.throw(f"Project cannot be confirmed without project details.")
        # else:
        #     first_item = task.custom_project_plan_[0]
        #     if not first_item:
        #         return_ms=frappe.throw(f"Project cannot be confirmed without project details.")
            return return_ms

@frappe.whitelist()
def validate_unique_passport(doc, method):
    if doc.passport_number:
        # existing_candidate = frappe.db.exists('Candidate', {'passport_number': doc.passport_number,'given_name':('!=',doc.given_name)})
        existing_candidate = frappe.db.exists('Candidate', {'passport_number': doc.passport_number})       
        if existing_candidate:
            frappe.throw(f"A candidate with Passport Number {doc.passport_number} already exists.")                   

@frappe.whitelist()
def validate_passport(passport):
    if passport:
        existing_candidate = frappe.db.exists('Candidate', {'passport_number': passport})
        if existing_candidate:
            frappe.throw(f"A candidate with Passport Number {passport} already exists.")
        
@frappe.whitelist()     
def on_click_create_mail(project,name,task,vac):
    create=frappe.get_doc("Project",name)
    table = '<table text-align="center" border="1" width="25%" style="border-collapse: collapse;">'
    table += '<tr style="background-color: #87CEFA"><td style= width="1%";font-weight: bold;">Task ID</td><td style="width:1%; font-weight: bold;">Subject</td><td style="width: 1%; font-weight: bold;">No of Vacancies</td><td style="width:1%; font-weight: bold;">SP</td></tr>'
    task=frappe.db.get_all("Task",{'project':name},['name','subject','vac','sp'])
    for i in task:
        tot_sp=frappe.db.sql("""SELECT sum(sp) as sp from `tabTask` where project=%s group by project""",(create.name),as_dict=True)[0]
        tot_vac=frappe.db.sql("""SELECT sum(vac) as vac from `tabTask` where project=%s group by project""",(create.name),as_dict=True)[0]
        table+="""<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>"""%(i.name,i.subject,i.vac,i.sp)
    table+="""<tr><td colspan=2 style="text-align: center;">Total</td><td>%s</td><td>%s</td></tr>"""%(round(tot_vac['vac']) or '',round(tot_sp['sp']) or '')
    table+='</table>'
    subject="Action Created-  %s" % nowdate()
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Action Created Details:<p><b>Project Name:</b>{}<p><b>Project ID:</b>{}– a new project has been created for your further action.<br>{}<br></p><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(create.project_name,create.name,table)
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["sangeetha.a@groupteampro.com","dineshbabu.k@groupteampro.com","sangeetha.s@groupteampro.com"],
        subject=subject,
        message=message
    )
@frappe.whitelist()     
def on_click_confirm_mail(project,name):
    create=frappe.get_doc("Project",name)
    # det=frappe.db.get_all("Task",{'project_name':'project','project':'name'})
    table= '<table text-align="center" border="1" width="25%" style="border-collapse: collapse;">'
    table += '<tr style="background-color: #87CEFA"><td style="width:1%; font-weight: bold;">Task ID</td><td style="width: 1%; font-weight: bold;">Subject</td><td style="width: 1%; font-weight: bold;">No of Vacancies</td><td style="width: 1%; font-weight: bold;">#SP</td><td style="width: 1%; font-weight: bold;">Date Batch-1</td><td style="width: 1%; font-weight: bold;">Date Batch-2</td></tr>'
    task=frappe.db.get_all("Task",{'project':name},['name','subject','vac','sp'])
    manager=frappe.db.get_value("Project",{'name':name},['account_manager'])
    for i in task:
        tot_sp=frappe.db.sql("""SELECT sum(sp) as sp from `tabTask` where project=%s group by project""",(create.name),as_dict=True)[0]
        tot_vac=frappe.db.sql("""SELECT sum(vac) as vac from `tabTask` where project=%s group by project""",(create.name),as_dict=True)[0]
        table+="""<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td></td><td></td></tr>"""%(i.name,i.subject,i.vac,i.sp)
    table+="""<tr><td colspan=2 style="text-align: center;">Total</td><td>%s</td><td>%s</td><td></td><td></td></tr>"""%(round(tot_vac['vac']) or '',round(tot_sp['sp']) or '')
    table+='</table>'
    subject="Action Confirmed-  %s" % nowdate()
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Action Created Details:<p><b>Project Name:{}</b><p><b>Project ID:</b>{}– has been confirmed for execution by OPS Team.<br>{}<br></p><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(create.project_name,create.name,table)
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["dineshbabu.k@groupteampro.com","sangeetha.s@groupteampro.com","annie.m@groupteampro.com",manager],
        subject = subject,
        message=message
    
        ) 

# @frappe.whitelist()
# def on_click_confirm_mail(project,name):
#     create = frappe.get_doc("Project",name)
#     tasks = frappe.db.get_all("Task", filters={'project': create.name}, fields=['name'])
#     manager=frappe.db.get_value("Project",{'name':create.name},['account_manager'])
#     details = []
#     if tasks:
#         task_names = [task['name'] for task in tasks]
#         details = frappe.get_all("Project Plan Child", filters={'parent': ['in', task_names]}, fields=['parent', 'date', 'count'])
#     table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
#     table += '<tr style="background-color: #87CEFA"><td style="width: 20%; font-weight: bold; text-align: center;">Task ID</td><td style="width: 20%; font-weight: bold; text-align: center;">Subject</td><td style="width: 20%; font-weight: bold; text-align: center;">No Of Vacancies</td><td style="width: 20%; font-weight: bold; text-align: center;">#SP</td><td style="width: 20%; font-weight: bold; text-align: center;">Date Batch-1</td><td style="width: 20%; font-weight: bold; text-align: center;">Date Batch-2</td></tr>'
#     task_dict = {task['name']: task for task in tasks}  # Dictionary to store task info by task name
#     for record in details:
#         task = task_dict.get(record['parent'])
#         table += f"<tr><td>{record['parent']}</td><td>{task['subject']}</td><td>{task['vac']}</td><td>{task['sp']}</td><td>{record['date']}</td><td>{record['count']}</td></tr>"
    
#     # for record in details:
#     #     table += f"<tr><td>{record['parent']}</td><td>{record['date']}</td><td>{record['count']}</td></tr>"
#     table += "</table>"
#     email_message = f"""   
#         Dear Sir/Mam,<br>
#         <p><b>Project Name: </b>{create.project_name} & <b>Project ID:</b>{create.name} – has been confirmed for execution by OPS Team.</p><br>
#         {table}<br>
#         Thanks & Regards<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
#     """
    
#     # Send the email
#     frappe.sendmail(
#         recipients=["divya.p@groupteampro.com"],
#         # recipients=["dineshbabu.k@groupteampro.com","sangeetha.s@groupteampro.com","annie.m@groupteampro.com",manager],
#         subject=f"New Action Confirmed - {nowdate()}",
#         message=email_message
#     )


@frappe.whitelist()     
def on_creation_of_psl_mail(doc,method):
    creates=frappe.get_doc("Closure",doc)
    candidate_mail=frappe.db.get_value("Candidate",{"name":doc.candidate},['mail_id'])
    acc=frappe.db.get_value("Candidate",{"name":doc.candidate},['task'])
    acc_manager=frappe.db.get_value("Task",{'name':acc},['account_manager'])
    spoc=frappe.db.get_value("Task",{'name':acc},['spoc'])
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["sangeetha.s@groupteampro.com","dc@groupteampro.com",acc_manager,spoc],
        subject = "New PSL Created -  %s" % nowdate(),
        message="""   
        Dear Sir/Mam,<br>
        <p><b>Closure ID: </b>%s  <b> Status :</b>%s  <b> Customer Name :</b>%s -a new PSL added in Closure for your further action.</p><br>
        
            Thanks & Regards<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
    """ % (doc.name,doc.status,doc.customer)
        ) 

@frappe.whitelist() 
def fp_candidate_list_send_mail():
    projects=frappe.get_all("Project",{'status':'Open','service':('in',['REC-D','REC-I'])},['*'])
    
    for i in projects:
        s_no=0
        table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
        table += '<tr style="background-color: #87CEFA"><td style="width: 45%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">CDID</td><td style="width: 25%; font-weight: bold; text-align: center;">Candidate Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Candidate Name</td><td style="width: 25%; font-weight: bold; text-align: center;">Position</td><td style="width: 25%; font-weight: bold; text-align: center;">Project Name</td><td style="width: 30%; font-weight: bold; text-align: center;">Project ID</td><td style="width: 25%; font-weight: bold; text-align: center;">Customer Name</td><td style="width: 25%; font-weight: bold; text-align: center;">Age</td><td style="width: 25%; font-weight: bold; text-align: center;">Next Contact On</td></tr>'
        tasks = frappe.get_all("Task", {'status': ('in', ['Open', 'Working','Overdue','Pending Review']),'project':i.name},['*'],group_by= "spoc")
        acc_manager=frappe.db.get_value("Task",{'project':i.name},['account_manager'])
        spoc=frappe.db.get_value("Task",{'project':i.name},['spoc'])
        task_count=frappe.db.count("Task", {'status': ('in', ['Open', 'Working','Overdue','Pending Review']),'project':i.name})
        # table+="""<tr><td></td><td></td><td></td><td></td><td></td><td>%s</td><td></td><td></td><td></td><td></td></tr>"""%(i.project_name)
        print(acc_manager)
        print(spoc)
        print(task_count)
        if task_count>0:
            row=0
            for j in tasks:
                candidate=frappe.get_all("Candidate",{'pending_for':('not in',['IDB','Sourced','Proposed PSL']),'task':j.name},['name','pending_for','given_name','position','project_name','project','customer','age_of_cv','custom_next_contact_on'],group_by= "customer ASC")
                for ca in candidate:
                    row+=1
            # table+="""<tr><td colspan=5></td><td>%s</td><td colspan=4></td></tr>"""%(i.project_name)
            table+="""<tr><td colspan=5></td><td rowspan=%s>%s</td><td colspan=3></tr>"""%(row+1,i.project_name)
            for j in tasks:
                candidate=frappe.get_all("Candidate",{'pending_for':('not in',['IDB','Sourced','Proposed PSL']),'task':j.name},['name','pending_for','given_name','position','project_name','project','customer','age_of_cv','custom_next_contact_on'],group_by= "customer ASC")
                for ca in candidate:
                    s_no+=1
                    table+="""<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" % (s_no,ca.name,ca.pending_for,ca.given_name,ca.position,ca.project,ca.customer,ca.age_of_cv,ca.custom_next_contact_on or '')
            table += '</table>'
            subject = "FP List -  %s" % nowdate()
            message = """
            Dear Sir/Madam,<br><br>
            Kindly find the below list of  your FP List :<br><br>{}<br><br>
            Thanks & Regards,<br>TEAM ERP<br>
            <i>This email has been automatically generated. Please do not reply</i>
            """.format(table)

            frappe.sendmail(
                # recipients=["divya.p@groupteampro.com"],
                recipients=[acc_manager,spoc],
                subject=subject,
                message=message,
            )
@frappe.whitelist()
def send_mail_to_creation_adv(name):
    pro = frappe.get_doc("Project",name)
    tasks = frappe.get_all("Task",{'project': pro.name},['*'])
    t=frappe.db.get_all("Task",{'project': pro.name},['food'],group_by='food')
    food_count=len(t)
    qualification=frappe.get_all("Task",{'project': pro.name},['qualification_type'],group_by='qualification_type')
    qual_count=len(qualification)
    experience=frappe.get_all("Task",{'project': pro.name},['total_experience'],group_by='total_experience')
    exp_count=len(experience)
    g_experience=frappe.get_all("Task",{'project': pro.name},['gulf_experience'],group_by='gulf_experience')
    g_exp=len(g_experience)
    interview=frappe.get_all("Task",{'project': pro.name},['mode_of_interview'],group_by='mode_of_interview')
    int_count=len(interview)
    acc=frappe.get_all("Task",{'project': pro.name},['accommodation'],group_by='accommodation')
    a_count=len(acc)
    transport=frappe.get_all("Task",{'project': pro.name},['transportation'],group_by='transportation')
    trans_count=len(transport)
    visa=frappe.get_all("Task",{'project': pro.name},['visa_type'],group_by='visa_type')
    v_count=len(visa)
    con=frappe.get_all("Task",{'project': pro.name},['contract_period_year'],group_by='contract_period_year')
    con_count=len(con)
    categorys=frappe.get_all("Task",{'project': pro.name},['category'],group_by='category')
    ca_count=len(categorys)
    keys=frappe.get_all("Task",{'project': pro.name},['custom_major_key_skills'],group_by='custom_major_key_skills')
    key_count=len(keys)
    rec=frappe.get_all("Task",{'project': pro.name},['custom_free_recruitment'],group_by='custom_free_recruitment')
    rec_count=len(rec)
    task_count=(frappe.db.count("Task",{'project': pro.name}))
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td style="width: 10%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">Title</td><td style="width: 60%; font-weight: bold; text-align: center;">Details</td></tr>'
    table += """<tr><td></td><td>Project ID</td><td>{}</td></tr>""".format(pro.name or '')
    table += """<tr><td></td><td>Date</td><td>{}</td></tr>""".format(pro.custom_actionconfirmed_datetime or '')
    table += """<tr><td>1</td><td>Country</td><td>{}</td></tr>""".format(pro.territory or '')
    table += """<tr><td></td><td>Client</td><td>{}</td></tr>""".format(pro.customer or '')
    s_no = 0
    for i in tasks:
        if tasks.index(i)==0:
            table += """<tr><td rowspan={}>2</td><td rowspan={}>Positions</td><td>{}</td></tr>""".format(task_count,task_count,i.subject or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(i.subject or '')
    for k in keys:
        if keys.index(k)==0:
            table += """<tr><td rowspan={}>3</td><td rowspan={}>Major Key Skills</td><td>{}</td></tr>""".format(key_count,key_count,k.custom_major_key_skills or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(k.custom_major_key_skills or '')
    for d in qualification:
        if qualification.index(d)==0:
            table += """<tr><td rowspan={}>4</td><td rowspan={}>Qualification</td><td>{}</td></tr>""".format(qual_count,qual_count,d.qualification_type or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(d.qualification_type or '')
    for e in experience:
        if experience.index(e)==0:
            table += """<tr><td rowspan={}>5</td><td rowspan={}>Experience</td><td>{}</td></tr>""".format(exp_count,exp_count,e.total_experience or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(e.total_experience or '')
    for g in g_experience:
        if g_experience.index(g)==0:
            table += """<tr><td rowspan={}>6</td><td rowspan={}>GCC Experience</td><td>{}</td></tr>""".format(g_exp,g_exp,g.gulf_experience or '')
        else:
            table +="""<tr><td>{}</td></tr>""".format(g.gulf_experience or '')
    for r in rec:
        if rec.index(r)==0:
            table += """<tr><td rowspan={}>7</td><td rowspan={}>Free Recruitment</td><td>{}</td></tr>""".format(rec_count,rec_count,r.custom_free_recruitment or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(r.custom_free_recruitment or '')
    for m in interview:
        if interview.index(m)==0:
            table += """<tr><td rowspan={}>8</td><td rowspan={}>Mode Of Interview</td><td>{}</td></tr>""".format(int_count,int_count,m.mode_of_interview or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(m.mode_of_interview or '')
    table += """<tr><td>9</td><td>If Direct Client Interview - Location & Date</td><td></td></tr>"""
    for j in t:
            if t.index(j)==0:
                table += """<tr><td rowspan={}>10</td><td rowspan={}>Food</td><td>{}</td></tr>""".format(food_count,food_count,j.food or '')
            else:
                table +=  """<tr><td>{}</td></tr>""".format(j.food or '')
    for a in acc:
            if acc.index(a)==0:
                table += """<tr><td rowspan={}>11</td><td rowspan={}>Accomodation</td><td>{}</td></tr>""".format(a_count,a_count,a.accommodation or '')
            else:
                table +=  """<tr><td>{}</td></tr>""".format(a.accommodation or '')
    for tr in transport:
            if transport.index(tr)==0:
                table += """<tr><td rowspan={}>12</td><td rowspan={}>Transportation</td><td>{}</td></tr>""".format(trans_count,trans_count,tr.transportation or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(tr.transportation or '')
    table += """<tr><td>13</td><td>Contact Number</td><td></td></tr>"""
    table += """<tr><td>14</td><td>Mail ID</td><td></td></tr>"""
    for v in visa:
            if visa.index(v)==0:
                table += """<tr><td rowspan={}>15</td><td rowspan={}>Visa Type</td><td>{}</td></tr>""".format(v_count,v_count,v.visa_type or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(v.visa_type or '')
    for cons in con:
            if con.index(cons)==0:
                table += """<tr><td rowspan={}>16</td><td rowspan={}>Contract</td><td>{}</td></tr>""".format(con_count,con_count,cons.contract_period_year or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(cons.contract_period_year or '')
    for cat in categorys:
            if categorys.index(cat)==0:
                table += """<tr><td rowspan={}>17</td><td rowspan={}>ECR/ECNR</td><td>{}</td></tr>""".format(ca_count,ca_count,cat.category or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(cat.category or '')
    table += """<tr><td>18</td><td>Special Remarks</td><td>Attractive Salary</td></tr>"""
    table += """<tr><td>19</td><td>Common Version 1.0</td><td>Company Name + Logo + RA Licence + Location + Website + Common Number (7305056202) + Common Mail ID</td></tr>"""
    table += '</table>'
    subject = " Action Confirmed -  {}".format(frappe.utils.nowdate())
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Action Confirmed Details:<br><br>{}<br><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(table)
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["dineshbabu.k@groupteampro.com","dm@groupteampro.com","annie.m@groupteampro.com"],
        subject=subject,
        message=message
    )


@frappe.whitelist()     
def send_mail_to_drop(name):
    create=frappe.get_doc("Closure",name)
    candidate=frappe.db.get_all("Candidate",{"name":create.candidate},['task'])
    spoc=frappe.db.get_value("Task",{'name':candidate},['spoc'])
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=[spoc],
        subject = "Candidate Droped-  %s" % nowdate(),
        message="""   
        Dear Sir/Mam,<br>
        <p><b>Closure ID: </b>%s – has been Droped.Additional CV is Required</p><br>
        
            Thanks & Regards<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
    """ % (create.name)
        ) 
    
@frappe.whitelist() 
def fp_candidate_list_send_mail_to_spoc():
    startdate=nowdate()
    next_day=add_days(startdate,1)
    projects = frappe.get_all("Project", filters={'status': 'Open', 'service': ['in', ['REC-D', 'REC-I']]}, fields=['name', 'spoc'])
    spoc_set = set()
    candidate_count=0
    for project in projects:
        if project.get('spoc'):
            spoc_set.add(project['spoc'])
    spoc_list = list(spoc_set)
    # table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    print(spoc_list)
    for spoc in spoc_list:
        spoc_projects = frappe.get_all("Project", filters={'status': 'Open', 'service': ['in', ['REC-D', 'REC-I']], 'spoc': spoc,'custom_spoc__next_contact_on':next_day}, fields=['*'])
        table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
        for project in spoc_projects:
            s_no = 0
            tasks = frappe.get_all("Task", filters={'status': ('in', ['Open', 'Working', 'Overdue', 'Pending Review']), 'project': project['name'], 'service': ('in', ['REC-D', 'REC-I'])}, fields=['name'])
            candidate_count = frappe.db.count("Candidate", filters={'project': project['name'], 'pending_for': ('not in', ['IDB', 'Sourced', 'Proposed PSL'])})
            if candidate_count > 0:
                row = 0
                table+="""<tr style="text-align: center;"><td style="border-left: none; border-right: none;"colspan=10 %s>%s</td></tr>"""%(row+1,project['project_name'])
                table += '<tr style="background-color: #87CEFA"><td style="width: 15%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">CDID</td><td style="width: 25%; font-weight: bold; text-align: center;">Candidate Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Given Name/Surname</td><td style="width: 30%; font-weight: bold; text-align: center;">Passport No</td><td style="width: 25%; font-weight: bold; text-align: center;">Position</td><td style="width: 30%; font-weight: bold; text-align: center;">Candidate Owner</td><td style="width: 40%; font-weight: bold; text-align: center;">Project ID</td><td style="width: 40%; font-weight: bold; text-align: center;">Customer Name</td><td style="width: 25%; font-weight: bold; text-align: center;">Age</td><td style="width: 30%; font-weight: bold; text-align: center;">Next Contact On</td></tr>'
                for task in tasks:
                    s_no += 1
                    candidates = frappe.get_all("Candidate", filters={'pending_for': ('not in', ['IDB', 'Sourced', 'Proposed PSL']), 'task': task['name']}, fields=['name', 'pending_for', 'given_name','passport_number', 'position', 'candidate_created_by', 'project', 'customer', 'age_of_cv', 'custom_next_contact_on'])
                    for candidate in candidates:
                        row += 1
                        table += """<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" % (s_no, candidate['name'], candidate['pending_for'], candidate['given_name'],candidate['passport_number'], candidate['position'], candidate['candidate_created_by'], candidate['project'], candidate['customer'], candidate['age_of_cv'], candidate['custom_next_contact_on'] or '')
        table += '</table>'
        subject = "FP List - %s" % nowdate()
        message = """
        Dear Sir/Madam,<br><br>
        Kindly find the below list of your FP List for Next Contact On<br><br>{}<br><br>
        Thanks & Regards,<br>TEAM ERP<br>
        <i>This email has been automatically generated. Please do not reply</i>
        """.format(table)
        if candidate_count>0:
            frappe.sendmail(
                recipients=[spoc],
                # recipients=["divya.p@groupteampro.com"],
                subject=subject,
                message=message,
            )


@frappe.whitelist() 
def update_task():
    project=frappe.get_all("Project",{'status':'Open','service':('in',['REC-D','REC-I'])},['*'])
    for i in project:
        tasks = frappe.get_all("Task", {'status': ('in', ['Open', 'Working','Overdue','Pending Review']),'project':i.name},['*'])
        for j in tasks:
            submitted_client = frappe.db.count('Candidate', {'task': j.name, 'pending_for':'Submitted(Client)'}) or 0
            print(submitted_client)
            submitted_spoc = frappe.db.count('Candidate', {'task': j.name, 'pending_for':'Submit(SPOC)'}) or 0          
            psl = frappe.db.count('Candidate', {'task': j.name, 'pending_for': ('in', ('Client Offered', 'Proposed PSL'))}) or 0
            # shortlisted = frappe.db.count('Candidate', {'task': j.name, 'pending_for':'Shortlisted'}) or 0
            # linedup = frappe.db.count('Candidate', {'task': j.name, 'pending_for': 'Linedup'}) or 0
            interviewed = frappe.db.count('Candidate', {'task': j.name, 'pending_for': 'Interviewed'}) or 0
            # result_pending = frappe.db.count('Candidate', {'task': j.name, 'pending_for': 'Result Pending'}) or 0
            total=(submitted_client + interviewed + submitted_spoc)
            # frappe.db.set_value('Task', j.name, 'psl', psl)
            frappe.db.set_value('Task', j.name, 'fp',total)
            # frappe.db.set_value('Task', j.name, 'sl', shortlisted)
            # frappe.db.set_value('Task', j.name, 'custom_lp',linedup)
            # task_status = frappe.db.get_value('Task',j.name, 'status')
            # if task_status in ('Completed', 'Cancelled'):
            #     frappe.db.set_value('Task', j.name, 'sp', 0)
            # else:
            #     vac = frappe.db.get_value('Task', j.name, 'vac')
            #     prop = frappe.db.get_value('Task', j.name, 'prop')
            #     pps = (vac - psl) * prop - (submitted_client + submitted_spoc +interviewed + shortlisted +linedup)
            #     frappe.db.set_value('Task', j.name, 'sp', pps)

@frappe.whitelist() 
def update_project():
    project=frappe.get_all("Project",{'status':'Open','service':('in',['REC-D','REC-I'])},['*'])
    for i in project:
        tot_fp=frappe.db.sql("""SELECT sum(fp) as fp from `tabTask` where project=%s group by project""",(i.name),as_dict=True)[0]
        tot_psl=frappe.db.sql("""SELECT sum(psl) as psl from `tabTask` where project=%s group by project""",(i.name),as_dict=True)[0]
        tot_sl=frappe.db.sql("""SELECT sum(sl) as sl from `tabTask` where project=%s group by project""",(i.name),as_dict=True)[0]
        tot_sp=frappe.db.sql("""SELECT sum(sp) as sp from `tabTask` where project=%s group by project""",(i.name),as_dict=True)[0]
        frappe.db.set_value("Project",i.name,'tfp',tot_fp['fp'])
        frappe.db.set_value("Project",i.name,'tpsl',tot_psl['psl'])
        frappe.db.set_value("Project",i.name,'tsl',tot_sl['sl'])
        frappe.db.set_value("Project",i.name,'tsp',tot_sp['sp'])
        

@frappe.whitelist() 
def update_project_count(doc,method):
    if doc.project:
        tot_fp=frappe.db.sql("""SELECT sum(fp) as fp from `tabTask` where project=%s """,(doc.project),as_dict=True)[0]
        tot_psl=frappe.db.sql("""SELECT sum(psl) as psl from `tabTask` where project=%s """,(doc.project),as_dict=True)[0]
        tot_sl=frappe.db.sql("""SELECT sum(sl) as sl from `tabTask` where project=%s """,(doc.project),as_dict=True)[0]
        tot_sp=frappe.db.sql("""SELECT sum(sp) as sp from `tabTask` where project=%s """,(doc.project),as_dict=True)[0]
        tot_lp=frappe.db.sql("""SELECT sum(custom_lp) as lp from `tabTask` where project=%s""",(doc.project),as_dict=True)[0]
        tot_rp=frappe.db.sql("""SELECT sum(custom_rp) as rp from `tabTask` where project=%s""",(doc.project),as_dict=True)[0]
        if tot_fp['fp'] is not None:
            frappe.db.set_value("Project",doc.project,'tfp',tot_fp['fp'])
        if tot_psl['psl'] is not None:
            frappe.db.set_value("Project",doc.project,'tpsl',tot_psl['psl'])
        if tot_sl['sl'] is not None:
            frappe.db.set_value("Project",doc.project,'tsl',tot_sl['sl'])
        if tot_sp['sp'] is not None:
            frappe.db.set_value("Project",doc.project,'tsp',tot_sp['sp'])
        if tot_lp['lp'] is not None:
            frappe.db.set_value("Project",doc.project,'custom_t_lp',tot_lp['lp'])  
        if tot_rp['rp'] is not None:
            frappe.db.set_value("Project",doc.project,'custom_t_rp',tot_rp['rp'])              
        # frappe.db.set_value("Project", doc.project, 'tfp', tot_fp[0].get('fp', 0))
        # frappe.db.set_value("Project", doc.project, 'tpsl', tot_psl[0].get('psl', 0))
        # frappe.db.set_value("Project", doc.project, 'tsl', tot_sl[0].get('sl', 0))
        # frappe.db.set_value("Project", doc.project, 'tsp', tot_sp[0].get('sp', 0))

@frappe.whitelist() 
def fp_candidate_list_send_mails():
    projects=frappe.get_all("Project",{'status':'Open','service':('in',['REC-D','REC-I'])},['*'])
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    for i in projects:
        s_no=0

        # row=0
        tasks = frappe.get_all("Task", {'status': ('in', ['Open', 'Working','Overdue','Pending Review']),'project':i.name,'service':('in',['REC-D','REC-I'])},['*'])
        # acc_manager=frappe.db.get_value("Task",{'project':i.name},['account_manager'])
        # spoc=frappe.db.get_value("Task",{'project':i.name},['spoc'])
        task_count=frappe.db.count("Task", {'status': ('in', ['Open', 'Working','Overdue','Pending Review']),'project':i.name,'service':('in',['REC-D','REC-I'])})
        candidate_count=frappe.db.count("Candidate", {'project':i.name,'pending_for':('not in',['IDB','Sourced','Proposed PSL'])})
        if candidate_count>0:
            row=0
            table+="""<tr style="text-align: center;"><td style="border-left: none; border-right: none;"colspan=10 %s>%s</td></tr>"""%(row+1,i.project_name)
            # table += '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
            table += '<tr style="background-color: #87CEFA"><td style="width: 15%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">CDID</td><td style="width: 25%; font-weight: bold; text-align: center;">Candidate Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Given Name/Surname</td><td style="width: 25%; font-weight: bold; text-align: center;">Position</td><td style="width: 30%; font-weight: bold; text-align: center;">Candidate Owner</td><td style="width: 40%; font-weight: bold; text-align: center;">Project ID</td><td style="width: 40%; font-weight: bold; text-align: center;">Customer Name</td><td style="width: 25%; font-weight: bold; text-align: center;">Age</td><td style="width: 30%; font-weight: bold; text-align: center;">Next Contact On</td></tr>'
            for j in tasks:
                s_no+=1
                candidate=frappe.get_all("Candidate",{'pending_for':('not in',['IDB','Sourced','Proposed PSL']),'task':j.name},['name','pending_for','given_name','position','candidate_created_by','project_name','project','customer','age_of_cv','custom_next_contact_on'])
                for ca in candidate:
                    row+=1
                    # s_no+=1
                    table+="""<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" % (s_no,ca.name,ca.pending_for,ca.given_name,ca.position,ca.candidate_created_by,ca.project,ca.customer,ca.age_of_cv,ca.custom_next_contact_on or '')
    table += '</table>'
    subject = "FP List -  %s" % nowdate()
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below list of  your FP List :<br><br>{}<br><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(table)
    # if row>1:
    frappe.sendmail(
        recipients=["sangeetha.a@groupteampro.com"],
        # recipients=["divya.p@groupteampro.com"],
        subject=subject,
        message=message,
    )

import frappe
from frappe.utils import nowdate

import frappe

@frappe.whitelist() 
def fp_candidate_to_spoc():
    projects = frappe.get_all("Project", filters={'status': 'Open', 'service': ['in', ['REC-D', 'REC-I']]}, fields=['name', 'spoc'])
    spoc_set = set()
    for project in projects:
        if project.get('spoc'):
            spoc_set.add(project['spoc'])
    spoc_list = list(spoc_set)
    # table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    print(spoc_list)
    for spoc in spoc_list:
        spoc_projects = frappe.get_all("Project", filters={'status': 'Open', 'service': ['in', ['REC-D', 'REC-I']], 'spoc': spoc}, fields=['*'])
        table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
       
        for project in spoc_projects:
            s_no = 0
            tasks = frappe.get_all("Task", filters={'status': ('in', ['Open', 'Working', 'Overdue', 'Pending Review']), 'project': project['name'], 'service': ('in', ['REC-D', 'REC-I'])}, fields=['name'])
            candidate_count = frappe.db.count("Candidate", filters={'project': project['name'], 'pending_for': ('not in', ['IDB', 'Sourced', 'Proposed PSL'])})
            if candidate_count > 0:
                row = 0
                table+="""<tr style="text-align: center;"><td style="border-left: none; border-right: none;"colspan=10 %s>%s</td></tr>"""%(row+1,project['project_name'])
                table += '<tr style="background-color: #87CEFA"><td style="width: 15%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">CDID</td><td style="width: 25%; font-weight: bold; text-align: center;">Candidate Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Given Name/Surname</td><td style="width: 25%; font-weight: bold; text-align: center;">Passport No</td><td style="width: 25%; font-weight: bold; text-align: center;">Position</td><td style="width: 30%; font-weight: bold; text-align: center;">Candidate Owner</td><td style="width: 40%; font-weight: bold; text-align: center;">Project ID</td><td style="width: 40%; font-weight: bold; text-align: center;">Customer Name</td><td style="width: 25%; font-weight: bold; text-align: center;">Age</td><td style="width: 30%; font-weight: bold; text-align: center;">Next Contact On</td></tr>'
                for task in tasks:
                    s_no += 1
                    candidates = frappe.get_all("Candidate", filters={'pending_for': ('not in', ['IDB', 'Sourced', 'Proposed PSL']), 'task': task['name']}, fields=['name', 'pending_for', 'given_name','passport_number', 'position', 'candidate_created_by', 'project', 'customer', 'age_of_cv', 'custom_next_contact_on'])
                    for candidate in candidates:
                        row += 1
                        table += """<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" % (s_no, candidate['name'], candidate['pending_for'],candidate['given_name'] ,candidate['passport_number'],candidate['position'], candidate['candidate_created_by'], candidate['project'], candidate['customer'], candidate['age_of_cv'], candidate['custom_next_contact_on'] or '')
        table += '</table>'
        subject = "FP List - %s" % nowdate()
        message = """
        Dear Sir/Madam,<br><br>
        Kindly find the below list of your FP List:<br><br>{}<br><br>
        Thanks & Regards,<br>TEAM ERP<br>
        <i>This email has been automatically generated. Please do not reply</i>
        """.format(table)
        frappe.sendmail(
            recipients=[spoc],
            # recipients=["divya.p@groupteampro.com"],
            subject=subject,
            message=message,
        )

@frappe.whitelist() 
def fp_candidate_to_acc_manager():
    projects = frappe.get_all("Project", filters={'status': 'Open', 'service': ['in', ['REC-D', 'REC-I']]}, fields=['name', 'account_manager'])
    spoc_set = set()
    for project in projects:
        if project.get('account_manager'):
            spoc_set.add(project['account_manager'])
    spoc_list = list(spoc_set)
    # table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    print(spoc_list)
    for spoc in spoc_list:
        spoc_projects = frappe.get_all("Project", filters={'status': 'Open', 'service': ['in', ['REC-D', 'REC-I']], 'account_manager': spoc}, fields=['name','project_name'])
        table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
       
        for project in spoc_projects:
            s_no = 0
            tasks = frappe.get_all("Task", filters={'status': ('in', ['Open', 'Working', 'Overdue', 'Pending Review']), 'project': project['name'], 'service': ('in', ['REC-D', 'REC-I'])}, fields=['name'])
            candidate_count = frappe.db.count("Candidate", filters={'project': project['name'], 'pending_for': ('not in', ['IDB', 'Sourced', 'Proposed PSL'])})
            if candidate_count > 0:
                row = 0
                table+="""<tr style="text-align: center;"><td style="border-left: none; border-right: none;"colspan=10 %s>%s</td></tr>"""%(row+1,project['project_name'])
                table += '<tr style="background-color: #87CEFA"><td style="width: 15%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">CDID</td><td style="width: 25%; font-weight: bold; text-align: center;">Candidate Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Given Name/Surname</td><td style="width: 25%; font-weight: bold; text-align: center;">Passport No</td><td style="width: 25%; font-weight: bold; text-align: center;">Position</td><td style="width: 30%; font-weight: bold; text-align: center;">Candidate Owner</td><td style="width: 40%; font-weight: bold; text-align: center;">Project ID</td><td style="width: 40%; font-weight: bold; text-align: center;">Customer Name</td><td style="width: 25%; font-weight: bold; text-align: center;">Age</td><td style="width: 30%; font-weight: bold; text-align: center;">Next Contact On</td></tr>'
                for task in tasks:
                    s_no += 1
                    candidates = frappe.get_all("Candidate", filters={'pending_for': ('not in', ['IDB', 'Sourced', 'Proposed PSL']), 'task': task['name']}, fields=['name', 'pending_for', 'given_name','passport_number', 'position', 'candidate_created_by', 'project', 'customer', 'age_of_cv', 'custom_next_contact_on'])
                    for candidate in candidates:
                        row += 1
                        table += """<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" % (s_no, candidate['name'], candidate['pending_for'], candidate['given_name'],candidate['passport_number'], candidate['position'], candidate['candidate_created_by'], candidate['project'], candidate['customer'], candidate['age_of_cv'], candidate['custom_next_contact_on'] or '')
        table += '</table>'
        subject = "FP List - %s" % nowdate()
        message = """
        Dear Sir/Madam,<br><br>
        Kindly find the below list of your FP List:<br><br>{}<br><br>
        Thanks & Regards,<br>TEAM ERP<br>
        <i>This email has been automatically generated. Please do not reply</i>
        """.format(table)
        frappe.sendmail(
            # recipients=["jothi.m@groupteampro.com"],
            recipients=[spoc],
            subject=subject,
            message=message,
        )


import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill


@frappe.whitelist()
def download():
    filename = "Candidate Details"
    build_xlsx_response(filename)

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
 

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20 
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20 
    ws.column_dimensions['K'].width = 20 
    ws.column_dimensions['L'].width = 20 
    ws.append(["Candidate ID", "PP Number", "Candidate Name", "Qualification", "Total Yrs of Exp",
               "Overseas Exp", "Current Employer", "Current Salary", "Exp. Salary", "Current Location", 
               "Notice Period", "Remarks"])
    data1= get_data(args)
    for row in data1:
        ws.append(row)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file
    
def get_data(args):
    data = []
    candidates = frappe.get_all("Candidate", filters={'custom_date':nowdate(),'candidate_created_by':args.candidate_created_by},
                                fields=["name", "given_name", "candidate_created_by", "pending_for", "customer", 
                                        "project_name", "passport_number", "total_experience", "overseas_experience", 
                                        "current_employer", "current_ctc", "expected_ctc", "location", 
                                        "notice_period_months", "highest_degree", "remarks_1"]) 
    for c in candidates:
        data.append([
                c.name, c.passport_number, c.given_name, c.highest_degree,
                c.total_experience, c.overseas_experience, c.current_employer,
                c.current_ctc, c.expected_ctc, c.location, c.notice_period_months,
                c.remarks_1
            ])
    return data

@frappe.whitelist() 
def send_mail_to_candidate(candidate_id):
    candidate=frappe.db.get_all("Candidate",{'name':candidate_id},['mail_id','territory','position','given_name','interview_location','interview_date','passport_number'])
    candidate_mail=frappe.db.get_value("Candidate",{'name':candidate_id},['mail_id'])
    if candidate:
        candidate = candidate[0]
    subject = " Acknowledgement of Receipt - Original Passport for ACGC, {}  {}".format(candidate['territory'],candidate['position']) 
    message = """
    Dear {},
    <br><br>
    Greetings from TEAMPRO!
    <br><br>
    Following your interview for the position of {} with ACGC, {} in {} on {}, we are delighted to inform you that you have been shortlisted for further consideration.
    <br><br>
    This email serves as confirmation of the receipt of your original passport (Passport No: {}), which we will securely hold until the next stage of the process.
    <br><br>
    Should you have any questions or require further clarification, please don't hesitate to contact us.
    <br><br>
    We appreciate the opportunity to be of service to you.
    <br><br>
    Note: This acknowledgment does not guarantee employment and is subject to confirmation from the client's side.
    <br><br>
    With Best Wishes & Regards,
    <br><br>
    TEAMPRO""".format(candidate['given_name'],candidate['position'],candidate['territory'],candidate['interview_location'],candidate['interview_date'],candidate['passport_number'])
    frappe.sendmail(
        # recipients=[candidate_mail],
        recipients=["sangeetha.a@groupteampro.com"],
        subject=subject,
        message=message,
    )

@frappe.whitelist() 
def send_mail_to_candidate_pass_return(candidate_id):
    candidate=frappe.db.get_all("Candidate",{'name':candidate_id},['mail_id','territory','position','given_name','interview_location','interview_date','passport_number'])
    candidate_mail=frappe.db.get_value("Candidate",{'name':candidate_id},['mail_id'])
    if candidate:
        candidate = candidate[0]
    subject = " Acknowledgement of Returned- Original Passport for ACGC, {}  {}".format(candidate['territory'],candidate['position']) 
    message = """
    Dear {},
    <br><br>
    Greetings from TEAMPRO!
    <br><br>
    Following your interview for the position of {} with ACGC, {} in {} on {}.
    <br><br>
    This email serves as confirmation of the return of your original passport (Passport No: {})
    <br><br>
    Should you have any questions or require further clarification, please don't hesitate to contact us.
    <br><br>
    We appreciate the opportunity to be of service to you.
    <br><br>
    Note: This acknowledgment does not guarantee employment and is subject to confirmation from the client's side.
    <br><br>
    With Best Wishes & Regards,
    <br><br>
    TEAMPRO""".format(candidate['given_name'],candidate['position'],candidate['territory'],candidate['interview_location'],candidate['interview_date'],candidate['passport_number'])
    frappe.sendmail(
        # recipients=[candidate_mail],
        recipients=["sangeetha.a@groupteampro.com"],
        subject=subject,
        message=message,
    )

@frappe.whitelist() 
def send_mail_to_draft(acc_manager,reason,name):
    subject="Project:{} is moved to draft".format(name)
    message="""
        <br>Dear Sir/Mam,<br>
        The Project:{} status is moved to Draft.The following queries need to be clarified.
        <br>Reason:{}
        <br><br>
        With Best Wishes & Regards,
        <br><br>
        TEAMPRO""".format(name,reason)
    frappe.sendmail(
        recipients=acc_manager,
        # recipients=["divya.p@groupteampro.com"],
        subject=subject,
        message=message,
    )

@frappe.whitelist() 
def send_notification_to_am_cofirmed(name):
    subject="Project:{} is conifirmed by Account Manager".format(name)
    message="""
            <br>Dear Sir/Mam,<br>
            The Project:{} is conifirmed by Account Manager for your further action.
            <br><br>
            With Best Wishes & Regards,
            <br><br>
            TEAMPRO""".format(name)
    frappe.sendmail(
        recipients=["sangeetha.a@groupteampro.com"],
        # recipients=["divya.p@groupteampro.com"],
        subject=subject,
        message=message,
    )

@frappe.whitelist() 
def candidate_idb_remarks(candidate_id,reason):
    subject="Candidate:{} is {}".format(candidate_id,reason)
    message="""
            <br>Dear Sir/Mam,<br>
            Candidate:{} is {}
            <br><br>
            With Best Wishes & Regards,
            <br><br>
            TEAMPRO""".format(candidate_id,reason)
    frappe.sendmail(
        recipients=["sangeetha.a@groupteampro.com"],
        # recipients=["divya.p@groupteampro.com"],
        subject=subject,
        message=message,
    )
    # return "ok"

# @frappe.whitelist()
# def candidate_update():
#     value=frappe.db.get_all("Candidate",{'pending_for':'Submitted(Internal)','submitted_date':('<','2024-08-16')},['*'])
#     ind=0
#     for i in value:
#         ind+=1
#     # print(ind)
#         frappe.db.sql("""update `tabCandidate` set pending_for = 'Submitted(Client)' where name = %s""",(i.name))


# @frappe.whitelist()
# def project_name_update():
#     tasks=frappe.db.get_all("Task",{"project":"PROJ-1713"},["name"])
#     count=0
#     for i in tasks:
#         count+=1
#         frappe.db.set_value("Task",i.name,"project_name","ONEIC_16.08.2024")
#         print(i.name)
#     print(count)

# @frappe.whitelist()
# def closure_update():
#     frappe.db.set_value("Closure",{"name":"CL03059"},"status","Visa")

@frappe.whitelist()
def quotation_validate(doc, method):
    if not doc.party_name:
        frappe.throw("You cannot create a Quotation without Opportunity")

@frappe.whitelist()
def download_excel():
    filename = "Candidate Details"
    build_xlsx_response_new(filename)

def build_xlsx_response_new(filename):
    xlsx_file = make_xlsx_new(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_new(data, sheet_name="Candidates", wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    
    if wb is None:
        wb = Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    for col in 'ABCDEFGHIJKL':
        ws.column_dimensions[col].width = 20
    
    ws.append(["Candidate ID", "PP Number", "Candidate Name", "Qualification", "Total Yrs of Exp",
               "Overseas Exp", "Current Employer", "Current Salary", "Exp. Salary", "Current Location", 
               "Notice Period", "Remarks"])
    
    data1 = get_data_new(args)
    for row in data1:
        ws.append(row)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

import json
import frappe

def get_data_new(args):
# def get_data_new():

    # args = {
    #     'cmd': 'jobpro.custom.download_excel', 
    #     'args': '{"custom_date_filter":{"condition":"fiscal year","value":"2024-2025"},"custom_candidate_status_filter":{"condition":"like","value":"%kama%"}}'
    # }
    if args is None:
        args = {}

    if isinstance(args.get('args'), str):
        try:
            args['filters'] = json.loads(args['args'])
        except json.JSONDecodeError as e:
            frappe.log_error(title='JSON Decode Error', message=str(e))
            return []
    
    args['filters'] = args.get('filters', {})

    data = []
    date_filter = args['filters'].get('custom_date_filter')
    candidate_created_by_filter = args['filters'].get('custom_candidate_status_filter')
    candidate_status = args['filters'].get('custom_status_filter')
    # print(candidate_created_by_filter)
    # print(date_filter)
    if not date_filter or not candidate_created_by_filter:
        frappe.log_error(title='Missing filters', message='Date filter or candidate_created_by filter is missing.')
        return data

    filters = {}

    candidate_condition = candidate_created_by_filter.get('condition')
    candidate_value = candidate_created_by_filter.get('value')

    if candidate_condition in ('!=', '=',):
        filters['candidate_created_by'] = [candidate_condition, candidate_value]
    elif candidate_condition in ('like', 'not like') and isinstance(candidate_value, str):
        filters['candidate_created_by'] = [candidate_condition, candidate_value]
    elif candidate_condition == 'is' and candidate_value == 'set':
        filters['candidate_created_by'] = ["is", "set"]
    elif candidate_condition == 'is' and candidate_value == 'not set':
        filters['candidate_created_by'] = ["is", "not set"]
    elif candidate_condition in ('in', 'not in') and isinstance(candidate_value, list):
        filters['candidate_created_by'] = [candidate_condition, candidate_value]
    else:
        return data

    date_condition = date_filter.get('condition')
    date_value = date_filter.get('value')

    if date_condition == 'Between' and isinstance(date_value, list) and len(date_value) == 2:
        filters['submitted_date'] = ['between', [date_value[0], date_value[1]]]
    elif date_condition == 'in' and isinstance(date_value, list):
        filters['submitted_date'] = ['in', date_value]
    elif date_condition == 'not in' and isinstance(date_value, list):
        filters['submitted_date'] = ['not in', date_value]
    elif date_condition == 'is' and date_value == 'set':
        filters['submitted_date'] = ['is', 'set'] 
    elif date_condition == 'is' and date_value == 'not set':
        filters['submitted_date'] = ['is', 'not set'] 
    elif date_condition in ('<', '<=', '>', '>=') and isinstance(date_value, str):
        filters['submitted_date'] = [date_condition, date_value]
    elif date_condition in ('=', '!=') and isinstance(date_value, str):
        filters['submitted_date'] = [date_condition, date_value]  
    elif date_condition == 'Timespan' and isinstance(date_value, str):
        # print(date_value)
        from_date, to_date = get_timespan_custom(date_value)
        filters['submitted_date'] = ['between', [from_date, to_date]]
    elif date_condition == 'fiscal year' and isinstance(date_value, str):
        fiscal_year_start, fiscal_year_end = get_fiscal_year_custom(date_value)
        filters['submitted_date'] = ['between', [fiscal_year_start, fiscal_year_end]]
    else:
        # frappe.log_error(title='Invalid Date Filter', message='Date filter is not set properly.')
        return data
    
    if candidate_status:
        status_condition = candidate_status.get('condition')
        status_value = candidate_status.get('value')
        if status_condition and status_value:
            if status_value =="Submit(SPOC)" or status_value == "Submitted(Client)":
                filters['pending_for'] = [status_condition, status_value]
    

    candidates = frappe.get_all(
        "Candidate", 
        filters=filters,
        fields=[
            "name", "passport_number", "given_name", "highest_degree", 
            "total_experience", "overseas_experience", "current_employer", 
            "current_ctc", "expected_ctc", "location", "notice_period_months", 
            "remarks_1"
        ]
    )
    # index = 0
    for c in candidates:   
        # index+=1
        data.append([
            c.name, c.passport_number, c.given_name, c.highest_degree,
            c.total_experience, c.overseas_experience, c.current_employer,
            c.current_ctc, c.expected_ctc, c.location, c.notice_period_months,
            c.remarks_1
        ])
    # print(index)
    return data


def get_timespan_custom(timespan):
    print(nowdate())
    today = nowdate()
    if timespan == "last week":
        start_date = add_days(today, -7)
        end_date = today
    elif timespan == "last month":
        start_date = add_months(today, -1)
        end_date = today
    elif timespan == "last quarter":
        start_date = add_months(today, -3)
        end_date = today
    elif timespan == "last year":
        start_date = add_months(today, -12)
        end_date = today
    elif timespan == "last 6 months":
        start_date = add_months(today, -6)
        end_date = today
    elif timespan == "today":
        start_date = end_date = today
    elif timespan == "yesterday":
        start_date = end_date = add_days(today, -1)
    elif timespan == "tomorrow":
        start_date = end_date = add_days(today, 1)
    elif timespan == "next month":
        start_date = add_months(today, 1)
        end_date = add_days(add_months(today, 1), -1)
    elif timespan == "next week":
        start_date = today
        end_date = add_days(today, 7)
    elif timespan == "next quarter":
        start_date = today
        end_date = add_months(today, 3)
    elif timespan == "next year":
        start_date = today
        end_date = add_months(today, 12)
    elif timespan == "next 6 months":
        start_date = today
        end_date = add_months(today, 6)
    elif timespan == "this week":
        start_date = get_first_day(today, "week")  
        end_date = add_days(start_date, 6) 
    elif timespan == "this month":
        start_date = get_first_day(today, "month")  
        end_date = get_last_day(today, "month")  
    elif timespan == "last month":
        start_date = add_months(today, -1)
        end_date = today
    elif timespan == "this quarter":
        start_date = get_first_day(today, "quarter")  
        end_date = get_last_day(today, "quarter")  
    elif timespan == "this year":
        start_date = get_first_day(today, "year")  
        end_date = get_last_day(today, "year")
    else:
        raise ValueError(f"Unsupported timespan: {timespan}")
    
    return start_date, end_date

def get_fiscal_year_custom(fiscal_year):
    fiscal_year_split = fiscal_year.split('-')
    start_year = fiscal_year_split[0]
    end_year = fiscal_year_split[1]

    start_date = date(int(start_year), 1, 1) 
    end_date = date(int(end_year), 12, 31)    

    return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')


@frappe.whitelist()
def show_candidate_live_status(project):
    date = nowdate()
    print(date)
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:5%;"><b>Position</b></td>
            <td style="width:5%;"><b>Linedup</b></td>
            <td style="width:5%;"><b>Reported</b></td>
            <td style="width:5%;"><b>Attended</b></td>
            <td style="width:5%;"><b>Proposed PSL</b></td>
        </tr>
    '''
    project_name = frappe.db.get_all("Task", {"project": project}, ["name", "subject"])
    
    for i in project_name:
        # Retrieve the counts
        linedup_count = frappe.db.sql("""
            SELECT COUNT(cs.status) AS status_count
            FROM `tabCandidate` c
            INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
            WHERE c.project = %s
            AND cs.status = %s
            AND c.subject = %s
            AND c.task = %s
            AND DATE(cs.sourced_date) = %s
        """, (project, "Linedup", i.subject, i.name, date), as_dict=True)

        reported_count = frappe.db.sql("""
            SELECT COUNT(cs.status) AS status_count
            FROM `tabCandidate` c
            INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
            WHERE c.project = %s
            AND cs.status = %s
            AND c.subject = %s
            AND c.task = %s
            AND DATE(cs.sourced_date) = %s
        """, (project, "Reported", i.subject, i.name, date), as_dict=True)

        attended_count = frappe.db.sql("""
            SELECT COUNT(cs.status) AS status_count
            FROM `tabCandidate` c
            INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
            WHERE c.project = %s
            AND cs.status = %s
            AND c.subject = %s
            AND c.task = %s
            AND DATE(cs.sourced_date) = %s
        """, (project, "Attended", i.subject, i.name, date), as_dict=True)

        ppl_count = frappe.db.sql("""
            SELECT COUNT(cs.status) AS status_count
            FROM `tabCandidate` c
            INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
            WHERE c.project = %s
            AND cs.status = %s
            AND c.subject = %s
            AND c.task = %s
            AND DATE(cs.sourced_date) = %s
        """, (project, "Proposed PSL", i.subject, i.name, date), as_dict=True)

        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
            i.subject,
            linedup_count[0].status_count if linedup_count else '0',
            reported_count[0].status_count if reported_count else '0',
            attended_count[0].status_count if attended_count else '0',
            ppl_count[0].status_count if ppl_count else '0'
        )
    
    data += '</table>'
    return data

@frappe.whitelist()
def show_candidate_live_status_total(project):
    date=nowdate()
    print(date)
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:4.25%;"><b>Linedup</b></td>
            <td style="width:4.25%;"><b>Reported</b></td>
            <td style="width:4.25%;"><b>Attended</b></td>
            <td style="width:4.25%;"><b>Proposed PSL</b></td>
        </tr>
    '''
    linedup_count = frappe.db.sql("""
        SELECT COUNT(cs.status) AS status_count
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.project = %s
        AND cs.status = %s
        AND DATE(cs.sourced_date) = %s
    """, (project, "Linedup",date), as_dict=True)

    reported_count = frappe.db.sql("""
        SELECT COUNT(cs.status) AS status_count
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.project = %s
        AND cs.status = %s
        AND DATE(cs.sourced_date) = %s
    """, (project, "Reported",date), as_dict=True)
    attended_count = frappe.db.sql("""
        SELECT COUNT(cs.status) AS status_count
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.project = %s
        AND cs.status = %s
        AND DATE(cs.sourced_date) = %s
    """, (project, "Attended",date), as_dict=True)
    ppl_count = frappe.db.sql("""
        SELECT COUNT(cs.status) AS status_count
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.project = %s
        AND cs.status = %s
        AND DATE(cs.sourced_date) = %s
    """, (project, "Attended",date), as_dict=True)
    data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(linedup_count[0].status_count if linedup_count else '0',reported_count[0].status_count if reported_count else '0',attended_count[0].status_count if attended_count else '0',ppl_count[0].status_count if ppl_count else '0')
    data += '</table>'
    return data
@frappe.whitelist()
def move_to_sfp():
    tot=0
    sfp=frappe.db.get_all("Sales Follow Up",{"lead":['!=',''],'creation':['between',('2023-12-01','2023-12-31')]},['*'])
    for i in sfp:
        sales=frappe.get_doc("Sales Follow Up",{'name':i.name})
        count=0
        for s in sales.contacts:
            count+=1
        if count <= 0:
            tot+=1
            print(i.lead)
            # if frappe.db.exists("Lead",{'name':i.lead}):
            lead=frappe.get_doc("Lead",{'name':i.lead})
            print("lead")
            for k in lead.lead_contacts:
                print(sales.name)
                print(k)
                sales.append('contacts',{
                'person_name': k.person_name,
                'mobile':k.mobile,
                'is_primary':k.is_primary,
                'email_id':k.email_id,
                'is_primaryemail':k.is_primaryemail,
                'has_whatsapp':k.has_whatsapp,
                'service':k.service
                })
            sales.save()
    # print(tot)
@frappe.whitelist()
def change_visited_status(lead=None, completed=None, taken=None):
    lead_doc = frappe.get_doc("Lead", lead)
    lead_doc.visit_status = 'Visited'
    lead_doc.visit_date = completed
    lead_doc.custom_visited_by = taken
    lead_doc.save()
    frappe.db.commit() 
    
@frappe.whitelist()
def update_sfp():
    count=0
    doc=frappe.db.get_all("Sales Follow Up",{'follow_up_to':'Lead','lead':'','organization_name':['!=','']},['*'])
    for i in doc:
        if frappe.db.exists("Lead",{'company_name':i.organization_name}):
            count+=1
            lname=frappe.db.get_value("Lead",{'company_name':i.organization_name},['name'])
            frappe.db.set_value("Sales Follow Up",i.name,'lead',lname)
    return count

@frappe.whitelist()
def update_address_details(doc, method):
    if doc.appointment_with and doc.party:
        if doc.appointment_with == 'Lead':
            ldoc = frappe.db.get_doc("Lead", doc.party)
            if ldoc:
                for contact in ldoc.lead_contacts:
                    if contact.is_primary == 1:
                        if contact.mobile:
                            frappe.db.set_value('Appointment', doc.name, 'customer_phone_number', contact.mobile)
                        break  

@frappe.whitelist()
def update_sfp_type():
    sales=frappe.db.get_all("Sales Follow Up",{'follow_up_to':"Customer"},["*"])
    count=0
    for i in sales:
        owner=frappe.db.get_value("Customer",{'name':i.customer},['account_manager'])
        frappe.db.set_value("Sales Follow Up",i.name,'account_manager_lead_owner',owner)
        frappe.db.set_value("Sales Follow Up",i.name,'party_from',"Customer")
        frappe.db.set_value("Sales Follow Up",i.name,'party_name',i.customer)
        count+=1
    print(count)

@frappe.whitelist()
def app_team_dpr_excel_format():
    filename = "APP_Team_DPR" + today() + ".xlsx"
    # file_path = make_xlsx_for_app_team_dpr(filename)
    # with open(file_path, "rb") as file_content:
    #     app_team_dpr_mail(filename, file_content.read())
    xlsx_file = make_xlsx_for_app_team_dpr(filename)
    app_team_dpr_mail(filename, xlsx_file.getvalue())

def app_team_dpr_mail(filename, file_content):
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00023','user_id':('not in',['sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'])},['*'])
    recievers=[]
    reciever=[]
    custom_date = today()
    date_obj = datetime.strptime(str(custom_date), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date= today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    print(formatted_next_date)
    data=[]
    for j in emp:
        reciever.append(j.user_id)
    for i in emp:
        recievers.append(i.user_id)
    recievers.append('anil.p@groupteampro.com')
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="9"><b>APP & Team DPR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Appointments</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Opportunity</b></td>
            <td style="width:13%;"><b>Customer</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''

    appointment_count = 0
    appointment_list = []
    app_individual=[]
    todo_list_i=[]
    todo_i=[]
    todos=[]
    todo_list=[]
    for user_email in recievers:
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s
        """, (user_email, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)[0].count or 0
        
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Customer"})
        
        # appointment_count = frappe.db.count("Appointment",{"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":user_email})
        todo_count=frappe.db.count("ToDo",{"allocated_to":user_email,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , lead_count if lead_count else '0' , open_count if open_count else '0', replied_count if replied_count else '0', interested_count if interested_count else '0', opportunity_count if opportunity_count else '0', customer_count if customer_count else'0',todo_count if todo_count else '0'
            )
        appt = frappe.db.sql("""select p.scheduled_time as time, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where date(p.scheduled_time) between '%s' and '%s' and c.user='%s'""" %(formatted_next_date, formatted_next_date,user_email),as_dict=1)
        appointment_list.append(appt)

        todos=frappe.db.get_all("ToDo",{"custom_production_date":formatted_next_date,"allocated_to":user_email},["*"])
        todo_list.append(todos)
    if todo_list:
        data += '''
              <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
        '''
        s_no=0
        for todo_group in todo_list:
            for i in todo_group:
                user_email=i['allocated_to']
                s_no+=1
                short_code = frappe.db.get_value("Employee", {"user_id": i.allocated_to}, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="6" style="text-align: left; padding-left: 50px;">{}</td><td colspan="1">{}</td></tr>'.format(short_code,i.name,i.custom_subject,i.status)
    if appointment_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>Appointment</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Customer</td></tr>
        '''
        for appt_group in appointment_list:
            for i in appt_group:  # each 'i' is a dictionary with appointment details
                user_email = i['user']
                short_code = frappe.db.get_value("Employee", {"user_id":i.user }, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code,i.time,i.name)
            
    data += '</table>'



#     frappe.sendmail(
#                 # recipients=recievers,
#                 recipients=['divya.p@groupteampro.com'],
#                 # recipients=['anil.p@groupteampro.com'],
#                 # cc='dineshbabu.k@groupteampro.com',
#                 subject='APP & Team DPR %s -Reg' % formatted_date,
#                 message = """
#                 <b>Dear Team,</b><br><br>
# Please find the below DPR for {} for your kind reference and action.<br><br>

#             {}<br><br>
#                 Thanks & Regards,<br>TEAM ERP<br>
                
#                 <i>This email has been automatically generated. Please do not reply</i>
#                 """.format(formatted_date,data),
#                 attachments=[
#             {"fname": filename, "fcontent": file_content}
#             ]
#             )
    for user_i in reciever:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="text-align:center;"><td colspan="8"><b>APP & Team DPR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Appointments</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Opportunity</b></td>
                <td style="width:13%;"><b>Customer</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
            </tr>
        '''

        # Fetch appointment counts for each user
        # appointments = frappe.get_all("Appointment", filters={"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]}, fields=["*"])
        # appointment_count = frappe.db.count("Appointment", {"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]], "owner": user_email})
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user = %s AND p.scheduled_time BETWEEN %s AND %s
    """, (user_i, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)[0].count or 0
        # Other counts
        short_code = frappe.db.get_value("Employee", {"user_id": user_i}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_i, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Lead"})
        open_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_i, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Open"})
        replied_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_i, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Replied"})
        interested_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_i, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Interested"})
        opportunity_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_i, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Opportunity"})
        customer_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_i, "next_contact_date": formatted_next_date, "follow_up_to": "Customer"})
        todo_count = frappe.db.count("ToDo", {"allocated_to": user_i, "custom_production_date": formatted_next_date})

        # Populate table row for the specific user
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
            short_code, appointment_count, lead_count, open_count, replied_count, interested_count, opportunity_count, customer_count, todo_count
        )
        app_i= frappe.db.sql("""select p.scheduled_time as time, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where c.user='%s' and date(p.scheduled_time) between '%s' and '%s'""" %(user_i,formatted_next_date, formatted_next_date),as_dict=1)
        app_individual.append(app_i)
        
        # ToDo Section
        todo_i = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": user_i}, ["*"])
        todo_list_i.append(todo_i)
        if todo_list_i:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>ToDo</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
            '''
            for todo_group in todo_list:
                for todo in todo_group:
                    user_i=todo['allocated_to']

                    short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject)

        # Appointment Section
        print(user_i)

        if app_individual:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>Appointment</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Customer</td></tr>
            '''
            for appt_group in app_individual:
                for i in appt_group:  # each 'i' is a dictionary with appointment details
                    user_i = i['user']
                    short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, i.time, i.name)

        data += '</table>'
        # frappe.sendmail(
        #     # recipients=[user_emails],
        #     recipients=['divya.p@groupteampro.com'],
        #     subject='APP & Team DPR %s - Reg' % formatted_date,
        #     message="""
        #         <b>Dear {user},</b><br><br>
        #         Please find the below DPR for {date} for your kind reference and action.<br><br>
        #         {table}<br><br>
        #         Thanks & Regards,<br>TEAM ERP<br>
        #         <i>This email has been automatically generated. Please do not reply</i>
        #     """.format(user=user_email.split('@')[0], date=formatted_date, table=data),
        #     attachments=[
        #     {"fname": filename, "fcontent": file_content},
        # ]
        # )


import frappe
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from frappe.utils import today
from datetime import datetime

def make_xlsx_for_app_team_dpr(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "APP Team DPR"
    
    # Column headers
    headers = ["Exe", "Appointments", "Lead", "Open", "Replied", "Interested", "Opportunity", "Customer", "ToDo"]
    ws.append(["APP & Team DPR"])
    ws.append(headers)
    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    sub_header_fill=PatternFill(start_color="87cefa", end_color="87cefa", fill_type="solid")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    # Set column widths
    default_column_widths = [10, 15, 10, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(default_column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = black_border

    # Initialize Date
    formatted_date = datetime.strptime(today(), '%Y-%m-%d').strftime('%d/%m/%Y')
    formatted_next_date = today()

    # Fetch employee list
    emp = frappe.db.get_all("Employee", {'status': 'Active', 'reports_to': 'TI00023', 'user_id': ('not in', ['sivarenisha.m@groupteampro.com', 'jeniba.a@groupteampro.com'])}, ['*'])
    recievers = [employee.user_id for employee in emp]
    recievers.append("anil.p@groupteampro.com")

    # DPR Data
    for user_email in recievers:
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        appointment_count = frappe.db.sql("""SELECT COUNT(DISTINCT p.name) AS count FROM `tabAppointment` p INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s""", (user_email, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)[0].count or 0
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Lead"})
        open_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Open"})
        replied_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Replied"})
        interested_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Interested"})
        opportunity_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Opportunity"})
        customer_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Customer"})
        todo_count = frappe.db.count("ToDo", {"allocated_to": user_email, "custom_production_date": formatted_next_date})

        # Add employee's DPR data row
        ws.append([short_code, appointment_count, lead_count, open_count, replied_count, interested_count, opportunity_count, customer_count, todo_count])

    # Add section for ToDo
    ws.append(["ToDo"])
    for cell in ws[ws.max_row]:  # Apply to ToDo header row
        cell.fill = header_fill
        cell.font = header_font
        cell.border = black_border

    ws.append(["Exe", "ID", "TODO", "Status"])
    for cell in ws[ws.max_row]:  # Apply to ToDo header row
        cell.fill = sub_header_fill
        cell.font = header_font
        cell.border = black_border
    # Retrieve ToDo data and add rows
    for user_emails in recievers:
        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": user_emails}, ["name", "custom_subject", "status"])
        short_code = frappe.db.get_value("Employee", {"user_id": user_emails}, "short_code")
        for todo in todo_list:
            ws.append([short_code, todo.name, todo.custom_subject, todo.status])

    # Add section for Appointments
    ws.append(["Appointment"])
    for cell in ws[ws.max_row]:  # Apply to Appointment header row
        cell.fill = header_fill
        cell.font = header_font
        cell.border = black_border

    ws.append(["Exe", "ID", "Csutomer"])
    for cell in ws[ws.max_row]:  # Apply to ToDo header row
        cell.fill = sub_header_fill
        cell.font = header_font
        cell.border = black_border
    # Retrieve Appointment data and add rows
    for user_emails in recievers:
        appointment_list = frappe.db.sql("""SELECT p.scheduled_time AS time, p.name AS name FROM `tabAppointment` p INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s""", (user_email, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)
        short_code = frappe.db.get_value("Employee", {"user_id": user_emails}, "short_code")
        for appointment in appointment_list:
            ws.append([short_code, appointment.name, appointment.customer_name])

    # Save workbook
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

@frappe.whitelist()
def get_contact_details(party_name,party_type,name):
    cont=''
    email=''
    person=''
    if party_type=='Sales Follow Up':
        sfp=frappe.get_doc('Sales Follow Up',party_name)
        for contact in sfp.contacts:
            if contact.is_primary == 1:
                cont=contact.mobile
                email=contact.is_primaryemail
                person=contact.person_name
    return {
        'cont': cont,
        'email': email,
        'person': person
    }

@frappe.whitelist()
def create_appointment(email,customer,taken,schedule,name,sfp_type,sfp_name):
    app=frappe.new_doc("Appointment")
    app.custom_appointment_from='Sales Follow Up'
    app.scheduled_time=schedule
    app.customer_name=customer
    app.customer_email=email
    # taken=list(taken)
    
    app.append("custom_appointment_to_be_created_by",{
        'user':taken
    })
    # app.custom_appointment_to_be_created_by=list(taken)
    if sfp_type=="Lead":
        app.appointment_with='Lead'
        app.party=sfp_name
    if sfp_type=="Customer":
        app.appointment_with='Customer'
        app.party=sfp_name
    sfp=frappe.get_doc("Sales Follow Up",name)
    for s in sfp.contacts:
        if s.is_primary == 1:
            if s.mobile:
                app.customer_phone_number=s.mobile
    # app.insert()
    app.save(ignore_permissions=True)
    return "OK"



@frappe.whitelist()
def create_sale_order_in_bulk(closure):
    doc_name = json.loads(closure)
    closure_candidate_docs = []
    closure_client_docs=[]
    closure_associate_docs=[]
    closure_both_docs=[]
    closure_both_doc=[]
    for j in doc_name:
        closure_doc = frappe.get_doc("Closure", j)
        parent_territory = frappe.get_value('Territory', closure_doc.territory, 'parent_territory')
        items = []
        if closure_doc.payment:
            item_candidate_id = frappe.db.get_value("Item", {"name": closure_doc.mobile})
            item_pp_id = frappe.db.get_value("Item", {"name": closure_doc.passport_no})
            
            if item_candidate_id or item_pp_id:
                item = frappe.get_doc("Item", item_pp_id)
            else:
                item = frappe.new_doc("Item")
                if parent_territory == 'India':
                    item.item_code = closure_doc.given_name
                    item.append("taxes", {
                        "item_tax_template": "T - GST @ 18% - THIS",
                        "tax_category": "Professional Service - GST",
                        "valid_from": today()
                    })
                else:
                    item.item_code = closure_doc.passport_no
                    item.is_non_gst = "0"
                item.item_name = closure_doc.passport_no + ":" + closure_doc.given_name
                if closure_doc.candidate_owner:
                    item.candidate_owner = closure_doc.candidate_owner
                if closure_doc.sa_id:
                    item.sa_id = closure_doc.sa_id
                item.item_group = "Candidates"
                item.stock_uom = "Nos"
                item.qty = "1"
                item.gst_hsn_code = '998519'
                item.is_stock_item = "0"
                item.include_item_in_manufacturing = "0"
                item.description = closure_doc.customer
                item.append("item_defaults", {
                    "company": "TeamPRO HR & IT Services Pvt. Ltd."
                })
                item.append("customer_items", {
                    "customer_name": closure_doc.customer,
                    "ref_code": closure_doc.mobile
                })
                item.insert()
            if closure_doc.territory != 'India' or parent_territory != 'India':
                if closure_doc.payment == "Candidate":
                    items.append({
                        "item_code": item.item_code,
                        "item_name": item.item_name,
                        "candidate_owner": item.candidate_owner,
                        "sa_id": item.sa_id,
                        "payment_type": "Candidate",
                        "description": item.description,
                        "uom": item.stock_uom,
                        "is_stock_item": "0",
                        "passport_no": closure_doc.passport_no,
                        "delivery_date": closure_doc.expected_doj or '',
                        "qty": "1",
                        "rate": closure_doc.candidate_si,
                        "sc1": closure_doc.candidate_service_charge,
                        "cost_center": "Main - THIS",
                    })
                    closure_candidate_docs.append((closure_doc, items))
                elif closure_doc.payment == "Client":
                    items.append({
                        "item_code": item.item_code,
                        "item_name": item.item_name,
                        "candidate_owner": item.candidate_owner,
                        "sa_id": item.sa_id,
                        "payment_type": "Candidate",
                        "description": item.description,
                        "uom": item.stock_uom,
                        "is_stock_item": "0",
                        "passport_no": closure_doc.passport_no,
                        "delivery_date": closure_doc.expected_doj or '',
                        "qty": "1",
                        "rate": closure_doc.client_si,
                        # "sc1": closure_doc.client_sc or '',
                        "cost_center": "Main - THIS",
                    })
                    closure_client_docs.append((closure_doc, items))
                elif closure_doc.payment == "Associate":
                    items.append({
                        "item_code": item.item_code,
                        "item_name": item.item_name,
                        "candidate_owner": item.candidate_owner,
                        "sa_id": item.sa_id,
                        "description": item.description,
                        "uom": item.stock_uom,
                        "is_stock_item": "0",
                        "passport_no": closure_doc.passport_no,
                        "delivery_date": closure_doc.expected_doj or '',
                        "qty": "1",
                        "rate": closure_doc.associate_si,
                        "cost_center": "Main - THIS",
                    })
                    closure_associate_docs.append((closure_doc, items))
                elif closure_doc.payment == "Both":
                    items.append({
                        "item_code": item.item_code,
                        "item_name": item.item_name,
                        "candidate_owner": item.candidate_owner,
                        "sa_id": item.sa_id,
                        "description": item.description,
                        "uom": item.stock_uom,
                        "is_stock_item": "0",
                        "passport_no": closure_doc.passport_no,
                        "delivery_date": closure_doc.expected_doj or '',
                        "qty": "1",
                        "rate": closure_doc.client_si,
                        "payment_type": "Candidate",
                        # "sc1": closure_doc.client_sc,
                        "cost_center": "Main - THIS",
                    })
                    closure_both_docs.append((closure_doc, items))
                    items.append({
                        "item_code": item.item_code,
                        "item_name": item.item_name,
                        "candidate_owner": item.candidate_owner,
                        "sa_id": item.sa_id,
                        "description": item.description,
                        "payment_type": "Candidate",
                        "uom": item.stock_uom,
                        "is_stock_item": "0",
                        "passport_no": closure_doc.passport_no,
                        "delivery_date": closure_doc.expected_doj or '',
                        "qty": "1",
                        "rate": closure_doc.candidate_si,
                        "sc1": closure_doc.candidate_service_charge,
                        "cost_center": "Main - THIS",
                    })
                    closure_both_doc.append((closure_doc, items))
    if closure_candidate_docs:
        candidate_customer = frappe.new_doc("Customer")
        candidate_customer.customer_name = closure_doc.given_name + '-' + closure_doc.passport_no
        candidate_customer.customer_type = "Individual"
        candidate_customer.customer_group = "Individual"
        candidate_customer.territory = closure_doc.territory
        candidate_customer.insert()
        candidate_customer.save(ignore_permissions=True)
        frappe.db.commit()
        sales_order = frappe.new_doc("Sales Order")
        first_closure_doc, items = closure_candidate_docs[0]
        sales_order.customer = first_closure_doc.customer
        sales_order.reference_customer_ = first_closure_doc.customer
        sales_order.customer_group = "Individual"
        sales_order.passport_number = first_closure_doc.passport_no
        sales_order.account_manager = first_closure_doc.account_manager
        sales_order.delivery_manager = first_closure_doc.candidate_owner
        sales_order.service = first_closure_doc.service
        sales_order.currency = "INR"
        sales_order.company_address = "TEAMPRO HR & IT Services Pvt. Ltd."
        for closure_doc, items in closure_candidate_docs:
            for item in items:
                sales_order.append("items", item)
        sales_order.save(ignore_permissions=True)
        frappe.db.commit()
        for closure_doc, _ in closure_candidate_docs:
            frappe.set_value('Closure', closure_doc.name, 'so_created', 1)
            frappe.set_value('Closure', closure_doc.name, 'so_confirmed_date', today())
    if closure_client_docs:
        sales_order = frappe.new_doc("Sales Order")
        first_closure_doc, items = closure_client_docs[0]
        sales_order.customer = first_closure_doc.customer
        sales_order.reference_customer_ = first_closure_doc.customer
        sales_order.passport_number = first_closure_doc.passport_no
        sales_order.account_manager = first_closure_doc.account_manager
        sales_order.delivery_manager = first_closure_doc.candidate_owner
        sales_order.closure_project=first_closure_doc.project
        sales_order.supplier=first_closure_doc.sa_id
        sales_order.service = first_closure_doc.service
        sales_order.currency = first_closure_doc.billing_currency
        sales_order.company_address = "TEAMPRO HR & IT Services Pvt. Ltd."
        for closure_doc, items in closure_client_docs:
            for item in items:
                sales_order.append("items", item)
        sales_order.save(ignore_permissions=True)
        frappe.db.commit()
        for closure_doc, _ in closure_client_docs:
            frappe.set_value('Closure', closure_doc.name, 'so_created', 1)
            frappe.set_value('Closure', closure_doc.name, 'so_confirmed_date', today())

    if closure_associate_docs:
        sales_order = frappe.new_doc("Sales Order")
        first_closure_doc, items = closure_associate_docs[0]
        sales_order.customer = first_closure_doc.associate
        sales_order.reference_customer_ = first_closure_doc.customer
        sales_order.customer_group="Associate"
        sales_order.service = first_closure_doc.service
        sales_order.company_address = "TEAMPRO HR & IT Services Pvt. Ltd."
        for closure_doc, items in closure_associate_docs:
            for item in items:
                sales_order.append("items", item)
        sales_order.save(ignore_permissions=True)
        frappe.db.commit()
        for closure_doc, _ in closure_associate_docs:
            frappe.set_value('Closure', closure_doc.name, 'so_created', 1)
            frappe.set_value('Closure', closure_doc.name, 'so_confirmed_date', today())

    if closure_both_docs:
        # Create a Sales Order for "Client"
        sales_order_client = frappe.new_doc("Sales Order")
        sales_order_client.customer = closure_both_docs[0][0].customer
        sales_order_client.reference_customer_ = closure_both_docs[0][0].customer
        sales_order_client.currency = closure_both_docs[0][0].billing_currency
        sales_order_client.delivery_manager = closure_both_docs[0][0].candidate_owner
        sales_order_client.passport_number = closure_both_docs[0][0].passport_no
        sales_order_client.account_manager = closure_both_docs[0][0].account_manager
        sales_order_client.service = closure_both_docs[0][0].service
        sales_order_client.company_address = "TEAMPRO HR & IT Services Pvt. Ltd."

        # Add unique items for "Client"
        added_items_client = set()
        for closure_doc, items in closure_both_docs:
            for item in items:
                if item["item_code"] not in added_items_client:
                    sales_order_client.append("items", item)
                    added_items_client.add(item["item_code"])
        
        sales_order_client.save(ignore_permissions=True)
        frappe.db.commit()
        for closure_doc, _ in closure_both_docs:
            frappe.set_value('Closure', closure_doc.name, 'so_created', 1)
            frappe.set_value('Closure', closure_doc.name, 'so_confirmed_date', today())

        added_customers = set()  
        for closure_doc, items in closure_both_docs:
            candidate_name = f"{closure_doc.given_name}-{closure_doc.passport_no}"
            if candidate_name not in added_customers:
                candidate_customer = frappe.new_doc("Customer")
                candidate_customer.customer_name = candidate_name
                candidate_customer.customer_type = "Individual"
                candidate_customer.customer_group = "Individual"
                candidate_customer.territory = closure_doc.territory
                candidate_customer.insert(ignore_permissions=True)
                frappe.db.commit()
                added_customers.add(candidate_name)
        # Create a Sales Order for "Candidate"
        sales_order_candidate = frappe.new_doc("Sales Order")
        sales_order_candidate.customer = closure_both_docs[0][0].customer
        sales_order_candidate.reference_customer_ = closure_both_docs[0][0].customer
        sales_order_candidate.currency = "INR"
        sales_order_candidate.delivery_manager = closure_both_docs[0][0].candidate_owner
        sales_order_candidate.passport_number = closure_both_docs[0][0].passport_no
        sales_order_candidate.account_manager = closure_both_docs[0][0].account_manager
        sales_order_candidate.customer_group = "Individual"
        sales_order_candidate.service = closure_both_docs[0][0].service
        sales_order_candidate.company_address = "TEAMPRO HR & IT Services Pvt. Ltd."

        # Add unique items for "Candidate"
        added_items_candidate = set()
        for closure_doc, items in closure_both_doc:
            for item in items:
                if item["item_code"] not in added_items_candidate:
                    sales_order_candidate.append("items", item)
                    added_items_candidate.add(item["item_code"])
        
        sales_order_candidate.save(ignore_permissions=True)
        frappe.db.commit()
        for closure_doc, _ in closure_both_docs:
            frappe.set_value('Closure', closure_doc.name, 'so_created', 1)
            frappe.set_value('Closure', closure_doc.name, 'so_confirmed_date', today())
            frappe.set_value('Closure', closure_doc.name, 'status', 'Visa')
    
# Define a method to calculate the total outstanding amount
@frappe.whitelist(allow_guest=True)
def get_sales_invoice_outstanding():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabSales Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Return', 'Credit Note Issued') AND company='TEAMPRO HR & IT Services Pvt. Ltd.'
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

# Define a method to calculate the total outstanding amount
@frappe.whitelist(allow_guest=True)
def get_sales_invoice_outstanding_tfp():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabSales Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Return', 'Credit Note Issued') AND company='TEAMPRO Food Products'
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

# Define a method to calculate the total outstanding amount
@frappe.whitelist(allow_guest=True)
def get_sales_invoice_outstanding_tgt():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabSales Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Return', 'Credit Note Issued') AND company='TEAMPRO General Trading'
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

# Define a method to calculate the total outstanding amount
@frappe.whitelist(allow_guest=True)
def get_sales_invoice_outstanding_tot():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabSales Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Return', 'Credit Note Issued')
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_pi_outstanding_this():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabPurchase Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Debit Note Issued','Return') AND company='TEAMPRO HR & IT Services Pvt. Ltd.'
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_pi_outstanding_tfp():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabPurchase Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Debit Note Issued','Return') AND company='TEAMPRO Food Products'
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_pi_outstanding_tgt():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabPurchase Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Debit Note Issued','Return') AND company='TEAMPRO General Trading'
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_pi_outstanding_tot():
    total_outstanding = frappe.db.sql("""
        SELECT SUM(outstanding_amount) 
        FROM `tabPurchase Invoice`
        WHERE status NOT IN ('Paid', 'Cancelled', 'Debit Note Issued','Return')
    """, as_dict=True)
    return {
        "value": total_outstanding[0].get("SUM(outstanding_amount)", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_so_outstanding_this():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - ((base_grand_total * amount_billed) + advance_paid)) AS total_outstanding
        FROM `tabSales Order`
        WHERE status NOT IN ('Closed', 'Cancelled', 'On Hold', 'Completed', 'To Deliver') AND company='TEAMPRO HR & IT Services Pvt. Ltd.'
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_so_outstanding_tfp():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - ((base_grand_total * amount_billed) + advance_paid)) AS total_outstanding
        FROM `tabSales Order`
        WHERE status NOT IN ('Closed', 'Cancelled', 'On Hold', 'Completed', 'To Deliver') AND company='TEAMPRO Food Products'
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_so_outstanding_tgt():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - ((base_grand_total * amount_billed) + advance_paid)) AS total_outstanding
        FROM `tabSales Order`
        WHERE status NOT IN ('Closed', 'Cancelled', 'On Hold', 'Completed', 'To Deliver') AND company='TEAMPRO General Trading'
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_so_outstanding_tot():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - ((base_grand_total * amount_billed) + advance_paid)) AS total_outstanding
        FROM `tabSales Order`
        WHERE status NOT IN ('Closed', 'Cancelled', 'On Hold', 'Completed', 'To Deliver')
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_po_outstanding_this():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - advance_paid) AS total_outstanding
        FROM `tabPurchase Order`
        WHERE status NOT IN ('On Hold', 'Cancelled', 'To Receive', 'Completed', 'Closed') AND company='TEAMPRO HR & IT Services Pvt. Ltd.'
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_po_outstanding_tfp():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - advance_paid) AS total_outstanding
        FROM `tabPurchase Order`
        WHERE status NOT IN ('On Hold', 'Cancelled', 'To Receive', 'Completed', 'Closed') AND company='TEAMPRO Food Products'
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_po_outstanding_tgt():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - advance_paid) AS total_outstanding
        FROM `tabPurchase Order`
        WHERE status NOT IN ('On Hold', 'Cancelled', 'To Receive', 'Completed', 'Closed') AND company='TEAMPRO General Trading'
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }

@frappe.whitelist(allow_guest=True)
def get_po_outstanding_tot():
    # Query to calculate the total outstanding based on the formula
    total_outstanding = frappe.db.sql("""
        SELECT 
            SUM(base_grand_total - advance_paid) AS total_outstanding
        FROM `tabPurchase Order`
        WHERE status NOT IN ('On Hold', 'Cancelled', 'To Receive', 'Completed', 'Closed')
    """, as_dict=True)

    # Return the computed outstanding amount in the required format
    return {
        "value": total_outstanding[0].get("total_outstanding", 0) if total_outstanding else 0,
        "fieldtype": "Currency",  # Specify that the value is in currency format
        "route_options": {},  # Optional, any extra parameters for routing
        "route": ["query-report", "Permitted Documents For User"]  # Optional route for additional data
    }


@frappe.whitelist()
def test_download():
    filename = "Candidate Details"
    build_xlsx_response_test(filename)

def build_xlsx_response_test(filename):
    xlsx_file = make_xlsx_test(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
 

def make_xlsx_test(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    # Set column widths
    for col in range(ord('A'), ord('M')):  # Columns A to L
        ws.column_dimensions[chr(col)].width = 20

    # Define headers
    headers = ["Candidate ID", "PP Number", "Candidate Name", "Qualification", 
               "Total Yrs of Exp", "Overseas Exp", "Current Employer", 
               "Current Salary", "Exp. Salary", "Current Location", 
               "Notice Period", "Remarks"]

    # Define styles
    position_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    position_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="98D7F5", end_color="98D7F5", fill_type="solid")
    header_font = Font(bold=True)
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    # Fetch candidate data grouped by positions
    position_candidates = get_data_grouped_by_position(args)

    for position, candidates in position_candidates.items():
        # Add position row
        position_row = ws.max_row + 1
        ws.merge_cells(start_row=position_row, start_column=1, end_row=position_row, end_column=12)
        cell = ws.cell(row=position_row, column=1)
        cell.value = f"{position}"
        cell.fill = position_fill
        cell.font = position_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

        # Add headers
        header_row = ws.max_row + 1
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = black_border

        # Add details for the position
        for candidate in candidates:
            # ws.append(candidate)
            row_num = ws.max_row + 1
            for col_num, value in enumerate(candidate, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = black_border  # Apply border to each cell

        # Add an empty row for separation
        ws.append([])

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_grouped_by_position(args):
    data = {}
    candidates = frappe.get_all(
        "Candidate",
        filters={'submitted_date': nowdate(), 'candidate_created_by': args.candidate_created_by},
        fields=["name", "passport_number", "given_name", "highest_degree",
                "total_experience", "overseas_experience", "current_employer",
                "current_ctc", "expected_ctc", "location", "notice_period_months",
                "remarks_1", "position","currency_ctc"]
    )
    
    for candidate in candidates:
        position = candidate.get("position", "")
        currency=candidate.currency_ctc
        formatted_ctc = f"{currency} {candidate.current_ctc}" if candidate.current_ctc else "0"
        if position not in data:
            data[position] = []
        data[position].append([
            candidate.name, candidate.passport_number, candidate.given_name,
            candidate.highest_degree, candidate.total_experience, 
            candidate.overseas_experience, candidate.current_employer,
            formatted_ctc, candidate.expected_ctc, candidate.location, 
            candidate.notice_period_months, candidate.remarks_1
        ])

    return data



@frappe.whitelist()
def update_sa_details_in_task(doc,method):
    if doc.name and doc.service=="REC-I":
        tasks = frappe.db.get_all("Task", filters={"project":doc.name}, fields=["name"])
        # pro=frappe.get_doc("Project",project)
        for task in tasks:
            task_doc = frappe.get_doc("Task", task["name"])
            existing_sa_ids = {i.sa_id for i in task_doc.sa_detail}
            for row in doc.custom_sa_details:
                if row.sa_id not in existing_sa_ids:
                    task_doc.append(
                        "sa_detail",
                        {
                            "sa_id": row.sa_id,
                            "sa_name": row.sa_name,
                            "sa_contact_number": row.sa_contact_number,
                            "variable_ec": row.variable_ec,
                            "received_count": row.received_count,
                        },
                    )

            # Save the task document to persist changes
            task_doc.save()
            frappe.db.commit()
            task_doc.reload()
        



@frappe.whitelist()
def send_closure_mail():
    current_date = datetime.now().strftime("%d-%m-%Y")
    ind=0
    s_no=1
    data = '<table  text-align: center; border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    data += '<tr style="font-weight: bold;background-color: #98d7f5;"><td width=5%>S.No</td><td width=15%>Closure ID</td><td width=25%>Candidate Name</td><td width=15%>Passport Number</td><td width=25%>Customer Name</td><td width=25%>Status</td><td width=35%>Latest Remarks</td></tr>'
    closure=frappe.db.get_all("Closure",{"status":["in",["PSL","Signed Offer Letter","Premedical","PCC","Final Medical","Biometric","Visa Stamping","Emigration"]]},["name","given_name","passport_no","customer","status","remark"])
    for i in closure:
        data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (s_no,i.name,i.given_name,i.passport_no, i.customer, i.status,i.remark)
        ind+=1
        s_no+=1
    data += '</table>'
    subject = "Closure Status Report -  %s" % current_date
    message="""
                Dear Sir/Madam,<br>
                Kindly Find the below attached Closure Status Report  <br>{}<br>
                Thanks & Regards,<br>
                TEAM ERP<br>
                "This email has been automatically generated. Please do not reply"
            """.format(data)
    if ind>0:
        frappe.sendmail(
            # recipients=['divya.p@groupteampro.com'],
            recipients=['sangeetha.s@groupteampro.com','dc@groupteampro.com','keerthana.k@groupteampro.com','sangeetha.a@groupteampro.com','dineshbabu.k@groupteampro.com'],
            subject=subject,
            message=message
        )

@frappe.whitelist()
def mesg_for_permission(doc,method):
    frappe.msgprint("Kindly submit the document to confirm your permission")



@frappe.whitelist()
def update_sfp_status():
    sfp=frappe.db.get_all("Lead",{'company_name':['!=','']},['name','company_name'])
    for s in sfp:
        if not frappe.db.exists('Existing Leads',{'lead_name':s.company_name}):
            exist=frappe.new_doc("Existing Leads")
            exist.lead_id=s.name
            exist.lead_name=s.company_name
            exist.insert()
            exist.save(ignore_permissions=True)
            frappe.db.commit()

@frappe.whitelist()
def assign_task_to_users(tasks, users,rc):
    tasks = json.loads(tasks) 
    users = json.loads(users) 
    rc=json.loads(rc)
    for task in tasks:
        doc = frappe.get_doc("Task", task)
        doc.set("custom_assign_task", [])
        for user in users:
            new_user=user.get('assigned_to')
            doc.append("custom_assign_task", {"assigned_to": new_user})
        doc.custom_required_count=rc
        # doc.custom_production_date=p_date
        doc.save()
        frappe.db.commit()

@frappe.whitelist()
def update_appt():
    count=0
    sfp=frappe.db.get_all("Sales Follow Up",{'app_status':''},['name'])
    for s in sfp:
        frappe.db.set_value("Sales Follow Up",s.name,'app_status','Yet to visit(YTV)')
    return count

@frappe.whitelist()
def update_so_details():
    so_ietm_name=frappe.db.get_value("Sales Order Item",{'parent':"SAL-ORD-2023-00440"},['name'])
    print(so_ietm_name)

@frappe.whitelist()
def sales_ami_team_dsr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00007','user_id':('not in',['dm@groupteampro.com']),"name":('!=',("TC00058"))},['*'])
    emp_emails=[]
    date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date=nowdate()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    user_mails=[]
    for i in emp:
        emp_emails.append(i.user_id)
    for j in emp:
        user_mails.append(j.user_id)

    emp_emails.append('annie.m@groupteampro.com')
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="11"><b>ANI & Team DSR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    for c in emp_emails:
        
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":c})
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": c,"appointment_created_on":formatted_next_date})
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})

        short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
    appointment_list = frappe.db.get_all("Sales Follow Up",{"appointment_created_on":formatted_next_date},["*"])
    if appointment_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Fixed</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1">Exe</td><td colspan="2" style="text-align:center;">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
        for i in appointment_list:
            short_code = frappe.db.get_value("Employee", {"user_id": i.next_contact_by}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.custom_name12,i.app_status,i.custom_details1)
    data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
            
        </tr>
    '''
    for user in emp_emails:
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": c,"appointment_created_on":formatted_next_date})
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"]],"owner":user})
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":user, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":user},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else '0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

    todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": ["in", emp_emails], "status": ('not in',['Cancelled'])}, ["*"])
    if todo_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
        '''
        for todo in todo_list:
            short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
            data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
    data += '</table>'
    frappe.sendmail(
                recipients=['divya.p@groupteampro.com'],
                # recipients=['annie.m@groupteampro.com'], 
                # cc='dineshbabu.k@groupteampro.com',
                subject='ANI & Team DSR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    for d in user_mails:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="text-align:center;"><td colspan="11"><b>ANI & Team DSR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Apt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
                <td style="width:10%;"><b>OR%</b></b></td>
                <td style="width:10%;"><b>PR%</b></b></td>
            </tr>
        '''
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":d})
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": d,"appointment_created_on":formatted_next_date})

        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":d, "status": ('not in',['Cancelled'])})

        short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' ,effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
        # appointment_list = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":d},["*"])
        appointment_list = frappe.db.get_all("Sales Follow Up",{"appointment_created_on":formatted_next_date},["*"])
        if appointment_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="11";><b>Appointment Fixed</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1">Exe</td><td colspan="1" style="text-align:center;">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
            '''
        for i in appointment_list:
            # short_code = frappe.db.get_value("Employee", {"user_id": i.owner}, "short_code")
            # data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
            short_code = frappe.db.get_value("Employee", {"user_id": i.next_contact_by}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.custom_name12,i.app_status,i.custom_details1)

        data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Apt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
                <td style="width:10%;"><b>OR%</b></b></td>
                <td style="width:10%;"><b>PR%</b></b></td>
            </tr>
        '''
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"]],"owner":d})
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": d,"appointment_created_on":formatted_next_date})

        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":d, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": d, "status": ('not in',['Cancelled'])}, ["*"])
        if todo_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="11";"><b>ToDo</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
            '''
            for todo in todo_list:
                short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
                data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
        data += '</table>'

        frappe.sendmail(
                # recipients=[d],
                recipients=['divya.p@groupteampro.com'],
                subject='ANI & Team DSR  %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )


# @frappe.whitelist()
# def send_mail_to_sams():
#     from datetime import datetime
    
#     posting_date = datetime.now().strftime("%d-%m-%Y")
#     pro = frappe.get_doc("Project", "PROJ-1784")
#     tasks = frappe.get_all("Task", {'project': pro.name}, ['*'])
    
#     # Generate the HTML table
#     table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
#     table +='<tr><td colspan="7" style="text-align: center; font-weight: bold;">Territory: {}</td></tr>'.format(pro.territory or '')
#     table += '<tr style="background-color: #87CEFA"><td style="width: 10%; font-weight: bold; text-align: center;">Positions</td><td style="width: 30%; font-weight: bold; text-align: center;"># Vac </td><td style="width: 60%; font-weight: bold; text-align: center;">Salary</td><td style="width: 60%; font-weight: bold; text-align: center;">Food</td><td style="width: 60%; font-weight: bold; text-align: center;">Accommodation</td><td style="width: 60%; font-weight: bold; text-align: center;">ECR/ECNR</td><td style="width: 60%; font-weight: bold; text-align: center;">Major Key Skills</td></tr>'
    
#     for i in tasks:
#         vac = (i.vac * 3)
#         table += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
#             i.subject, vac, i.amount, i.food, i.accommodation, i.category, i.custom_major_key_skills
#         )
#     table += '</table>'
#     sams=frappe.get_all("SAMS",{"sa_status": ["!=", "Do Not Contact"]},["email_address","name"])
#     ind=0
#     for user in sams:
#         frappe.sendmail(
#             recipients=['divya.p@groupteampro.com'],
#             # recipients=[user.email_address],
#             subject=f'Position Details {posting_date} - Reg',
#             message=f"""
#             <b>Dear Team,</b><br><br>
#             Please find the below positions for {pro.name} for your kind reference and action.<br><br>
#             {table}<br><br>
#             Thanks & Regards,<br>TEAM ERP<br>
#             <i>This email has been automatically generated. Please do not reply</i>
#             """
#         )
@frappe.whitelist()
def send_mail_to_sams(name):
    from datetime import datetime
    
    posting_date = datetime.now().strftime("%d-%m-%Y")
    pro = frappe.get_doc("Project",name)
    tasks = frappe.get_all("Task", {'project': pro.name}, ['*'])
    
    # Generate the HTML table
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table +='<tr><td colspan="7" style="text-align: center; font-weight: bold;">Territory: {}</td></tr>'.format(pro.territory or '')
    table += '<tr style="background-color: #87CEFA"><td style="width: 5%; font-weight: bold; text-align: center;">Positions</td><td style="width: 4%; font-weight: bold; text-align: center;"># Vac </td><td style="width: 10%; font-weight: bold; text-align: center;">Salary</td><td style="width: 5%; font-weight: bold; text-align: center;">Food</td><td style="width: 5%; font-weight: bold; text-align: center;">Accommodation</td><td style="width: 5%; font-weight: bold; text-align: center;">ECR/ECNR</td><td style="width: 10%; font-weight: bold; text-align: center;">Major Key Skills</td></tr>'
    
    for i in tasks:
        vac = (i.vac * 3)
        table += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
            i.subject, vac, i.amount, i.food, i.accommodation, i.category, i.custom_major_key_skills
        )
    table += '</table>'
    sams=frappe.get_all("SAMS",{"sa_status": ["!=", "Do Not Contact"]},["email_address","name"])
    ind=0
    frappe.sendmail(
        recipients=['sangeetha.a@groupteampro.com'],
        # recipients=[user.email_address],
        subject=f'Position Details {posting_date} - Reg',
        message=f"""
        <b>Dear Sir/Mam,</b><br><br>
        Please find the below positions for {pro.name} for your kind reference and action.<br><br>
        {table}<br><br>
        Thanks & Regards,<br>TEAM ERP<br>
        <i>This email has been automatically generated. Please do not reply</i>
        """
    )

#  If the “Next Action Date” is Old date, either LUO or NAD is blank give a mail alert.
from datetime import datetime
import frappe
from frappe.utils import today, getdate

@frappe.whitelist()
def sendmail_luo_nad_alert():
    today_date = getdate()  # This ensures today_date is a date object
    formatted_date = today_date.strftime('%d/%m/%Y')
    
    count = 1
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
        <tr style="background-color: #0f1568 ;text-align:center;color: white;"><b>
            <td style='width:5%'><b>Sr</b></td>
            <td style='width:10%'><b>ID</b></td>
            <td style='width:15%'><b>Passport Number </b></td>
            <td style='width:20%'><b>Given Name/Surname</b></td>
            <td style='width:13%'><b>Status</b></td>
            <td style='width:25%'><b>Customer Name</b></td>
            <td style='width:7%'><b>Position</b></td>
            <td style='width:13%'><b>Next Action</b></td>
            <td style='width:20%'><b>Next Action On</b></td>
            <td style='width:20%'><b>Remarks</b></td>
            <td style='width:20%'><b>Last Updated On</b></td>

        </b></tr>
        '''
        
    closure = frappe.db.get_all("Closure", {"status": ("not in", ["Dropped", "Arrived"])}, ["*"])
    formatted_next_action=''
    formatted_last_action=''
    for i in closure:
        if i.custom_next_follow_up_on:
            formatted_next_action = i.custom_next_follow_up_on.strftime('%d/%m/%Y')
        if i.last_updated_on:
            formatted_last_action=i.last_updated_on.strftime('%d/%m/%Y')
        next_follow_up_date = getdate(i.custom_next_follow_up_on) if i.custom_next_follow_up_on else None
        if next_follow_up_date < today_date or i.last_updated_on is None or next_follow_up_date is None:
            data+='<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'%(count,i.name ,i.passport_no or '-',i.given_name,i.status,i.customer,i.task_subject,i.std_remarks,formatted_next_action or '-',i.remark,formatted_last_action or '-')
            count += 1

    data += '</table>'
    
    frappe.sendmail(
        recipients=['dineshbabu.k@groupteampro.com','sangeetha.s@groupteampro.com','sangeetha.a@groupteampro.com','keerthana.k@groupteampro.com','dc@groupteampro.com'],
        # recipients='divya.p@groupteampro.com',
        subject=f'Action Required: Closure with Outdated or Missing Follow-Up Details - {formatted_date}',
        message=f"""
        <b>Dear Team,</b><br><br>

        This is a reminder regarding closure records that have outdated or missing follow-up details as of {formatted_date}. 
        Please review the list below and take the necessary action.<br><br>

        {data}<br><br>
        Thanks & Regards,<br>
        TEAM ERP<br><br>

        <i>This email was automatically generated. Please do not reply.</i>
        """
    )

@frappe.whitelist()
def dnd_dsr():
    today_date=today()
    dsr_date=add_days(today_date,-1)
    date_obj = datetime.strptime(today_date, '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    count=1
    closure=frappe.db.get_all("Closure",{"last_updated_on":dsr_date},["*"])
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
        <tr style="background-color: #0f1568 ;text-align:center;color: white;"><b>
            <td style='width:5%'><b>Sr</b></td>
            <td style='width:10%'><b>ID</b></td>
            <td style='width:20%'><b>Given Name/Surname</b></td>
            <td style='width:15%'><b>Passport Number </b></td>
            <td style='width:25%'><b>Customer Name</b></td>
            <td style='width:13%'><b>Status</b></td>
            <td style='width:20%'><b>Remarks</b></td>
        </b></tr>
        '''
    for i in closure:
        data+='<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'%(count,i.name ,i.given_name,i.passport_no,i.customer,i.status,i.remark)
        count+=1
    data += '</table>'
    if count>1:
        frappe.sendmail(
                        recipients='dc@groupteampro.com',
                        cc=['dineshbabu.k@groupteampro.com','sangeetha.s@groupteampro.com','aruna.g@groupteampro.com','lokeshkumar.a@groupteampro.com','sangeetha.a@groupteampro.com','keerthana.k@groupteampro.com'],
                        # recipients='divya.p@groupteampro.com',
                        subject = f'DND- DSR {formatted_date} -Reg',
                        message = """
                        <b>Dear Team,</b><br><br>
        Please find the below DSR for {} for your kind reference and action.<br><br>

                    {}<br><br>
                        Thanks & Regards,<br>TEAM ERP<br>
                        
                        <i>This email has been automatically generated. Please do not reply</i>
                        """.format(formatted_date,data)
                    )


@frappe.whitelist()
def dnd_dpr():
    today_date=today()
    date_obj = datetime.strptime(today_date, '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    count=1
    closure=frappe.db.get_all("Closure",{"custom_next_follow_up_on":today_date},["*"])
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
        <tr style="background-color: #0f1568 ;text-align:center;color: white;"><b>
            <td style='width:5%'><b>Sr</b></td>
            <td style='width:10%'><b>ID</b></td>
            <td style='width:20%'><b>Given Name/Surname</b></td>
            <td style='width:15%'><b>Passport Number </b></td>
            <td style='width:25%'><b>Customer Name</b></td>
            <td style='width:13%'><b>Status</b></td>
            <td style='width:20%'><b>Remarks</b></td>
        </b></tr>
        '''
    for i in closure:
        data+='<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'%(count,i.name ,i.given_name,i.passport_no,i.customer,i.status,i.remark)
        count+=1
    data += '</table>'
    if count>1:
        frappe.sendmail(
                        recipients='dc@groupteampro.com',
                        cc=['dineshbabu.k@groupteampro.com','sangeetha.s@groupteampro.com','aruna.g@groupteampro.com','lokeshkumar.a@groupteampro.com','sangeetha.a@groupteampro.com','keerthana.k@groupteampro.com'],
                        # recipients='divya.p@groupteampro.com',
                        subject = f'DND- DPR {formatted_date} -Reg',
                        message = """
                        <b>Dear Team,</b><br><br>
        Please find the below DPR for {} for your kind reference and action.<br><br>

                    {}<br><br>
                        Thanks & Regards,<br>TEAM ERP<br>
                        
                        <i>This email has been automatically generated. Please do not reply</i>
                        """.format(formatted_date,data)
                    )

# @frappe.whitelist()
# def update_candidate_status():
#     frappe.db.set_value("Closure","CL03582","status","Onboarding")

# @frappe.whitelist()
# def create_item_closure():
#     item_candidate_id = frappe.db.get_value("Item", {"name": "7462862561"})
#     item_pp_id = frappe.db.get_value("Item", {"name": "C8825536"})
#     if item_candidate_id or item_pp_id:
#         item = frappe.get_doc("Item",item_pp_id)
#     else:
#         item = frappe.new_doc("Item")
#         item.item_code = "C8825536"
#         item.is_non_gst = "0"
#         item.item_name = "C8825536" + ":"+"MOHAMMED FARMAN SAFI"
#         candidate_owner="aruna.g@groupteampro.com"
#         if candidate_owner:
#             item.candidate_owner = candidate_owner
#         item.item_group = "Candidates"
#         item.stock_uom = "Nos"
#         item.qty = "1"
#         item.gst_hsn_code = '998519'
#         item.is_stock_item = "0"
#         item.include_item_in_manufacturing = "0"
#         item.description = "Arabian Castles for General Contracting Co., LLC"
#         item.append("item_defaults", {
#             "company": "TeamPRO HR & IT Services Pvt. Ltd."
#         })
#         item.append("customer_items", {
#             "customer_name": "Arabian Castles for General Contracting Co., LLC",
#             "ref_code": "8248117664"
#         })
#         item.insert()
#         item.save(ignore_permissions=True)



from datetime import datetime
@frappe.whitelist()
def dpr_mail(name,date,service,task_type):
    formatted_date = datetime.strptime(date, "%Y-%m-%d").strftime("%d-%m-%Y")
    daily = frappe.db.get_value(
        "Daily Monitor", 
        {"service": "IT-SW", "task_type": "CS"}, 
        ["*"]
    )
    
    task_data = frappe.get_doc("Daily Monitor",name)

    if task_type == "CS":
        # Table Header
        task_table = '''
        <table border="1" width="100%" style="border-collapse: collapse;">
            <tr style="background-color: #0f1568; text-align:center; color: white;">
                <td style='width:5%'><b>SI NO</b></td>
                <td style='width:10%'><b>ID</b></td>
                <td style='width:15%'><b>Project</b></td>
                <td style='width:20%'><b>Subject</b></td>
                <td style='width:5%'><b>CB</b></td>
                <td style='width:10%'><b>Status</b></td>
                <td style='width:5%'><b>Revision</b></td>
                <td style='width:5%'><b>AT</b></td>
                <td style='width:5%'><b>ET</b></td>
                <td style='width:5%'><b>RT</b></td>
                <td style='width:7%'><b>Priority</b></td>
                <td style='width:13%'><b>Allocated On</b></td>
            </tr>
        '''

        count = 1
        for task in task_data.task_details:
            task_table += f'''
            <tr>
                <td>{count}</td>
                <td>{task.id or task.issue}</td>
                <td>{task.project_name or '-'}</td>
                <td>{task.subject}</td>
                <td>{task.cb}</td>
                <td>{task.status}</td>
                <td>-</td>  
                <td>-</td>  
                <td>-</td>  
                <td>-</td>  
                <td>{task.priority}</td>
                <td>{task.allocated_on or ''}</td>
            </tr>
            '''
            count += 1

        task_table += '</table>'  # Closing the Task Table

        # -------------------- ISSUE LOOP --------------------
        issue_table = '''
        <table border="1" width="50%" style="border-collapse: collapse; text-align:center;">
            <tr style="background-color: #0f1568; color: white; text-align:center; font-size: 12px;">
                <td><b>SPOC</b></td>
                <td><b>APH</b></td>
                <td><b>PR[#/hr]</b></td>
                <td><b>Working [#/hr]</b></td>
                <td><b>CR[#/hr]</b></td>
                <td><b>Issue[#/hr]</b></td>
                <td><b>RT</b></td>
                <td><b>RT Vs APH %</b></td>
            </tr>
        '''

        task_det = frappe.db.get_all(
            "Task",
            {
                "custom_production_date_cs": date,
                "type": task_type,
                "service": service,
                "status": ("in", ["Working", "Pending Review", "Client Review"])
            },
            ['*'], 
            order_by='spoc asc', 
            group_by='spoc asc'
        )

        aph_total, pending_total,pr_sum,working_sum,client_sum, working_total, client_review, issue_overall, total,issue_value,working_value,pr_value,client_value,issue_total=0,0,0,0,0,0,0, 0,0, 0, 0, 0, 0, 0

        for k in task_det:
            emp_cb = frappe.db.get_value('Employee', {'user_id': k.spoc}, ['short_code'])
            actual_aph = frappe.db.get_value('Employee', {'short_code': emp_cb}, ['custom_aph'])

            pr_time = frappe.db.sql("""
                SELECT SUM(pr_expected_time) AS pr_time 
                FROM `tabTask`
                WHERE spoc=%s AND custom_production_date_cs=%s AND status='Pending Review'
            """, (k.spoc, date), as_dict=True)

            working_time = frappe.db.sql("""
                SELECT SUM(pr_expected_time) AS working_time 
                FROM `tabTask`
                WHERE spoc=%s AND custom_production_date_cs=%s AND status='Working'
            """, (k.spoc, date), as_dict=True)

            cr_time = frappe.db.sql("""
                SELECT SUM(pr_expected_time) AS cr_time 
                FROM `tabTask`
                WHERE spoc=%s AND custom_production_date_cs=%s AND status='Client Review'
            """, (k.spoc, date), as_dict=True)

            sum_rt = frappe.db.sql("""
                SELECT SUM(pr_expected_time) AS rt 
                FROM `tabTask`
                WHERE spoc=%s AND custom_production_date_cs=%s 
                AND type='CS' AND status!="Client Review"
                GROUP BY spoc
            """, (k.spoc, date), as_dict=True)

            pr_count = frappe.db.count("Task", {
                "custom_production_date_cs": date, 
                "type": "CS", 
                "service": "IT-SW", 
                "spoc": k.spoc, 
                "status": "Pending Review"
            })

            working_count = frappe.db.count("Task", {
                "custom_production_date_cs": date, 
                "type": "CS", 
                "service": "IT-SW", 
                "spoc": k.spoc, 
                "status": "Working"
            })

            client_count = frappe.db.count("Task", {
                "custom_production_date_cs": date, 
                "type": "CS", 
                "service": "IT-SW", 
                "spoc": k.spoc, 
                "status": "Client Review"
            })

            issue_data = frappe.db.sql("""
                SELECT SUM(custom_excepted_time_cs) AS issue_1 
                FROM `tabIssue`
                WHERE custom_spoc=%s AND custom_production_date=%s
            """, (k.spoc, date), as_dict=True)
            issue_count = frappe.db.count("Issue", {"custom_production_date": date, "custom_spoc": k.spoc})

            issue_value = f"{issue_data[0].issue_1:.1f}" if issue_data and issue_data[0].issue_1 else "0"
            pr_value= f"{pr_time[0].pr_time:.1f}" if pr_time and pr_time[0].pr_time else "0"
            working_value= f"{working_time[0].working_time:.1f}" if working_time and working_time[0].working_time else "0"
            client_value= f"{cr_time[0].cr_time:.1f}" if cr_time and cr_time[0].cr_time else "0"
            pending_total += pr_count
            working_total += working_count
            client_review += client_count
            issue_overall+=issue_count
            issue_total+=float(issue_value)
            pr_sum+=float(pr_value)
            working_sum+=float(working_value)
            client_sum+=float(client_value)
            

            if sum_rt:
                total += sum_rt[0].rt

            aph_total += float(actual_aph or 0)
            percent = (float(sum_rt[0].rt) / float(actual_aph)) * 100 if sum_rt and actual_aph else 0

            issue_table += f'''
            <tr style="font-size: 14px;">
                <td>{emp_cb}</td>
                <td>{actual_aph or '8'}</td>
                <td>{pr_count or '0'} / {pr_value}</td>
                <td>{working_count or '0'} /{working_value}</td>
                <td>{client_count or '0'} / {client_value}</td>
                <td>{issue_count or '0'} / {issue_value}</td>
                <td>{sum_rt[0].rt if sum_rt else '0'}</td>
                <td>{round(percent, 2)}</td>
            </tr>
            '''

        issue_table += f'''
        <tr style="font-size: 14px;">
            <td colspan=1><b>Total</b></td>
            <td>{aph_total}</td>
            <td>{pending_total} / {pr_sum if pr_sum else '0'}</td>
            <td>{working_total} / {working_sum if working_sum else '0'}</td>
            <td>{client_review} / {client_sum if client_sum else '0'}</td>
            <td>{issue_overall} / {issue_total if issue_total else '0'}</td>
            <td>{round(total, 2) if total else '0'}</td>
            <td>{round((total/aph_total)*100, 2) if aph_total else '0'}</td>
        </tr>
        </table>
        '''  # Closing the Issue Table

        # Sending Email
        frappe.sendmail(
            sender='sarath.v@groupteampro.com',
            recipients="siva.m@groupteampro.com",
            subject = f'{"IT-SW"} - {"CS"} DPR {formatted_date} -Reg',
            message="""
                <b>Dear Team,</b><br><br>
                Please find the below DPR for {} for your kind reference and action. Ensure all the tasks are allocated on time and as per the requirement.<br><br>
                {}<br><br>
                {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                <i>This email has been automatically generated. Please do not reply</i>
            """ .format(formatted_date,issue_table if issue_table else '',task_table if task_table else '')
        )

        frappe.msgprint("DPR mail has been successfully sent")
        task_data.dm_status = 'DPR Completed'
        task_data.dpr_submitted_on = today()
        task_data.save()
        frappe.db.commit()


from datetime import datetime
@frappe.whitelist()
def dsr_mail(name,service,task_type,date):
    formatted_date = datetime.strptime(date, "%Y-%m-%d").strftime("%d-%m-%Y")
    task_data=frappe.get_doc("Daily Monitor",name)
    if task_data.dsr_check==1:
        if task_type =="CS":
            sum_et=0
            table=''
            count=1
            aph_totals=0
            total_at=0
            data = '<table border="1" width="100%" style="border-collapse: collapse;">'
            data += '''
            <tr style="background-color: #0f1568 ;color: white;text-align:center;font-size: 12px;">
                <td style='width:4%'><b>SI NO</b></td>
                <td style='width:6%'><b>Task/Issue ID</b></td>
                <td style='width:12%'><b>Project </b></td>
                <td style='width:18%'><b>Subject</b></td>
                <td style='width:4%'><b>CB</b></td>
                <td style='width:7%'><b>Status</b></td>
                <td style='width:4%'><b>Revision</b></td>
                <td style='width:6%'><b>Priority</b></td>
                <td style='width:8%'><b>Allocated On</b></td>
                <td style='width:4%'><b>Time Taken</b></td>
                <td style='width:10%'><b>Remarks</b></td>
                <td style='width:9%'><b>TL Remarks</b></td>
            </tr>
            '''
            table = '<table border="1" width="70%" style="border-collapse: collapse;text-align:center;">'
            table += '''
            <tr style="background-color: #0f1568 ;color: white;text-align:center;font-size: 12px;">
                <td style='width:1%'><b>SPOC</b></td>
                <td style='width:1%'><b>APH</b></td>
                <td style='width:1%'><b>RT</b></td>
                <td style='width:1%'><b>Actual Time Taken</b></td>
                <td style='width:1%'><b>RT Vs APH %</b></td>
                <td style='width:1%'><b>PR[#/hr]</b></td>
                <td style='width:1%'><b>Working [#/hr]</b></td>
                <td style='width:1%'><b>CR[#/hr]</b></td>
                <td style='width:1%'><b>Issue[#/hr]</b></td>
                <td style='width:1%'><b>OR %</b></td>
                <td style='width:1%'><b>PR %</b></td>
            </tr>
            '''
            cs_task = frappe.db.get_all("Task", {"custom_production_date":date,"type":task_type,"service":service,"status": ["not in", ["Working", "Open"]]}, ['*'], order_by='spoc asc',group_by='spoc asc')
            task = frappe.db.get_all("Task", {"custom_production_date":date,"type":task_type,"service":service}, ['*'], order_by='cb asc',group_by='custom_allocated_to asc')
            
            sorted_task_details = sorted(task_data.task_details, key=lambda i: i.cb or '')
            count = 1
            for i in sorted_task_details:
                vtaken = float(i.at)
                value_taken = round(vtaken, 3)
                if i.at_taken:
                    t_taken = float(i.at_taken)
                    today_taken = round(t_taken, 3)
                else:
                    t_taken='0'
                    today_taken='0'
                remark = '-' if i.remark is None else i.remark
                tl_remark = '-' if i.tl_remark is None else i.tl_remark
                if i.id is not None:
                    id=i.id
                elif i.issue is not None:
                    id=i.issue
                elif i.meeting is not None:
                    id=i.meeting
                else:
                    id='-'
                data += '<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'% (count,id, i.project_name, i.subject, i.cb, i.status, i.revisions,i.priority, i.allocated_on, today_taken, remark, tl_remark)
                count += 1
            data += '</table>'
            pending_count_total=0
            working_total=0
            client_total=0
            issue_overall, total,issue_value,working_value,pr_value,client_value=0,0,0, 0,0, 0
            for j in cs_task:
                employee_id=frappe.db.get_value('Employee',{'user_id':j.spoc},['name'])
                emp_cb=frappe.db.get_value('Employee',{'user_id':j.spoc},['short_code'])
                timesheet = frappe.db.get_value("Timesheet", {'start_date': date, 'employee':employee_id}, ['total_hours'])  
                actual_aph=frappe.db.get_value('Employee',{'short_code':emp_cb},['custom_aph'])
                pending_review_count=frappe.db.count("Task",{"custom_production_date":date,"type":task_type,"service":service,"spoc":j.spoc,"status":"Pending Review"})
                working_count=frappe.db.count("Task",{"custom_production_date":date,"type":task_type,"service":service,"spoc":j.spoc,"status":"Working"})
                client_count=frappe.db.count("Task",{"custom_production_date":date,"type":task_type,"service":service,"spoc":j.spoc,"status":"Client Review"})
                pr_time = frappe.db.sql("""
                    SELECT SUM(pr_expected_time) AS pr_time 
                    FROM `tabTask`
                    WHERE spoc=%s AND custom_production_date_cs=%s AND status='Pending Review'
                """, (j.spoc, date), as_dict=True)

                working_time = frappe.db.sql("""
                    SELECT SUM(pr_expected_time) AS working_time 
                    FROM `tabTask`
                    WHERE spoc=%s AND custom_production_date_cs=%s AND status='Working'
                """, (j.spoc, date), as_dict=True)

                cr_time = frappe.db.sql("""
                    SELECT SUM(pr_expected_time) AS cr_time 
                    FROM `tabTask`
                    WHERE spoc=%s AND custom_production_date_cs=%s AND status='Client Review'
                """, (j.spoc, date), as_dict=True)

                issue_data = frappe.db.sql("""
                    SELECT SUM(custom_excepted_time_cs) AS issue_time 
                    FROM `tabIssue`
                    WHERE custom_spoc=%s AND custom_production_date=%s
                """, (j.spoc, date), as_dict=True)

                issue_count = frappe.db.count("Issue", {"custom_production_date": date, "custom_spoc": j.spoc})

                issue_value = f"{issue_data[0].issue_1:.1f}" if issue_data and issue_data[0].issue_1 else "0"
                pr_value= f"{pr_time[0].pr_time:.1f}" if pr_time and pr_time[0].pr_time else "0"
                working_value= f"{working_time[0].working_time:.1f}" if working_time and working_time[0].working_time else "0"
                client_value= f"{cr_time[0].cr_time:.1f}" if cr_time and cr_time[0].cr_time else "0"

                pending_count_total+=float(pending_review_count)
                working_total+=float(working_count)
                client_total +=float(client_count)
                issue_overall+=issue_count
                for datas in task_data.dm_summary:
                    if datas.d_cb == emp_cb:
                        sum_et=datas.d_rt
                if sum_et:
                    total+=float(sum_et)

                if actual_aph:
                    aph_totals+=float(actual_aph)
                if timesheet is not None:
                    total_at+=float(timesheet)
                if actual_aph and timesheet:
                    percent=(float(sum_et)/float(actual_aph))*100
                    value=actual_aph
                    or_count=float(timesheet)/float(value)*100
                if timesheet and sum_et is not None:
                    pr_count=float(sum_et)/float(timesheet)*100

                if total_at:
                    or_total=float(total_at)/float(aph_totals)*100
                    pr_total=float(total)/float(total_at)*100
                if aph_totals!=0:
                    total_count=float(total)/float(aph_totals)*100
                if percent:
                    table+='<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>' %(emp_cb,actual_aph or '8',sum_et if sum_et else '0',round(timesheet,2) if timesheet is not None else '0',round(percent) if timesheet is not None else '0',f"{pending_review_count or '0'}/{pr_value or '0'}",f"{working_count or '0'}/{working_value or '0'}",f"{client_count or '0'}/{client_value or '0'}",f"{issue_count or '0'}/{issue_overall or '0'}",round(or_count) if timesheet is not None else '0',round(pr_count) if timesheet is not None else '0')
                else:
                    table+='<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>' %(emp_cb,actual_aph or '8',sum_et if sum_et else '0',round(timesheet,2) if timesheet is not None else'0','0',f"{pending_review_count or '0'}/{pr_value or '0'}",f"{working_count or '0'}/{working_value or '0'}",f"{client_count or '0'}/{client_value or '0'}",f"{issue_count or '0'}/{issue_overall or '0'}",round(or_count,2) or '0',round(pr_count,2) or '0')
            table+='<tr style="font-size: 14px;" ><td colspan=1>Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'%(aph_totals,total,round(total_at,2),round(total_count),pending_count_total,working_total,client_total,issue_overall,round(or_total),round(pr_total))
            table += "</table>"
            frappe.sendmail(
                    sender='sarath.v@groupteampro.com',
                    recipients="siva.m@groupteampro.com",
                    # recipients=['sarath.v@groupteampro.com','sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'],
                    # cc=['dineshbabu.k@groupteampro.com','anil.p@groupteampro.com','abdulla.pi@groupteampro.com'],
                    subject = f'{service} - {task_type} DSR {formatted_date} -Reg',
                    message = """
                <b>Dear Team,</b><br><br>
                    Please find the below DSR for {} for your kind reference.<br><br>
                    {}<br><br>
                    {}<br><br>
                    Thanks & Regards,<br>TEAM ERP<br>
                    <i>This email has been automatically generated. Please do not reply</i>
                    """.format(formatted_date,table if table else '',data)
                )
            frappe.msgprint("DSR mail has been successfully sent.")
            # task_data.dm_status='Submitted'
            task_data.dsr_submitted_on=today()
            task_data.save()
            frappe.db.commit()

import frappe
from frappe.utils import getdate, nowdate, date_diff

@frappe.whitelist()
def check_closure_tat_and_alert():
    settings = frappe.get_doc("Closure Settings")
    tat_map = {row.status: row.tat_days for row in settings.closure_tat_days}

    alert_rows = []
    index = 1

    closures = frappe.get_all("Closure", filters={"status":['not in', ['Dropped','Arrived']]}, fields=["name","given_name","passport_no"])

    for closure in closures:
        closure_doc = frappe.get_doc("Closure", closure.name)
        if closure_doc.custom_history:
            sorted_history = sorted(closure_doc.custom_history, key=lambda x: x.date, reverse=True)
            last_entry = sorted_history[0]
            last_status = last_entry.status
            last_date = getdate(last_entry.date)

            if last_status in tat_map:
                days_allowed = tat_map[last_status]
                current_days = date_diff(nowdate(), last_date)

                if current_days > days_allowed:
                    days_crossed = current_days - days_allowed
                    alert_rows.append(f"""
                        <tr>
                            <td>{index}</td>
                            <td>{closure_doc.name}</td>
                            <td>{closure_doc.given_name}</td>
                            <td>{closure_doc.passport_no}</td>
                            <td>{last_status}</td>
                            <td style="text-align: right;">{days_allowed}</td>
                            <td style="text-align: right;">{days_crossed}</td>
                        </tr>
                    """)
                    index += 1

    if alert_rows:
        table_html = f"""
            <p>Dear Sir/Madam,</p>
            <p>Kindly find below the Closure TAT Alert details:</p>
            <table border="1" cellpadding="5" cellspacing="0">
                <tr style="background-color: #0f1568; color: white; text-align: center;">
                    <th>S.No</th>
                    <th>ID</th>
                    <th>Name</th>
                    <th>Passport Number</th>
                    <th>Status</th>
                    <th >TAT Days</th>
                    <th >Days Crossed</th>
                </tr>
                {''.join(alert_rows)}
            </table>
        """

        frappe.sendmail(
            recipients=["divya.p@groupteampro.com","sangeetha.a@groupteampro.com","keerthana.k@groupteampro.com"],
            cc=['sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
            subject="Closure TAT Alert",
            message=table_html
        )

import frappe
from frappe import _
from frappe.utils import today
from datetime import datetime

@frappe.whitelist()
def send_proj_creation(name, account_manager=None, customer=None):
    custom_date = today()
    date_obj = datetime.strptime(str(custom_date), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')

    subject = f'New Project Created on {formatted_date} - Reg'

    message = f"""
    <b>Dear Team,</b><br><br>

    A new project has been created on kick of completion. Please find the project details below:<br><br>

    <table border="1" cellspacing="0" cellpadding="5">
        <tr><td><b>Project Name</b></td><td>{name}</td></tr>
        <tr><td><b>Customer</b></td><td>{customer or ''}</td></tr>
    """

    if account_manager:
        message += f"<tr><td><b>Account Manager</b></td><td>{account_manager}</td></tr>"

    message += """
    </table>
    <br><br>
    Kindly take the necessary actions.<br><br>

    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """

    recipients = ['sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com','sangeetha.a@groupteampro.com']
    if account_manager:
        recipients.append(account_manager)

    frappe.sendmail(
        recipients=recipients,
        subject=subject,
        message=message
    )

@frappe.whitelist()
def restrict_leave(doc,method):
    current_date = today()
    allowed=add_days(doc.from_date,1)
    while check_holiday(allowed, doc.employee):
        allowed = add_days(allowed, 1)
    if current_date >  allowed:
        frappe.throw("Leave applications should be raised within next 24 hours")
@frappe.whitelist()
def restrict_att_req(doc,method):
    current_date = today()
    allowed=add_days(doc.from_date,1)
    while check_holiday(allowed, doc.employee):
        allowed = add_days(allowed, 1)
    if current_date >  allowed:
        frappe.throw("Attendance Request should be raised within next 24 hours")
@frappe.whitelist()
def check_holiday(date, emp):
    holiday_list = frappe.db.get_value('Employee', emp, 'holiday_list')
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
    if holiday:
        if holiday[0].weekly_off == 1:
            return True
        else:
            return True
    else:
        return False  